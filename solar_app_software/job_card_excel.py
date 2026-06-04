"""
Job Card Excel generator — Power On Plus Solar Solutions
Reproduces the physical JOB CARD DETAILS form as a printable A4 Excel sheet.
"""
from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, PatternFill
)
from openpyxl.utils import get_column_letter
from openpyxl.styles.borders import BORDER_THIN, BORDER_MEDIUM
from datetime import date
import os

def thin(color="000000"):
    s = Side(style=BORDER_THIN, color=color)
    return s

def med():
    return Side(style=BORDER_MEDIUM, color="000000")

def all_thin(cell):
    s = thin()
    cell.border = Border(left=s, right=s, top=s, bottom=s)

def all_med(cell):
    m = med()
    cell.border = Border(left=m, right=m, top=m, bottom=m)

def label_style(cell, text, bold=False, size=8, align="left", wrap=False):
    cell.value = text
    cell.font = Font(name="Arial", size=size, bold=bold)
    cell.alignment = Alignment(
        horizontal=align, vertical="center", wrap_text=wrap
    )

def value_cell(cell, text="", size=8):
    cell.value = text
    cell.font = Font(name="Arial", size=size)
    cell.alignment = Alignment(horizontal="left", vertical="center")


def build_job_card(project=None, output_path="/tmp/job_card.xlsx"):
    """
    Build the job card Excel sheet.
    `project` is the SQLAlchemy Project object (or None for blank template).
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Job Card"

    # ── Page setup ──────────────────────────────────────────────────────────
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.page_setup.orientation = "portrait"
    ws.page_setup.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 1
    ws.page_margins.left   = 0.4
    ws.page_margins.right  = 0.4
    ws.page_margins.top    = 0.4
    ws.page_margins.bottom = 0.4
    ws.sheet_view.showGridLines = False

    # ── Column widths (A–N) ────────────────────────────────────────────────
    col_widths = {
        "A": 3, "B": 10, "C": 12, "D": 10,
        "E": 3,  "F": 8,  "G": 8,  "H": 6,
        "I": 4,  "J": 4,  "K": 4,  "L": 4,
        "M": 4,  "N": 4,
    }
    for col, w in col_widths.items():
        ws.column_dimensions[col].width = w

    def rh(row, height):
        ws.row_dimensions[row].height = height

    # ── Helpers ─────────────────────────────────────────────────────────────
    BG_HEADER = PatternFill("solid", fgColor="000000")   # black header
    BG_LIGHT  = PatternFill("solid", fgColor="F2F2F2")   # light gray section
    BG_WHITE  = PatternFill("solid", fgColor="FFFFFF")

    def merge(start, end, value="", bold=False, size=8,
              bg=None, color="000000", align="left", wrap=False, border=True):
        ws.merge_cells(f"{start}:{end}")
        cell = ws[start]
        cell.value = value
        cell.font = Font(name="Arial", size=size, bold=bold, color=color)
        cell.alignment = Alignment(
            horizontal=align, vertical="center", wrap_text=wrap
        )
        if bg:
            cell.fill = bg
        if border:
            # apply thin border to every cell in the merged range
            # openpyxl only lets us style individual cells for borders
            from openpyxl.utils import range_boundaries, get_column_letter
            c1, r1, c2, r2 = range_boundaries(f"{start}:{end}")
            for rr in range(r1, r2 + 1):
                for cc in range(c1, c2 + 1):
                    c = ws.cell(rr, cc)
                    s = thin()
                    c.border = Border(left=s, right=s, top=s, bottom=s)
        return cell

    def label(col, row, text, bold=False, size=8, bg=None, color="000000",
              align="left", wrap=False):
        cell = ws.cell(row=row, column=col)
        cell.value = text
        cell.font = Font(name="Arial", size=size, bold=bold, color=color)
        cell.alignment = Alignment(
            horizontal=align, vertical="center", wrap_text=wrap
        )
        if bg:
            cell.fill = bg
        s = thin()
        cell.border = Border(left=s, right=s, top=s, bottom=s)
        return cell

    def val(col, row, text="", size=8, bold=False, color="000000", align="left"):
        cell = ws.cell(row=row, column=col)
        cell.value = text
        cell.font = Font(name="Arial", size=size, bold=bold, color=color)
        cell.alignment = Alignment(horizontal=align, vertical="center")
        s = thin()
        cell.border = Border(left=s, right=s, top=s, bottom=s)
        return cell

    # ── Pull data from project object ────────────────────────────────────────
    p = project
    c = p.customer if p else None
    cd = p.connection_details if p else None
    ld = p.loan_detail if p else None
    pd_ = p.panel_details if p else None
    kseb = p.kseb_task if p else None
    sub = p.subsidy if p else None

    def safe(val, fmt=None):
        if val is None:
            return ""
        if fmt == "inr":
            try:
                return f"₹{float(val):,.0f}"
            except:
                return str(val)
        return str(val)

    serial_no = safe(p.project_code if p else "")
    today_str = date.today().strftime("%d-%m-%Y")

    # ──────────────────────────────────────────────────────────────────────────
    # ROW 1 — Header: POWER ON + | JOB CARD DETAILS
    # ──────────────────────────────────────────────────────────────────────────
    rh(1, 22)
    merge("A1", "H1", "POWER ON +", bold=True, size=14,
          bg=BG_WHITE, align="center")
    merge("I1", "N1", "JOB CARD DETAILS", bold=True, size=11,
          bg=BG_WHITE, align="center")

    # ──────────────────────────────────────────────────────────────────────────
    # ROW 2 — REG / Serial No + Date
    # ──────────────────────────────────────────────────────────────────────────
    rh(2, 14)
    merge("A2", "H2", "REG:", bold=False, size=8)
    label(9, 2, "Serial No:", bold=True, size=8, align="right")
    merge("J2", "N2", serial_no, bold=True, size=9, align="center")

    # ──────────────────────────────────────────────────────────────────────────
    # ROW 3 — MOB / Date
    # ──────────────────────────────────────────────────────────────────────────
    rh(3, 14)
    merge("A3", "H3", f"MOB:", size=8)
    label(9, 3, "Date:", bold=True, size=8, align="right")
    merge("J3", "N3", today_str, size=8, align="center")

    # ROW 4 — Handling / KW, Panel
    rh(4, 14)
    handling_name = safe(p.doc_staff.full_name if p and p.doc_staff else "")
    merge("A4", "D4", f"Handling: {handling_name}", size=8)
    merge("E4", "H4", "", size=8)
    label(9, 4, "KW:", bold=True, size=8)
    merge("J4", "K4", safe(p.inverter_capacity_kw if p else ""), size=8)
    label(12, 4, "Panel:", bold=True, size=8)
    merge("M4", "N4", safe(p.panel_capacity_kw if p else ""), size=8)

    # ROW 5 — C/O / Roof, Phase checkboxes
    rh(5, 14)
    coord_name = safe(p.coordinator.full_name if p and p.coordinator else "")
    merge("A5", "D5", "C/O:", size=8, bold=True)
    merge("E5", "H5", coord_name, size=8)

    # ROW 6 — Sub C/O / Phase
    rh(6, 14)
    merge("A6", "D6", "Sub C/O:", size=8, bold=True)
    merge("E6", "H6", safe(c.sub_co if c else ""), size=8)
    label(9, 6, "Phase:", bold=True, size=8)
    phase_val = ""
    if cd and cd.connection_type:
        phase_val = "3Ph" if "Three" in cd.connection_type else "1Ph"
    merge("J6", "N6", phase_val, size=8)

    # ──────────────────────────────────────────────────────────────────────────
    # ROW 7 — Name / Section
    # ──────────────────────────────────────────────────────────────────────────
    rh(7, 14)
    merge("A7", "D7", "Name:", size=8, bold=True)
    merge("E7", "H7", safe(c.name if c else ""), size=8, bold=True)
    label(9, 7, "Section:", bold=True, size=8)
    merge("J7", "N7", safe(cd.kseb_section if cd else ""), size=8)

    # ──────────────────────────────────────────────────────────────────────────
    # ROW 8 — Nickname / Category
    # ──────────────────────────────────────────────────────────────────────────
    rh(8, 14)
    merge("A8", "D8", "Nickname:", size=8, bold=True)
    merge("E8", "H8", "", size=8)
    label(9, 8, "Category:", bold=True, size=8)
    cat = safe(cd.category if cd else "")
    merge("J8", "N8", cat, size=8)

    # ──────────────────────────────────────────────────────────────────────────
    # ROW 9 — House Name / Loan
    # ──────────────────────────────────────────────────────────────────────────
    rh(9, 14)
    merge("A9", "D9", "House Name:", size=8, bold=True)
    merge("E9", "H9", safe(c.house_name if c else ""), size=8)
    label(9, 9, "Loan:", bold=True, size=8)
    loan_val = ""
    if p and p.project_type == "Loan":
        loan_val = safe(ld.bank_name if ld else "Yes")
    merge("J9", "N9", loan_val, size=8)

    # ──────────────────────────────────────────────────────────────────────────
    # ROW 10 — OTP No / Ownership
    # ──────────────────────────────────────────────────────────────────────────
    rh(10, 14)
    merge("A10", "D10", "OTP No:", size=8, bold=True)
    merge("E10", "H10", safe(c.phone if c else ""), size=8)
    label(9, 10, "Ownership:", bold=True, size=8)
    own_val = ""
    if cd:
        own_val = "Required" if cd.ownership_change_needed else "Not Required"
    merge("J10", "N10", own_val, size=8)

    # ──────────────────────────────────────────────────────────────────────────
    # ROW 11 — APP ID / Load Clearance
    # ──────────────────────────────────────────────────────────────────────────
    rh(11, 14)
    merge("A11", "D11", "APP ID:", size=8, bold=True)
    merge("E11", "H11", "", size=8)
    label(9, 11, "Load Clearance:", bold=True, size=8, wrap=True)
    lc_val = ""
    if cd:
        lc_val = "Required" if cd.load_clearance_needed else "Not Required"
    merge("J11", "N11", lc_val, size=8)

    # ──────────────────────────────────────────────────────────────────────────
    # ROW 12 — Consumer Number / Feasibility
    # ──────────────────────────────────────────────────────────────────────────
    rh(12, 14)
    merge("A12", "D12", "Consumer Number:", size=8, bold=True, wrap=True)
    merge("E12", "H12", safe(cd.consumer_number if cd else ""), size=8)
    label(9, 12, "Feasibility:", bold=True, size=8)
    doc_map = {}
    if p:
        doc_map = {d.doc_type: d for d in p.documents}
    feas = doc_map.get("Feasibility Receipt")
    feas_val = safe(feas.status if feas else "")
    merge("J12", "N12", feas_val, size=8)

    # build doc_map from documents
    doc_map = {}
    if p:
        doc_map = {d.doc_type: d for d in p.documents}

    def doc_status(key):
        d = doc_map.get(key)
        return d.status if d else ""

    # ROW 13 — Stamp Paper
    merge(f"J13", f"N13", doc_status('KSEB Stamp Paper'), size=8)

    # ROW 14 — B-Class Licence  
    merge(f"J14", f"N14", doc_status('B-Class Licence'), size=8)

    # ──────────────────────────────────────────────────────────────────────────
    # ROW 15 — Pin / Photos
    # ──────────────────────────────────────────────────────────────────────────
    rh(15, 14)
    merge("A15", "D15", "Pin:", size=8, bold=True)
    merge("E15", "H15", safe(c.pincode if c else ""), size=8)
    label(9, 15, "Photos:", bold=True, size=8)
    merge("J15", "N15", "", size=8)

    # ──────────────────────────────────────────────────────────────────────────
    # ROW 16 — Village / MNRE
    # ──────────────────────────────────────────────────────────────────────────
    rh(16, 14)
    merge("A16", "D16", "Village:", size=8, bold=True)
    merge("E16", "H16", safe(c.village if c else ""), size=8)
    label(9, 16, "MNRE:", bold=True, size=8)
    mnre_doc = doc_map.get("MNRE")
    merge("J16", "N16", safe(mnre_doc.status if mnre_doc else ""), size=8)

    # ──────────────────────────────────────────────────────────────────────────
    # ROW 17 — Taluk / KSEB Connection
    # ──────────────────────────────────────────────────────────────────────────
    rh(17, 14)
    merge("A17", "D17", "Taluk:", size=8, bold=True)
    merge("E17", "H17", safe(c.taluk if c else ""), size=8)
    label(9, 17, "KSEB Connection:", bold=True, size=8, wrap=True)
    kseb_conn = doc_map.get("KSEB Connection")
    merge("J17", "N17", safe(kseb_conn.status if kseb_conn else ""), size=8)

    # ROW 18 — KSEB Completion — now reads 'KSEB Connection Done'
    kseb_done = "✓" if doc_map.get('KSEB Connection Done') and doc_map['KSEB Connection Done'].status in ['Received','Completed'] else ""
    merge("J18", "N18", kseb_done, size=8)

    # ──────────────────────────────────────────────────────────────────────────
    # ROW 19 — Email ID / Net Meter
    # ──────────────────────────────────────────────────────────────────────────
    rh(19, 14)
    email = safe(c.email if c else "")
    merge("A19", "D19", "Email ID:", size=8, bold=True)
    merge("E19", "H19", email, size=8)
    label(9, 19, "Net Meter:", bold=True, size=8)
    nm_serial = safe(pd_.net_meter_serial_number if pd_ else "")
    merge("J19", "N19", nm_serial, size=8)

    # ──────────────────────────────────────────────────────────────────────────
    # ROW 20 — Password / App Installation
    # ──────────────────────────────────────────────────────────────────────────
    rh(20, 14)
    merge("A20", "D20", "Password:", size=8, bold=True)
    merge("E20", "H20", "", size=8)
    label(9, 20, "App Installation:", bold=True, size=8, wrap=True)
    app_doc = doc_map.get("App Installation")
    app_val = safe(app_doc.status if app_doc else "")
    merge("J20", "N20", app_val, size=8)

    # ──────────────────────────────────────────────────────────────────────────
    # ROW 21 — Subsidy Cheque / blank
    # ──────────────────────────────────────────────────────────────────────────
    rh(21, 14)
    merge("A21", "D21", "Subsidy Cheque:", size=8, bold=True)
    sub_status = safe(sub.status if sub else "")
    merge("E21", "H21", sub_status, size=8)
    label(9, 21, "Transportation-2:", bold=True, size=8, wrap=True)
    merge("J21", "N21", "", size=8)

    # ──────────────────────────────────────────────────────────────────────────
    # ROW 22 — Bank Details / blank
    # ──────────────────────────────────────────────────────────────────────────
    rh(22, 14)
    merge("A22", "D22", "Bank Details:", size=8, bold=True)
    bank_val = safe(ld.bank_name if ld else "")
    merge("E22", "H22", bank_val, size=8)
    merge("I22", "N22", "", size=8)

    # ──────────────────────────────────────────────────────────────────────────
    # ROW 23 — Transportation-1 / Structure
    # ──────────────────────────────────────────────────────────────────────────
    rh(23, 14)
    merge("A23", "D23", "Transportation-1:", size=8, bold=True)
    merge("E23", "H23", "", size=8)
    label(9, 23, "Structure:", bold=True, size=8)
    op = p.onsite_progress if p else None
    struct_val = safe(op.structure_work_status if op else "")
    merge("J23", "N23", struct_val, size=8)

    # ──────────────────────────────────────────────────────────────────────────
    # ROW 24 — Structure (left) / Electrical
    # ──────────────────────────────────────────────────────────────────────────
    rh(24, 14)
    merge("A24", "D24", "Structure:", size=8, bold=True)
    merge("E24", "H24", struct_val, size=8)
    label(9, 24, "Electrical:", bold=True, size=8)
    elec_val = safe(op.electrical_status if op else "")
    merge("J24", "N24", elec_val, size=8)

    # ──────────────────────────────────────────────────────────────────────────
    # ROW 25 — Wheeling / Wheeling
    # ──────────────────────────────────────────────────────────────────────────
    rh(25, 14)
    merge("A25", "D25", "Wheeling:", size=8, bold=True)
    merge("E25", "H25", "", size=8)
    label(9, 25, "Wheeling:", bold=True, size=8)
    merge("J25", "N25", "", size=8)

    # ──────────────────────────────────────────────────────────────────────────
    # ROW 26 — Credit Duration (merged)
    # ──────────────────────────────────────────────────────────────────────────
    rh(26, 14)
    merge("A26", "N26", "Credit Duration:", size=8, bold=True)

    # ──────────────────────────────────────────────────────────────────────────
    # ROW 27 — System Details header
    # ──────────────────────────────────────────────────────────────────────────
    rh(27, 16)
    merge("A27", "N27", "System Details",
          bold=True, size=10, bg=BG_LIGHT, align="left")

    # ──────────────────────────────────────────────────────────────────────────
    # ROWS 28-30 — Serial numbers
    # ──────────────────────────────────────────────────────────────────────────
    for row, lbl, data in [
        (28, "NET METER SERIAL NUMBER:", safe(pd_.net_meter_serial_number if pd_ else "")),
        (29, "ENERGY METER SERIAL NUMBER:", safe(pd_.energy_meter_serial_number if pd_ else "")),
        (30, "SYSTEM SERIAL:", safe(pd_.inverter_serial_number if pd_ else "")),
        (31, "PANEL DETAILS:", safe(pd_.panel_brand if pd_ else "")),
    ]:
        rh(row, 14)
        merge(f"A{row}", f"D{row}", lbl, size=8, bold=True)
        merge(f"E{row}", f"N{row}", data, size=8)

    # ROW 32-34 — extra blank rows for panel serials
    for r in range(32, 35):
        rh(r, 14)
        merge(f"A{r}", f"N{r}", "", size=8)

    # ──────────────────────────────────────────────────────────────────────────
    # ROW 35 — Payment Details header
    # ──────────────────────────────────────────────────────────────────────────
    rh(35, 16)
    merge("A35", "N35", "Payment Details",
          bold=True, size=10, bg=BG_LIGHT, align="left")

    # Amount header
    rh(36, 14)
    merge("A36", "H36", "Amount:", size=8, bold=True)
    merge("I36", "N36", "", size=8)

    # Payment rows: Lead, CD, NM, Total
    total_amount = 0
    if p:
        total_amount = float(p.total_amount or 0)

    cd_exp = None
    nm_exp = None
    if p:
        for exp in p.expenses:
            if exp.expense_type == "CD Payment":
                cd_exp = exp
            elif exp.expense_type == "Meter":
                nm_exp = exp

    pay_rows = [
        ("Lead", safe(p.total_amount if p else "", fmt="inr")),
        ("CD",   safe(cd_exp.amount if cd_exp else "", fmt="inr")),
        ("NM",   safe(nm_exp.amount if nm_exp else "", fmt="inr")),
        ("Total", safe(p.total_receivable if p else "", fmt="inr")),
    ]
    for i, (lbl, amt) in enumerate(pay_rows, 37):
        rh(i, 14)
        bold_row = lbl == "Total"
        merge(f"A{i}", f"H{i}", lbl, size=8, bold=bold_row)
        merge(f"I{i}", f"N{i}", amt, size=8, bold=bold_row)

    # ──────────────────────────────────────────────────────────────────────────
    # ROW 41 — Customer Signature
    # ──────────────────────────────────────────────────────────────────────────
    rh(41, 20)
    merge("A41", "H41", "", size=8)
    merge("I41", "N41", "Customer Signature",
          size=8, bold=True, align="center")

    wb.save(output_path)
    return output_path


if __name__ == "__main__":
    path = build_job_card(project=None, output_path="/tmp/job_card_blank.xlsx")
    print(f"Saved: {path}")