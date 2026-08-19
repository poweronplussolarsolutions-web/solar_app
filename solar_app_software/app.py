"""
Power on Plus Solar Solutions — Flask Web Application
=====================================================
Run:  python app.py
Deps: pip install flask flask-sqlalchemy flask-login flask-migrate pymysql flask-wtf flask-limiter
"""
import pymysql
pymysql.install_as_MySQLdb()
from flask import Flask, render_template, redirect, url_for, request, flash, jsonify, session, make_response
from flask_sqlalchemy import SQLAlchemy
from flask_login import LoginManager, UserMixin, login_user, logout_user, login_required, current_user
from werkzeug.security import generate_password_hash, check_password_hash
from functools import wraps
from datetime import datetime, date, timezone, timedelta
from decimal import Decimal
import os
import re
from pywebpush import webpush, WebPushException
import json as _json
from flask import send_from_directory
from flask import send_file
import tempfile, calendar
# from solar_app_software.logging_system import (
#     setup_logging, security_log,
#     log_login_attempt, log_lockout, log_password_change,
#     log_admin_action, log_access_denied,
# )
from sqlalchemy.orm import selectinload,joinedload,noload
from flask_login import user_loaded_from_request
# ── Security imports ──────────────────────────────────────────────────────────
from flask_wtf.csrf import CSRFProtect
from flask_limiter import Limiter
from flask_limiter.util import get_remote_address
from flask_wtf.csrf import generate_csrf
from dotenv import load_dotenv
load_dotenv()

DCR_SUBSIDY_AMOUNT = 78000
SERVICE_BACKFILL_FLAG_DAYS = 30

PROJECT_STAGES = [
    'Lead', 'Site Visit', 'Documentation',
    'Onsite Work', 'Connection', 'Subsidy', 'Payment',
]
STAGE_STATUS_MAP = {
    'Lead': 'Lead', 'Site Visit': 'InProgress', 'Documentation': 'InProgress',
    'Onsite Work': 'InProgress', 'Connection': 'InProgress',
    'Subsidy': 'InProgress', 'Payment': 'InProgress',
}

# ── Login-attempt tracking (in-memory; swap for Redis in production) ──────────
_login_attempts: dict = {}   # ip -> {'count': int, 'locked_until': datetime|None}
MAX_LOGIN_ATTEMPTS  = 5
LOCKOUT_MINUTES     = 15


def get_document_stages():
    return DocumentStage.query.filter_by(is_active=True).order_by(DocumentStage.sort_order).all()


def get_expected_docs(project_type, project_subtype=None, loan_subtype=None, work_category='Installation'):
    docs = []
    for stage in get_document_stages():
        if work_category == 'Outside':
            if stage.condition in ('always', 'outside'):
                docs.extend(stage.doc_list)
            continue
        
        if stage.condition == 'always':
            docs.extend(stage.doc_list)
        elif stage.condition == 'loan' and project_type == 'Loan':
            docs.extend(stage.doc_list)
        elif stage.condition == 'loan_self' and project_type == 'Loan' and loan_subtype != 'Assisted':
            docs.extend(stage.doc_list)
        elif stage.condition == 'dcr' and project_subtype == 'DCR':
            docs.extend(stage.doc_list)
    return docs




def get_doc_completion(project):
    expected = get_expected_docs(project.project_type, project.project_subtype,
                                  project.loan_subtype, project.work_category)
    recorded = {d.doc_type: d for d in project.documents}
    done_count = sum(1 for n in expected if n in recorded and recorded[n].status in ('Received','Sent','Completed'))
    return done_count, len(expected)


# ─────────────────────────────────────────────────────────────────────────────
# APP CONFIG
# ─────────────────────────────────────────────────────────────────────────────

app = Flask(__name__)

# ── Secret key: MUST be set via environment variable in production ────────────
_secret = os.environ.get('SECRET_KEY', '')
if not _secret:
    import warnings
    warnings.warn(
        "SECRET_KEY env var not set — using insecure fallback. "
        "Set SECRET_KEY before deploying to production.",
        stacklevel=1,
    )
    _secret = 'solar-dev-only-insecure-key-change-me'
app.config['SECRET_KEY'] = _secret
db_port=os.getenv("DB_PORT")or"3306"
app.config['SQLALCHEMY_DATABASE_URI']=os.getenv("DATABASE_URL")
app.config['SQLALCHEMY_ENGINE_OPTIONS'] = {
    'pool_recycle': 180,        
    'pool_pre_ping': True,      
    'pool_timeout': 5,
    'pool_size': 5,
    'max_overflow': 10,
    'connect_args':{'connect_timeout':5},
}
# @app.teardown_appcontext
# def shutdown_session(exception=None):
#     try:
#         db.session.remove()
#     except Exception as e:
#         print("Session cleanup error:", e)
# app.config['SQLALCHEMY_DATABASE_URI'] = (
#     f"mysql+pymysql://{os.getenv('DB_USER')}:{os.getenv('DB_PASSWORD')}"
#     f"@{os.getenv('DB_HOST')}:{os.getenv('DB_PORT')}/{os.getenv('DB_NAME')}"
# )
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False

# ── Session / cookie hardening ────────────────────────────────────────────────
app.config['SESSION_COOKIE_HTTPONLY'] = True          # JS cannot read cookie
app.config['SESSION_COOKIE_SAMESITE'] = 'Lax'        # CSRF mitigation
app.config['SESSION_COOKIE_SECURE']   = os.environ.get('FLASK_ENV') == 'production'
app.config['PERMANENT_SESSION_LIFETIME'] = timedelta(days=7)
app.config['WTF_CSRF_TIME_LIMIT']     = 3600          # CSRF token valid 1 h

db           = SQLAlchemy(app)
login_manager = LoginManager(app)
login_manager.login_view          = 'login'
login_manager.login_message_category = 'info'

def get_db_connection():
    return mysql.connector.connect(
        host=os.getenv("DB_HOST"),
        user=os.getenv("DB_USER"),
        password=os.getenv("DB_PASSWORD"),
        database=os.getenv("DB_NAME"),
        port=int(os.getenv("DB_PORT")),  
        ssl_disabled=False
    )

import mysql.connector
import os
# app.secret_key=os.getenv("SECRET_KEY")
# ── CSRF protection (covers all POST/PUT/DELETE forms automatically) ──────────
csrf = CSRFProtect(app)

# ── Rate limiter ─────────────────────────────────────────────────────────────
limiter = Limiter(
    key_func=get_remote_address,
    app=app,
    default_limits=[],          # no global limit; apply per-route
    storage_uri='memory://',    # swap to 'redis://localhost:6379' in production
)
# setup_logging(app)

# ── Security headers (injected on every response) ────────────────────────────
@app.after_request
def set_security_headers(response):
    response.headers['X-Content-Type-Options']    = 'nosniff'
    response.headers['X-Frame-Options']           = 'SAMEORIGIN'
    response.headers['X-XSS-Protection']          = '1; mode=block'
    response.headers['Referrer-Policy']           = 'strict-origin-when-cross-origin'
    response.headers['Permissions-Policy']        = 'geolocation=(), microphone=(), camera=()'
    # Tight CSP — adjust 'unsafe-inline' once you move styles/scripts to files
    response.headers['Content-Security-Policy'] = (
    "default-src 'self'; "
    "script-src 'self' 'unsafe-inline' https://cdn.jsdelivr.net https://cdnjs.cloudflare.com; "
    "style-src 'self' 'unsafe-inline' https://fonts.googleapis.com https://cdn.jsdelivr.net; "
    "font-src 'self' https://fonts.gstatic.com; "
    "img-src 'self' data: https://*.tile.openstreetmap.org; "
    "connect-src 'self' https://nominatim.openstreetmap.org https://fcm.googleapis.com; "
    "worker-src 'self'; "
    "manifest-src 'self';"
)
    return response


# ─────────────────────────────────────────────────────────────────────────────
# INPUT HELPERS
# ─────────────────────────────────────────────────────────────────────────────

def _clean(value: str, max_len: int = 255) -> str:
    """Strip whitespace and enforce a maximum length."""
    return (value or '').strip()[:max_len]


def _safe_float(value, default: float = 0.0) -> float:
    try:
        v = float(str(value).replace(',', '.').strip())
        if v < 0 or v > 1_000_000_000:
            return default
        return round(v, 2)
    except (TypeError, ValueError):
        return default


def _validate_password(password: str) -> list[str]:
    """Return list of unmet password requirements."""
    errors = []
    if len(password) < 8:
        errors.append('at least 8 characters')
    if not re.search(r'[A-Z]', password):
        errors.append('one uppercase letter')
    if not re.search(r'[0-9]', password):
        errors.append('one digit')
    return errors


# ─────────────────────────────────────────────────────────────────────────────
# MODELS
# ─────────────────────────────────────────────────────────────────────────────

class User(UserMixin, db.Model):
    __tablename__ = 'users'
    id         = db.Column(db.Integer, primary_key=True)
    username   = db.Column(db.String(80),  unique=True, nullable=False)
    email      = db.Column(db.String(120), nullable=False)
    phone=db.Column(db.String(20),unique=True,nullable=False)
    password   = db.Column(db.String(512), nullable=False)   
    full_name  = db.Column(db.String(120), nullable=False)
    role       = db.Column(db.Enum('admin','coordinator','documents','payments','onsite','service','office','documents_k','stocks','director'), nullable=False)
    is_active  = db.Column(db.Boolean, default=True)
    is_deleted  = db.Column(db.Boolean, default=False)
    created_at = db.Column(db.DateTime, default=datetime.utcnow)
    status     = db.Column(db.String(20), nullable=False, default='active')
    # Login-attempt tracking persisted to DB (supplements in-memory cache)
    failed_logins   = db.Column(db.Integer, default=0)
    locked_until    = db.Column(db.DateTime, nullable=True)

    def set_password(self, raw: str):
        self.password = generate_password_hash(raw, method='scrypt')

    def check_password(self, raw: str) -> bool:
        return check_password_hash(self.password, raw)

    def is_locked(self) -> bool:
        if self.locked_until and datetime.utcnow() < self.locked_until:
            return True
        return False

    def record_failed_login(self):
        self.failed_logins = (self.failed_logins or 0) + 1
        if self.failed_logins >= MAX_LOGIN_ATTEMPTS:
            self.locked_until = datetime.utcnow() + timedelta(minutes=LOCKOUT_MINUTES)

    def reset_login_attempts(self):
        self.failed_logins = 0
        self.locked_until  = None


class Customer(db.Model):
    __tablename__ = 'customers'
    id         = db.Column(db.Integer, primary_key=True)
    name       = db.Column(db.String(120), nullable=False)
    phone      = db.Column(db.String(20))
    email      = db.Column(db.String(120))
    house_name = db.Column(db.String(120))
    place      = db.Column(db.String(120))
    post       = db.Column(db.String(120))
    pincode    = db.Column(db.String(10))
    village    = db.Column(db.String(120))
    district   = db.Column(db.String(80))
    taluk      = db.Column(db.String(120))
    created_at = db.Column(db.DateTime, default=datetime.utcnow)
    projects   = db.relationship('Project', backref='customer', lazy=True)
    sub_co=db.Column(db.String(120), nullable=True)
    @property
    def full_address(self):
        parts = [self.house_name, self.place, self.post, self.village, self.district, self.pincode]
        return ', '.join(p for p in parts if p) or None

class Project(db.Model):
    __tablename__ = 'projects'
    id               = db.Column(db.Integer, primary_key=True)
    project_code     = db.Column(db.String(20), unique=True, nullable=False)
    customer_id      = db.Column(db.Integer, db.ForeignKey('customers.id'), nullable=False)
    inverter_capacity_kw = db.Column(db.Float, nullable=False)
    panel_capacity_kw    = db.Column(db.Float, nullable=False)
    structure_capacity_kw = db.Column(db.Float, nullable=True)
    project_type     = db.Column(db.Enum('Loan', 'Cash'), nullable=False)
    status           = db.Column(db.Enum('Lead','Created','InProgress','Completed','Delayed','Pending','Closed','OnHold','Cancelled'), default='Lead')
    stage            = db.Column(db.String(100), default='Lead')
    total_amount     = db.Column(db.Numeric(12, 2), default=0)
    collected_amount = db.Column(db.Numeric(12, 2), default=0)
    coordinator_id   = db.Column(db.Integer, db.ForeignKey('users.id'))
    doc_staff_id     = db.Column(db.Integer, db.ForeignKey('users.id'))
    notes            = db.Column(db.Text)
    created_at       = db.Column(db.DateTime, default=datetime.utcnow)
    updated_at       = db.Column(db.DateTime, default=datetime.utcnow, onupdate=datetime.utcnow)
    staged_changed_at = db.Column(db.DateTime, default=datetime.utcnow)
    project_subtype  = db.Column(db.Enum('DCR','Non-DCR'), nullable=True)
    coordinator      = db.relationship('User', foreign_keys=[coordinator_id], backref='coordinated_projects')
    doc_staff        = db.relationship('User', foreign_keys=[doc_staff_id],   backref='doc_projects')
    payments         = db.relationship('Payment',          backref='project', lazy=True)
    documents        = db.relationship('Document',         backref='project', lazy=True)
    logs             = db.relationship('ProjectLog',       backref='project', lazy=True)
    materials        = db.relationship('Material',         backref='project', lazy=True)
    assignments      = db.relationship('WorkerAssignment', backref='project', lazy=True)
    loan_subtype     = db.Column(db.Enum('Assisted','Self'), nullable=True)
    roof_type=db.Column(db.Enum('Flat','Sheet','Slope','Clay Tile','Tile','Other'),nullable=True)
    roof_type_other=db.Column(db.String(60), nullable=True)
    inverter_type = db.Column(db.Enum('Standard','Hybrid','String'), nullable=True)
    panel_items      = db.relationship('PanelItem', backref='project', lazy=True, cascade='all,delete-orphan')
    extra_materials  = db.relationship('ExtraMaterial', backref='project', lazy=True, cascade='all,delete-orphan')
    coordinator_name = db.Column(db.String(120), nullable=True)
    work_category = db.Column(db.Enum('Installation','Outside'), nullable=False, default='Installation')

    @property
    def contract_amount(self):
        return float(self.total_amount or 0)

    @property
    def company_paid_expenses(self):
        return [e for e in self.expenses if e.paid_by == 'Company']

    @property
    def company_expense_total(self):
        return sum(float(e.amount) for e in self.company_paid_expenses)

    @property
    def total_receivable(self):
        return self.contract_amount + self.company_expense_total

    @property
    def recovered_expense_total(self):
        return sum(float(e.amount) for e in self.company_paid_expenses if e.recovered)
    @property
    def total_waived(self):
        return sum(float(w.amount) for w in self.waivers)
    @property
    def pending_amount(self):
        sub_customer_share = 0
        if self.subsidy and self.subsidy.customer_share and self.subsidy.status == 'Received':
            sub_customer_share = float(self.subsidy.customer_share)
        return max(0, self.total_receivable - self.effective_collected
                   - sub_customer_share - self.total_waived)

    @property
    def effective_collected(self):
        company_share = 0
        if self.subsidy and self.subsidy.company_share and self.subsidy.status == 'Received':
            company_share = float(self.subsidy.company_share)
        return float(self.collected_amount or 0) + self.recovered_expense_total + company_share

    @property
    def payment_pct(self):
        t = self.total_receivable
        if t == 0:
            return 0
        return min(100, int(self.effective_collected / t * 100))

    @property
    def days_open(self):
        if self.status in ('Closed', 'Completed', 'Cancelled'):
            end = (self.updated_at or datetime.utcnow()).date()
        else:
            end = datetime.utcnow().date()
        return (end - self.created_at.date()).days

    @property
    def days_in_stage(self):
        if self.status in ('Closed', 'Completed', 'Cancelled'):
            end = (self.updated_at or datetime.utcnow()).replace(tzinfo=None)
        else:
            end = datetime.now(timezone.utc).replace(tzinfo=None)
        ref = self.staged_changed_at or self.updated_at or self.created_at
        if ref is None:
            return 0
        return (end - ref).days

    @property
    def bank_instalments(self):
        pays = [p for p in self.payments if p.payment_source == 'Bank']
        return {p.instalment: p for p in pays}

    @property
    def next_bank_instalment(self):
        done = self.bank_instalments
        if 'First' not in done:
            return 'First'
        if 'Second' not in done:
            return 'Second'
        return None
    @property
    def first_payment_received(self):
        if self.project_type != 'Loan':
            return None  # not applicable
        return any(
        p.payment_source == 'Bank' and p.instalment == 'First'
        for p in self.payments
    )
    @property
    def roof_type_display(self):
        if self.roof_type == 'Other':
            return self.roof_type_other or 'Other'
        if self.roof_type == 'Clay Tile':
            return 'Clay Tile (Oodu)'
        return self.roof_type or '—'
class ConnectionDetails(db.Model):
    __tablename__ = 'connection_details'
    id                       = db.Column(db.Integer, primary_key=True)
    project_id               = db.Column(db.Integer, db.ForeignKey('projects.id'), unique=True, nullable=False)
    connection_type          = db.Column(db.Enum('Single Phase', 'Three Phase'), nullable=True)
    category          = db.Column(db.Enum('Residential', 'Commercial'), nullable=True)
    consumer_number = db.Column(db.String(50),nullable=True)
    kseb_section = db.Column(db.String(100), nullable=True)
    ownership_change_needed  = db.Column(db.Boolean, default=False)
    ownership_change_status  = db.Column(db.Enum('Not Required', 'Pending', 'InProgress', 'Completed'), default='Not Required')
    load_clearance_needed    = db.Column(db.Boolean, default=False)
    load_clearance_status    = db.Column(db.Enum('Not Required', 'Pending', 'InProgress', 'Completed'), default='Not Required')
    notes                    = db.Column(db.Text, nullable=True)
    updated_at               = db.Column(db.DateTime, default=datetime.utcnow, onupdate=datetime.utcnow)
    updated_by               = db.Column(db.Integer, db.ForeignKey('users.id'), nullable=True)
    project                  = db.relationship('Project', backref=db.backref('connection_details', uselist=False))
    updater                  = db.relationship('User', foreign_keys=[updated_by])
class LoanDetail(db.Model):
    __tablename__ = 'loan_details'
    id          = db.Column(db.Integer, primary_key=True)
    project_id  = db.Column(db.Integer, db.ForeignKey('projects.id'), unique=True, nullable=False)
    bank_name   = db.Column(db.String(120), nullable=True)
    loan_amount = db.Column(db.Numeric(12, 2), nullable=True)
    notes       = db.Column(db.String(300), nullable=True)
    updated_at  = db.Column(db.DateTime, default=datetime.utcnow, onupdate=datetime.utcnow)
    updated_by  = db.Column(db.Integer, db.ForeignKey('users.id'), nullable=True)
    project     = db.relationship('Project', backref=db.backref('loan_detail', uselist=False))
    updater     = db.relationship('User', foreign_keys=[updated_by])
class Document(db.Model):
    __tablename__ = 'documents'
    id            = db.Column(db.Integer, primary_key=True)
    project_id    = db.Column(db.Integer, db.ForeignKey('projects.id'), nullable=False)
    doc_type      = db.Column(db.String(80), nullable=False)
    status        = db.Column(db.Enum('Pending','Received','Sent','Completed'), default='Pending')
    received_date = db.Column(db.Date)
    notes         = db.Column(db.Text)
    created_at    = db.Column(db.DateTime, default=datetime.utcnow)


class Payment(db.Model):
    __tablename__ = 'payments'
    id            = db.Column(db.Integer, primary_key=True)
    project_id    = db.Column(db.Integer, db.ForeignKey('projects.id'), nullable=False)
    amount        = db.Column(db.Numeric(12, 2), nullable=False)
    payment_type  = db.Column(db.Enum('Cash','Bank','Cheque','Online'), nullable=False)
    payment_source = db.Column(db.Enum('Customer','Bank'), nullable=False, default='Customer')
    instalment    = db.Column(db.Enum('Full','First','Second'), nullable=True)
    payment_date  = db.Column(db.Date, nullable=False)
    reference_no  = db.Column(db.String(80))
    received_by   = db.Column(db.Integer, db.ForeignKey('users.id'))
    notes         = db.Column(db.Text)
    created_at    = db.Column(db.DateTime, default=datetime.utcnow)
class CompanyBankAdvance(db.Model):
    __tablename__ = 'company_bank_advances'
    id             = db.Column(db.Integer, primary_key=True)
    project_id     = db.Column(db.Integer, db.ForeignKey('projects.id'), unique=True, nullable=False)
    amount         = db.Column(db.Numeric(12, 2), nullable=False)
    paid_date      = db.Column(db.Date, nullable=True)
    notes          = db.Column(db.String(300), nullable=True)
    recovered      = db.Column(db.Boolean, default=False)
    recovery_date  = db.Column(db.Date, nullable=True)
    recovery_method = db.Column(db.String(50), nullable=True)
    recovery_reference = db.Column(db.String(100), nullable=True)
    recovery_notes = db.Column(db.String(300), nullable=True)
    recorded_by    = db.Column(db.Integer, db.ForeignKey('users.id'), nullable=True)
    created_at     = db.Column(db.DateTime, default=datetime.utcnow)
    updated_at     = db.Column(db.DateTime, default=datetime.utcnow, onupdate=datetime.utcnow)
    project        = db.relationship('Project', backref=db.backref('bank_advance', uselist=False))
    recorder       = db.relationship('User', foreign_keys=[recorded_by])
class BankExcessReturn(db.Model):
    __tablename__ = 'bank_excess_returns'
    id               = db.Column(db.Integer, primary_key=True)
    project_id       = db.Column(db.Integer, db.ForeignKey('projects.id'), unique=True, nullable=False)
    excess_amount    = db.Column(db.Numeric(12, 2), nullable=False)
    received_date    = db.Column(db.Date, nullable=True)
    notes            = db.Column(db.String(300), nullable=True)
    returned         = db.Column(db.Boolean, default=False)     # kept for backward compat — True once settled
    action           = db.Column(db.Enum('Returned','Retained','Adjusted'), nullable=True)   # NEW
    returned_to      = db.Column(db.Enum('Customer','Bank'), nullable=True)                  # NEW
    returned_date    = db.Column(db.Date, nullable=True)
    returned_method  = db.Column(db.String(50), nullable=True)
    returned_reference = db.Column(db.String(100), nullable=True)
    returned_notes   = db.Column(db.String(300), nullable=True)
    recorded_by      = db.Column(db.Integer, db.ForeignKey('users.id'), nullable=True)
    created_at       = db.Column(db.DateTime, default=datetime.utcnow)
    updated_at       = db.Column(db.DateTime, default=datetime.utcnow, onupdate=datetime.utcnow)
    project          = db.relationship('Project', backref=db.backref('bank_excess', uselist=False))
    recorder         = db.relationship('User', foreign_keys=[recorded_by])

    @property
    def action_label(self):
        return {
            'Returned': f'Returned to {self.returned_to}' if self.returned_to else 'Returned to Bank',
            'Retained': 'Retained as company income',
            'Adjusted': 'Adjusted against other dues',
        }.get(self.action, 'Returned')
class PaymentWaiver(db.Model):
    __tablename__ = 'payment_waivers'
    id           = db.Column(db.Integer, primary_key=True)
    project_id   = db.Column(db.Integer, db.ForeignKey('projects.id'), nullable=False)
    amount       = db.Column(db.Numeric(12, 2), nullable=False)
    reason       = db.Column(db.String(500), nullable=False)
    waived_date  = db.Column(db.Date, default=date.today)
    waived_by    = db.Column(db.Integer, db.ForeignKey('users.id'), nullable=True)
    created_at   = db.Column(db.DateTime, default=datetime.utcnow)

    project = db.relationship('Project', backref='waivers')
    waiver  = db.relationship('User', foreign_keys=[waived_by])
class PaymentExcess(db.Model):
    __tablename__ = 'payment_excess'
    id                   = db.Column(db.Integer, primary_key=True)
    project_id           = db.Column(db.Integer, db.ForeignKey('projects.id'), nullable=False)
    amount               = db.Column(db.Numeric(12, 2), nullable=False)
    source               = db.Column(db.Enum('Customer','Bank'), nullable=False, default='Customer')
    detected_date        = db.Column(db.Date, default=date.today)
    notes                = db.Column(db.String(300), nullable=True)
    status               = db.Column(db.Enum('Pending','Settled'), default='Pending')
    action               = db.Column(db.Enum('Returned','Retained','Adjusted'), nullable=True)
    returned_to          = db.Column(db.Enum('Customer','Bank'), nullable=True)
    settlement_method    = db.Column(db.String(50), nullable=True)
    settlement_date      = db.Column(db.Date, nullable=True)
    settlement_reference = db.Column(db.String(100), nullable=True)
    settlement_notes     = db.Column(db.String(300), nullable=True)
    recorded_by          = db.Column(db.Integer, db.ForeignKey('users.id'), nullable=True)
    settled_by           = db.Column(db.Integer, db.ForeignKey('users.id'), nullable=True)
    created_at           = db.Column(db.DateTime, default=datetime.utcnow)
    updated_at           = db.Column(db.DateTime, default=datetime.utcnow, onupdate=datetime.utcnow)
    project               = db.relationship('Project', backref='payment_excesses')
    recorder              = db.relationship('User', foreign_keys=[recorded_by])
    settler                = db.relationship('User', foreign_keys=[settled_by])

    @property
    def action_label(self):
        return {
            'Returned': f'Returned to {self.returned_to}' if self.returned_to else 'Returned',
            'Retained': 'Retained as company income',
            'Adjusted': 'Adjusted against other dues',
        }.get(self.action, '—')
class Worker(db.Model):
    __tablename__ = 'workers'
    id           = db.Column(db.Integer, primary_key=True)
    name         = db.Column(db.String(100), nullable=False)
    phone        = db.Column(db.String(20))
    skill        = db.Column(db.String(80))
    rate_per_day = db.Column(db.Numeric(8, 2), default=0)
    is_active    = db.Column(db.Boolean, default=True)
    created_at   = db.Column(db.DateTime, default=datetime.utcnow)
    assignments  = db.relationship('WorkerAssignment', backref='worker', lazy=True)
    weekly_payments = db.relationship('WorkerWeeklyPayment', back_populates='worker',
                                      order_by='WorkerWeeklyPayment.week_start.desc()')


weekly_pay_project = db.Table(
    'weekly_pay_project',
    db.Column('payment_id', db.Integer, db.ForeignKey('worker_weekly_payment.id'), primary_key=True),
    db.Column('project_id', db.Integer, db.ForeignKey('projects.id'), primary_key=True),
)


class WorkerAssignment(db.Model):
    __tablename__ = 'worker_assignments'
    id          = db.Column(db.Integer, primary_key=True)
    project_id  = db.Column(db.Integer, db.ForeignKey('projects.id'), nullable=False)
    worker_id   = db.Column(db.Integer, db.ForeignKey('workers.id'),  nullable=False)
    start_date  = db.Column(db.Date)
    end_date    = db.Column(db.Date)
    days_worked = db.Column(db.Integer, default=0)
    work_phase  = db.Column(db.Enum('Structure','Electrical/Installation','Full Work'), default='Structure')
    status      = db.Column(db.Enum('Assigned','Active','Completed','Paid'), default='Assigned')
    created_at  = db.Column(db.DateTime, default=datetime.utcnow)


class WorkerWeeklyPayment(db.Model):
    __tablename__ = 'worker_weekly_payment'
    __table_args__ = (
        db.UniqueConstraint('worker_id', 'week_start', name='uq_worker_week'),
    )
    id           = db.Column(db.Integer, primary_key=True)
    worker_id    = db.Column(db.Integer, db.ForeignKey('workers.id'), nullable=False)
    week_start   = db.Column(db.Date, nullable=False)
    week_end     = db.Column(db.Date, nullable=False)
    days_worked  = db.Column(db.Numeric(4, 1), nullable=False)
    rate_per_day = db.Column(db.Numeric(10, 2), nullable=False)
    amount       = db.Column(db.Numeric(10, 2), nullable=False)
    paid_date    = db.Column(db.Date, nullable=False)
    payer_id     = db.Column(db.Integer, db.ForeignKey('users.id'), nullable=True)
    notes        = db.Column(db.String(300), nullable=True)
    projects     = db.relationship('Project', secondary='weekly_pay_project', lazy='select')
    worker       = db.relationship('Worker', back_populates='weekly_payments')
    payer        = db.relationship('User', foreign_keys=[payer_id])


class KSEBTask(db.Model):
    __tablename__ = 'kseb_tasks'
    id              = db.Column(db.Integer, primary_key=True)
    project_id      = db.Column(db.Integer, db.ForeignKey('projects.id'), unique=True, nullable=False)
    stamp_paper     = db.Column(db.Enum('Pending','Requested','Received'), default='Pending')
    b_class_licence = db.Column(db.Enum('Pending','Requested','Received'), default='Pending')
    file_sent       = db.Column(db.Boolean, default=False)
    file_sent_date  = db.Column(db.Date)
    inspection_date = db.Column(db.Date)
    inspection_done = db.Column(db.Boolean, default=False)
    cd_payment_done = db.Column(db.Boolean, default=False)
    cd_payment_date = db.Column(db.Date)
    connection_date = db.Column(db.Date)
    connection_done = db.Column(db.Boolean, default=False)
    meter_available = db.Column(db.Boolean, default=True)
    ae_completed    = db.Column(db.Boolean, default=False)
    notes           = db.Column(db.Text)
    updated_at      = db.Column(db.DateTime, default=datetime.utcnow, onupdate=datetime.utcnow)
    project         = db.relationship('Project', backref=db.backref('kseb_task', uselist=False))


class Subsidy(db.Model):
    __tablename__ = 'subsidy'
    id              = db.Column(db.Integer, primary_key=True)
    project_id      = db.Column(db.Integer, db.ForeignKey('projects.id'), unique=True, nullable=False)
    request_date    = db.Column(db.Date)
    expected_amount = db.Column(db.Numeric(10, 2), default=0)
    received_amount = db.Column(db.Numeric(10, 2), default=0)
    customer_share  = db.Column(db.Numeric(10, 2), default=0)
    company_share   = db.Column(db.Numeric(10, 2), default=0)
    status          = db.Column(db.Enum('NotStarted','Processing','Commissioned','Redeemed','Received'), default='NotStarted')
    notes           = db.Column(db.Text)
    updated_at      = db.Column(db.DateTime, default=datetime.utcnow, onupdate=datetime.utcnow)
    project         = db.relationship('Project', backref=db.backref('subsidy', uselist=False))


class AppInstallation(db.Model):
    __tablename__ = 'app_installations'
    id             = db.Column(db.Integer, primary_key=True)
    project_id     = db.Column(db.Integer, db.ForeignKey('projects.id'), unique=True, nullable=False)
    scheduled_date = db.Column(db.Date)
    completed_date = db.Column(db.Date)
    installed_by   = db.Column(db.Integer, db.ForeignKey('users.id'))
    status         = db.Column(db.Enum('Pending','Scheduled','Completed'), default='Pending')
    notes          = db.Column(db.Text)
    created_at     = db.Column(db.DateTime, default=datetime.utcnow)
    project        = db.relationship('Project', backref=db.backref('app_install', uselist=False))
class PanelDetails(db.Model):
    __tablename__ = 'panel_details'
    id                      = db.Column(db.Integer, primary_key=True)
    project_id              = db.Column(db.Integer, db.ForeignKey('projects.id'), unique=True, nullable=False)
    panel_brand             = db.Column(db.String(100), nullable=True)
    inverter_brand             = db.Column(db.String(100), nullable=True)
    num_panels              = db.Column(db.Integer,     nullable=True)
    panel_serial_numbers    = db.Column(db.Text,        nullable=True)   # newline-separated
    inverter_serial_number  = db.Column(db.String(100), nullable=True)
    net_meter_serial_number = db.Column(db.String(100), nullable=True)
    energy_meter_serial_number = db.Column(db.String(100), nullable=True)
    app_id                  = db.Column(db.String(100), nullable=True)   
    app_password             = db.Column(db.String(100), nullable=True)
    notes                   = db.Column(db.Text,        nullable=True)
    updated_at              = db.Column(db.DateTime, default=datetime.utcnow, onupdate=datetime.utcnow)
    updated_by              = db.Column(db.Integer, db.ForeignKey('users.id'), nullable=True)
    project                 = db.relationship('Project', backref=db.backref('panel_details', uselist=False))
    updater                 = db.relationship('User', foreign_keys=[updated_by])

class ServiceRecord(db.Model):
   
    __tablename__ = 'service_records'
 
    id              = db.Column(db.Integer, primary_key=True)
    project_id      = db.Column(db.Integer, db.ForeignKey('projects.id'), nullable=False)
 
    visit_number    = db.Column(db.Integer, nullable=False)          # 1–5
    scheduled_date  = db.Column(db.Date, nullable=False)
    completed_date  = db.Column(db.Date, nullable=True)
    conducted_by    = db.Column(db.Integer, db.ForeignKey('users.id'), nullable=True)
 
    status          = db.Column(
        db.Enum('Upcoming', 'Due', 'Overdue', 'Completed', 'Skipped'),
        default='Upcoming', nullable=False
    )
 
    # Checklist
    panel_cleaning         = db.Column(db.Boolean, default=False)
    
    notes                  = db.Column(db.Text, nullable=True)
 
    created_at  = db.Column(db.DateTime, default=datetime.utcnow)
    updated_at  = db.Column(db.DateTime, default=datetime.utcnow, onupdate=datetime.utcnow)
 
    project     = db.relationship('Project', backref='service_records')
    technician  = db.relationship('User', foreign_keys=[conducted_by])
 
    @property
    def checklist_pct(self):
        return 100 if self.panel_cleaning else 0
 
    @property
    def is_overdue(self):
        return (
            self.status not in ('Completed', 'Skipped')
            and self.scheduled_date < date.today()
        )


class SiteVisit(db.Model):
    __tablename__ = 'site_visits'
    id             = db.Column(db.Integer, primary_key=True)
    project_id     = db.Column(db.Integer, db.ForeignKey('projects.id'), nullable=False)
    scheduled_date = db.Column(db.Date)
    visited_date   = db.Column(db.Date)
    conducted_by   = db.Column(db.Integer, db.ForeignKey('users.id'))
    observations   = db.Column(db.Text)
    status         = db.Column(db.Enum('Scheduled','Completed','Cancelled'), default='Scheduled')
    created_at     = db.Column(db.DateTime, default=datetime.utcnow)
    project        = db.relationship('Project', backref='site_visits')


class ProjectLog(db.Model):
    __tablename__ = 'project_logs'
    id         = db.Column(db.Integer, primary_key=True)
    project_id = db.Column(db.Integer, db.ForeignKey('projects.id'), nullable=False)
    action     = db.Column(db.String(200), nullable=False)
    old_value  = db.Column(db.String(100))
    new_value  = db.Column(db.String(100))
    done_by    = db.Column(db.Integer, db.ForeignKey('users.id'))
    created_at = db.Column(db.DateTime, default=datetime.utcnow)
    user       = db.relationship('User', foreign_keys=[done_by])


class Material(db.Model):
    __tablename__ = 'materials'
    id              = db.Column(db.Integer, primary_key=True)
    project_id      = db.Column(db.Integer, db.ForeignKey('projects.id'), nullable=False)
    item_name       = db.Column(db.String(100), nullable=False)
    quantity        = db.Column(db.Numeric(8, 2))
    unit            = db.Column(db.String(20))
    dispatch_status = db.Column(db.Enum('Pending','Dispatched','Delivered'), default='Pending')
    dispatch_date   = db.Column(db.Date)
    received_date   = db.Column(db.Date)
    notes           = db.Column(db.Text)
    created_at      = db.Column(db.DateTime, default=datetime.utcnow)


class Notification(db.Model):
    __tablename__ = 'notifications'
    id         = db.Column(db.Integer, primary_key=True)
    user_id    = db.Column(db.Integer, db.ForeignKey('users.id'), nullable=False)
    project_id = db.Column(db.Integer, db.ForeignKey('projects.id'), nullable=True)
    message    = db.Column(db.String(255), nullable=False)
    notif_type = db.Column(db.String(80), default='info')
    is_read    = db.Column(db.Boolean, default=False)
    action_url = db.Column(db.String(255), nullable=True)
    created_at = db.Column(db.DateTime, default=datetime.utcnow)
    user       = db.relationship('User', foreign_keys=[user_id])
    project    = db.relationship('Project', foreign_keys=[project_id])


class OnsiteProgress(db.Model):
    __tablename__ = 'onsite_progress'
    id                     = db.Column(db.Integer, primary_key=True)
    project_id             = db.Column(db.Integer, db.ForeignKey('projects.id'), unique=True, nullable=False)
    materials_ordered        = db.Column(db.Boolean, default=False)
    materials_ordered_date   = db.Column(db.Date, nullable=True)
    materials_delivered      = db.Column(db.Boolean, default=False)
    materials_delivered_date = db.Column(db.Date, nullable=True)
    structure_work_status  = db.Column(db.Enum('NotStarted','InProgress','Completed'), default='NotStarted')
    structure_start_date   = db.Column(db.Date)
    structure_end_date     = db.Column(db.Date)
    structure_notes        = db.Column(db.Text)
    installation_status    = db.Column(db.Enum('NotStarted','InProgress','Completed'), default='NotStarted')
    installation_start_date = db.Column(db.Date)
    installation_end_date   = db.Column(db.Date)
    installation_notes     = db.Column(db.Text)
    electrical_status      = db.Column(db.Enum('NotStarted','InProgress','Completed'), default='NotStarted')
    electrical_start_date  = db.Column(db.Date)
    electrical_end_date    = db.Column(db.Date)
    electrical_notes       = db.Column(db.Text)
    updated_at             = db.Column(db.DateTime, default=datetime.utcnow, onupdate=datetime.utcnow)
    updated_by             = db.Column(db.Integer, db.ForeignKey('users.id'))
    project                = db.relationship('Project', backref=db.backref('onsite_progress', uselist=False))
    important_notes=db.Column(db.Text, nullable=True)


class OnsiteLog(db.Model):
    __tablename__ = 'onsite_logs'
    id         = db.Column(db.Integer, primary_key=True)
    project_id = db.Column(db.Integer, db.ForeignKey('projects.id'), nullable=False)
    log_date   = db.Column(db.Date, nullable=False, default=date.today)
    work_phase = db.Column(db.Enum('Structure','Full Work','Electrical/Installation'), nullable=False)
    note       = db.Column(db.String(500), nullable=False)
    logged_by  = db.Column(db.Integer, db.ForeignKey('users.id'))
    created_at = db.Column(db.DateTime, default=datetime.utcnow)
    project    = db.relationship('Project', backref='onsite_logs')
    logger     = db.relationship('User', foreign_keys=[logged_by])


class JobCard(db.Model):
    __tablename__ = 'job_cards'
    id           = db.Column(db.Integer, primary_key=True)
    project_id   = db.Column(db.Integer, db.ForeignKey('projects.id'), nullable=False)
    worker_id    = db.Column(db.Integer, db.ForeignKey('workers.id'),  nullable=False)
    work_phase   = db.Column(db.Enum('Structure','Electrical/Installation','Full Work'), nullable=False)
    description  = db.Column(db.String(300))
    agreed_amount   = db.Column(db.Numeric(10, 2), nullable=True)
    actual_days     = db.Column(db.Numeric(4, 1),  nullable=True)
    rate_per_day    = db.Column(db.Numeric(10, 2), nullable=True)
    final_amount    = db.Column(db.Numeric(10, 2), nullable=True)
    status = db.Column(db.Enum('Open','Approved','Paid','Voided'), default='Open')
    created_at   = db.Column(db.DateTime, default=datetime.utcnow)
    closed_at    = db.Column(db.DateTime, nullable=True)
    approved_at  = db.Column(db.DateTime, nullable=True)
    approved_by  = db.Column(db.Integer, db.ForeignKey('users.id'), nullable=True)
    closed_by    = db.Column(db.Integer, db.ForeignKey('users.id'), nullable=True)
    project      = db.relationship('Project', backref='job_cards')
    worker       = db.relationship('Worker',  backref='job_cards')
    approver     = db.relationship('User', foreign_keys=[approved_by])
    closer       = db.relationship('User', foreign_keys=[closed_by])


class WorkerAdvance(db.Model):
    __tablename__ = 'worker_advances'
    id               = db.Column(db.Integer, primary_key=True)
    worker_id        = db.Column(db.Integer, db.ForeignKey('workers.id'), nullable=False)
    project_id       = db.Column(db.Integer, db.ForeignKey('projects.id'), nullable=True)
    amount           = db.Column(db.Numeric(10, 2), nullable=False)
    given_date       = db.Column(db.Date, nullable=False)
    given_by         = db.Column(db.Integer, db.ForeignKey('users.id'))
    recovered_amount = db.Column(db.Numeric(10, 2), default=0)
    status           = db.Column(db.Enum('Outstanding','PartiallyRecovered','Cleared'), default='Outstanding')
    notes            = db.Column(db.String(300))
    created_at       = db.Column(db.DateTime, default=datetime.utcnow)
    worker  = db.relationship('Worker',  backref='advances')
    giver   = db.relationship('User',    foreign_keys=[given_by])
    project = db.relationship('Project', backref='worker_advances')


class WorkerLedger(db.Model):
    __tablename__ = 'worker_ledger'
    id             = db.Column(db.Integer, primary_key=True)
    worker_id      = db.Column(db.Integer, db.ForeignKey('workers.id'), nullable=False)
    entry_date     = db.Column(db.Date, nullable=False)
    entry_type     = db.Column(db.Enum('Earning','Advance','Deduction','Settlement','Bonus'), nullable=False)
    amount         = db.Column(db.Numeric(10, 2), nullable=False)
    direction      = db.Column(db.Enum('Credit','Debit'), nullable=False)
    reference_type = db.Column(db.String(40))
    reference_id   = db.Column(db.Integer)
    notes          = db.Column(db.String(300))
    balance_after  = db.Column(db.Numeric(10, 2))
    recorded_by    = db.Column(db.Integer, db.ForeignKey('users.id'))
    created_at     = db.Column(db.DateTime, default=datetime.utcnow)
    worker    = db.relationship('Worker', backref='ledger_entries')
    recorder  = db.relationship('User',   foreign_keys=[recorded_by])


class ProjectExpense(db.Model):
    __tablename__ = 'project_expenses'
    id           = db.Column(db.Integer, primary_key=True)
    project_id   = db.Column(db.Integer, db.ForeignKey('projects.id'), nullable=False)
    expense_type = db.Column(db.Enum('CD Payment', 'Meter', 'Load', 'Additional'), nullable=False)
    amount       = db.Column(db.Numeric(10, 2), nullable=False, default=0)
    paid_by      = db.Column(db.Enum('Customer', 'Company'), nullable=False, default='Customer')
    paid_date    = db.Column(db.Date, nullable=True)
    recovered    = db.Column(db.Boolean, default=False)
    recovered_date = db.Column(db.Date, nullable=True)
    recovery_method=db.Column(db.String(100),nullable=True)
    recovery_reference=db.Column(db.String(100),nullable=True)
    notes        = db.Column(db.Text, nullable=True)
    created_at   = db.Column(db.DateTime, default=datetime.utcnow)
    updated_at   = db.Column(db.DateTime, default=datetime.utcnow, onupdate=datetime.utcnow)
    recorded_by  = db.Column(db.Integer, db.ForeignKey('users.id'), nullable=True)
    project  = db.relationship('Project', backref='expenses')
    recorder = db.relationship('User', foreign_keys=[recorded_by])


class DocumentStage(db.Model):
    __tablename__ = 'document_stages'
    id         = db.Column(db.Integer, primary_key=True)
    name       = db.Column(db.String(100), nullable=False)
    condition  = db.Column(db.Enum('always', 'loan', 'loan_self', 'dcr'), nullable=False, default='always')
    docs       = db.Column(db.Text, nullable=False)
    sort_order = db.Column(db.Integer, default=0)
    is_active  = db.Column(db.Boolean, default=True)

    @property
    def doc_list(self):
        return [d.strip() for d in self.docs.split(',') if d.strip()]
class PanelItem(db.Model):
    id         = db.Column(db.Integer, primary_key=True)
    project_id = db.Column(db.Integer, db.ForeignKey('projects.id'), nullable=False)
    brand      = db.Column(db.String(20))   
    panel_type = db.Column(db.String(20))   
    wattage    = db.Column(db.Integer)
    quantity   = db.Column(db.Integer)

class ExtraMaterial(db.Model):
    id             = db.Column(db.Integer, primary_key=True)
    project_id     = db.Column(db.Integer, db.ForeignKey('projects.id'), nullable=False)
    description    = db.Column(db.String(200))
    quantity_label = db.Column(db.String(50))
class StockItem(db.Model):
    __tablename__ = 'stock_items'
    id             = db.Column(db.Integer, primary_key=True)
    category       = db.Column(db.String(50), nullable=False)   # Panel, Inverter, Battery, Other, Accessories, Welding...
    brand          = db.Column(db.String(60), nullable=True)    # Waaree, UTL, Microtek, Adani...
    subtype        = db.Column(db.String(40), nullable=True)    # DCR/Non-DCR, Ongrid/Offgrid/Hybrid, Lithium/Lead Acid...
    model_name     = db.Column(db.String(60), nullable=True)    # Gamma 1012, Sigma 5048...
    name           = db.Column(db.String(120), nullable=False)  # ACDB, Earth Rod, Welding Rod...
    unit           = db.Column(db.String(20), default='pcs')
    reorder_level  = db.Column(db.Numeric(10, 2), default=0)
    current_qty    = db.Column(db.Numeric(10, 2), default=0)    # cached — updated on every transaction
    is_active      = db.Column(db.Boolean, default=True)
    created_at     = db.Column(db.DateTime, default=datetime.utcnow)
    created_by     = db.Column(db.Integer, db.ForeignKey('users.id'), nullable=True)
    creator        = db.relationship('User', foreign_keys=[created_by])
 
    @property
    def display_name(self):
        parts = [self.brand, self.model_name, self.name]
        return ' — '.join(p for p in parts if p)
 
    @property
    def is_low(self):
        return float(self.current_qty or 0) <= float(self.reorder_level or 0)
 
 
class StockTransaction(db.Model):
    __tablename__ = 'stock_transactions'
    id             = db.Column(db.Integer, primary_key=True)
    stock_item_id  = db.Column(db.Integer, db.ForeignKey('stock_items.id'), nullable=False)
    txn_type       = db.Column(db.Enum('In', 'Out', 'Adjustment'), nullable=False)
    quantity       = db.Column(db.Numeric(10, 2), nullable=False)   # always positive; sign implied by txn_type
    source         = db.Column(db.String(30), default='Manual')      # Manual, PanelItem, ExtraMaterial, Material
    sale_type      = db.Column(db.Enum('B2B', 'B2C'), nullable=True)    
    project_id     = db.Column(db.Integer, db.ForeignKey('projects.id'), nullable=True)
    reference_id   = db.Column(db.Integer, nullable=True)           # id of the linked PanelItem/ExtraMaterial/Material row
    notes          = db.Column(db.String(300), nullable=True)
    balance_after  = db.Column(db.Numeric(10, 2), nullable=False)
    recorded_by    = db.Column(db.Integer, db.ForeignKey('users.id'))
    created_at     = db.Column(db.DateTime, default=datetime.utcnow)
    stock_item     = db.relationship('StockItem', backref='transactions')
    project        = db.relationship('Project', foreign_keys=[project_id])
    recorder       = db.relationship('User', foreign_keys=[recorded_by])
class ProjectGeoTag(db.Model):
    __tablename__ = 'project_geo_tags'
    id           = db.Column(db.Integer, primary_key=True)
    project_id   = db.Column(db.Integer, db.ForeignKey('projects.id'), unique=True, nullable=False)
    photo_path   = db.Column(db.String(255), nullable=True)      # ← Completion / geo-tag photo (AFTER work, onsite team)
    site_photo_path = db.Column(db.String(255), nullable=True)   # ← NEW: Scouting photo (BEFORE work, coordinator/docs)
    site_photo_uploaded_by = db.Column(db.Integer, db.ForeignKey('users.id'), nullable=True)
    site_photo_uploaded_at = db.Column(db.DateTime, nullable=True)
    maps_url     = db.Column(db.String(500), nullable=True)
    latitude     = db.Column(db.Float, nullable=True)
    longitude    = db.Column(db.Float, nullable=True)
    has_gps      = db.Column(db.Boolean, default=False)
    address      = db.Column(db.String(300), nullable=True)
    uploaded_by  = db.Column(db.Integer, db.ForeignKey('users.id'))
    uploaded_at  = db.Column(db.DateTime, default=datetime.utcnow)
    project      = db.relationship('Project', backref=db.backref('geo_tag', uselist=False))
    uploader     = db.relationship('User', foreign_keys=[uploaded_by])
    site_photo_uploader = db.relationship('User', foreign_keys=[site_photo_uploaded_by])
class GeocodeCache(db.Model):
    __tablename__ = 'geocode_cache'
    id         = db.Column(db.Integer, primary_key=True)
    query_text = db.Column(db.String(200), unique=True, nullable=False)
    latitude   = db.Column(db.Float, nullable=True)
    longitude  = db.Column(db.Float, nullable=True)
    created_at = db.Column(db.DateTime, default=datetime.utcnow)
class ProductCategory(db.Model):
    __tablename__ = 'product_categories'
    id         = db.Column(db.Integer, primary_key=True)
    name       = db.Column(db.String(120), unique=True, nullable=False)
    is_active  = db.Column(db.Boolean, default=True)
    created_at = db.Column(db.DateTime, default=datetime.utcnow)
    created_by = db.Column(db.Integer, db.ForeignKey('users.id'), nullable=True)
    creator    = db.relationship('User', foreign_keys=[created_by])

    @property
    def replacement_count(self):
        return ProductReplacement.query.filter_by(category_id=self.id).count()

    @property
    def pending_count(self):
        return ProductReplacement.query.filter_by(category_id=self.id, status='Pending').count()

class ProductReplacement(db.Model):
    __tablename__ = 'product_replacements'
    id                = db.Column(db.Integer, primary_key=True)
    category_id       = db.Column(db.Integer, db.ForeignKey('product_categories.id'), nullable=False)
    complaint_id      = db.Column(db.String(60), nullable=False)
    serial_number     = db.Column(db.String(100), nullable=False)   # serial of the faulty/replaced unit
    new_serial_number = db.Column(db.String(100), nullable=True)    # serial of the unit given as replacement
    replacement_date  = db.Column(db.Date, nullable=False, default=date.today)
    purchaser_name    = db.Column(db.String(120), nullable=False)
    purchaser_phone   = db.Column(db.String(20), nullable=True)
    order_id           = db.Column(db.String(20), nullable=True)    # optional free-text order ID
    notes             = db.Column(db.Text, nullable=True)
    status            = db.Column(db.Enum('Pending','Cleared'), nullable=False, default='Pending')  
    cleared_date      = db.Column(db.Date, nullable=True)            
    cleared_by        = db.Column(db.Integer, db.ForeignKey('users.id'), nullable=True)  
    recorded_by       = db.Column(db.Integer, db.ForeignKey('users.id'), nullable=True)
    created_at        = db.Column(db.DateTime, default=datetime.utcnow)
    updated_at        = db.Column(db.DateTime, default=datetime.utcnow, onupdate=datetime.utcnow)

    category = db.relationship('ProductCategory',
        backref=db.backref('replacements', lazy=True, order_by='ProductReplacement.replacement_date.desc()'))
    recorder = db.relationship('User', foreign_keys=[recorded_by])
    clearer  = db.relationship('User', foreign_keys=[cleared_by]) 
class PushSubscription(db.Model):
    __tablename__ = 'push_subscriptions'
    id         = db.Column(db.Integer, primary_key=True)
    user_id    = db.Column(db.Integer, db.ForeignKey('users.id'), nullable=False)
    endpoint   = db.Column(db.String(500), unique=True, nullable=False)
    p256dh     = db.Column(db.String(255), nullable=False)
    auth       = db.Column(db.String(255), nullable=False)
    user_agent = db.Column(db.String(255), nullable=True)
    created_at = db.Column(db.DateTime, default=datetime.utcnow)
    user       = db.relationship('User', backref='push_subscriptions') 
from urllib.parse import urlparse

ALLOWED_MAPS_HOSTS = ('google.com', 'goo.gl', 'maps.app.goo.gl')

def _is_allowed_maps_host(url):
    try:
        host = urlparse(url).netloc.lower()
    except Exception:
        return False
    return any(host == d or host.endswith('.' + d) for d in ALLOWED_MAPS_HOSTS)

def parse_gmaps_coords(url):
    """Extract (lat, lng) from any common Google Maps link format.
    Resolves shortened maps.app.goo.gl / goo.gl links first."""
    if not url:
        return None, None
    url = url.strip()
    if not _is_allowed_maps_host(url):
        return None, None

    # Shortened links need resolving to their real URL before we can regex it
    if 'goo.gl' in url:
        try:
            r = requests.head(url, allow_redirects=True, timeout=5,
                               headers={'User-Agent': 'Mozilla/5.0'})
            url = r.url
        except Exception:
            try:
                r = requests.get(url, allow_redirects=True, timeout=5,
                                  headers={'User-Agent': 'Mozilla/5.0'})
                url = r.url
            except Exception:
                return None, None

    patterns = [
        r'@(-?\d+\.\d+),(-?\d+\.\d+)',            # .../@11.140829,76.032104,15z
        r'[?&]q=(-?\d+\.\d+),(-?\d+\.\d+)',        # ?q=11.140829,76.032104
        r'!3d(-?\d+\.\d+)!4d(-?\d+\.\d+)',         # place URLs with embedded coords
        r'/place/(-?\d+\.\d+),\s*(-?\d+\.\d+)',    # /place/11.14,76.03
    ]
    for pat in patterns:
        m = re.search(pat, url)
        if m:
            lat, lng = float(m.group(1)), float(m.group(2))
            if 6 <= lat <= 36 and 68 <= lng <= 98:   # sanity-check within India
                return round(lat, 6), round(lng, 6)
    return None, None
# ─────────────────────────────────────────────────────────────────────────────
# HELPERS
# ─────────────────────────────────────────────────────────────────────────────
# _user_cache={}
from flask import g

@login_manager.user_loader
def load_user(user_id):
    if '_cached_user' not in g:
        g._cached_user = db.session.get(User, int(user_id))
    return g._cached_user
# @app.teardown_request
# def clear_user_cache(exc=None):
#     _user_cache.clear()

def roles_required(*roles):
    def decorator(f):
        @wraps(f)
        def wrapper(*args, **kwargs):
            if not current_user.is_authenticated or current_user.role not in roles:
                # log_access_denied(request.path, current_user.username)
                flash('Access denied.', 'danger')
                return redirect(url_for('dashboard'))
            return f(*args, **kwargs)
        return wrapper
    return decorator



def log_action(project_id, action, old_val=None, new_val=None):
    cutoff    = datetime.utcnow() - timedelta(seconds=5)
    duplicate = ProjectLog.query.filter_by(
        project_id=project_id,
        action=action,
        old_value=str(old_val) if old_val else None,
        new_value=str(new_val) if new_val else None,
        done_by=current_user.id,
    ).filter(ProjectLog.created_at >= cutoff).first()
    if duplicate:
        return
    entry = ProjectLog(
        project_id=project_id,
        action=action,
        old_value=str(old_val) if old_val else None,
        new_value=str(new_val) if new_val else None,
        done_by=current_user.id,
    )
    db.session.add(entry)
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
GEO_UPLOAD_DIR = os.environ.get(
    'GEO_UPLOAD_DIR',
    os.path.join(BASE_DIR, 'private_uploads', 'geo_photos')
)
@app.route('/projects/<int:pid>/geo_tag', methods=['POST'])
@login_required
@roles_required('admin', 'onsite')
def upload_geo_tag(pid):
    proj = Project.query.get_or_404(pid)
    file = request.files.get('geo_photo')
    if not file or file.filename == '':
        flash('Please select a photo to upload.', 'danger')
        return redirect(url_for('onsite_progress', pid=pid))

    ext = os.path.splitext(file.filename)[1].lower()
    if ext not in ('.jpg', '.jpeg', '.png'):
        flash('Please upload a JPG or PNG.', 'danger')
        return redirect(url_for('onsite_progress', pid=pid))

    os.makedirs(GEO_UPLOAD_DIR, exist_ok=True)
    fname = f'{proj.project_code}_{int(datetime.utcnow().timestamp())}{ext}'
    fpath = os.path.join(GEO_UPLOAD_DIR, fname)   # now always absolute
    file.save(fpath)

    tag = proj.geo_tag or ProjectGeoTag(project_id=pid)
    tag.photo_path  = fpath
    tag.uploaded_by = current_user.id
    tag.uploaded_at = datetime.utcnow()
    if not tag.id:
        db.session.add(tag)
    log_action(pid, 'Site photo uploaded', new_val=fname)
    db.session.commit()
    flash('Site photo uploaded.', 'success')
    return redirect(url_for('onsite_progress', pid=pid))

@app.route('/projects/<int:pid>/geo_tag/location', methods=['POST'])
@login_required
@roles_required('admin', 'onsite')
def save_geo_location(pid):
    proj = Project.query.get_or_404(pid)
    maps_url = _clean(request.form.get('maps_url', ''), 500)

    if not maps_url:
        flash('Please paste a Google Maps link.', 'danger')
        return redirect(url_for('onsite_progress', pid=pid))

    lat, lng = parse_gmaps_coords(maps_url)

    tag = proj.geo_tag or ProjectGeoTag(project_id=pid)
    tag.maps_url    = maps_url
    tag.uploaded_by = current_user.id
    tag.uploaded_at = datetime.utcnow()
    if not tag.id:
        db.session.add(tag)

    if lat is not None:
        tag.latitude, tag.longitude, tag.has_gps = lat, lng, True
        log_action(pid, 'Site location saved from Google Maps link', new_val=f'{lat},{lng}')
        db.session.commit()
        flash(f'Location captured: {lat}, {lng}', 'success')
    else:
        tag.has_gps = False
        db.session.commit()
        log_action(pid, 'Google Maps link saved but coordinates could not be parsed')
        flash('Link saved, but coordinates could not be read from it. Try sharing the location again '
              '(long-press the pin in Google Maps → Share → Copy link), or set a pin manually below.', 'warning')

    return redirect(url_for('onsite_progress', pid=pid))

@app.route('/geo_photo/<int:pid>')
@login_required
def serve_geo_photo(pid):
    tag = ProjectGeoTag.query.filter_by(project_id=pid).first_or_404()
    if not tag.photo_path or not os.path.isfile(tag.photo_path):
        abort(404)
    return send_file(tag.photo_path)

@app.route('/geo_photo/<int:pid>/download')
@login_required
def download_geo_photo(pid):
    tag = ProjectGeoTag.query.filter_by(project_id=pid).first_or_404()
    if not tag.photo_path or not os.path.isfile(tag.photo_path):
        abort(404)
    proj = tag.project
    ext  = os.path.splitext(tag.photo_path)[1]
    return send_file(tag.photo_path, as_attachment=True,
                      download_name=f'{proj.project_code}_site_photo{ext}')
@app.route('/projects/<int:pid>/geo_tag/manual', methods=['POST'])
@login_required
@roles_required('admin', 'onsite')
def set_geo_tag_manual(pid):
    tag = ProjectGeoTag.query.filter_by(project_id=pid).first()
    if not tag:
        tag = ProjectGeoTag(project_id=pid)
        db.session.add(tag)
    tag.latitude  = _safe_float(request.form.get('latitude'))
    tag.longitude = _safe_float(request.form.get('longitude'))
    tag.has_gps   = True
    tag.uploaded_by = current_user.id
    db.session.commit()
    flash('Location pin updated.', 'success')
    return redirect(url_for('onsite_progress', pid=pid))

from math import radians, cos, sin, asin, sqrt

def haversine_km(lat1, lng1, lat2, lng2):
    lat1, lng1, lat2, lng2 = map(radians, [lat1, lng1, lat2, lng2])
    a = sin((lat2-lat1)/2)**2 + cos(lat1)*cos(lat2)*sin((lng2-lng1)/2)**2
    return 2 * 6371 * asin(sqrt(a))

import requests

def geocode_place(query):
    try:
        r = requests.get('https://nominatim.openstreetmap.org/search',
            params={'q': query, 'format': 'json', 'limit': 1, 'countrycodes': 'in'},
            headers={'User-Agent': 'PowerOnPlusSolar/1.0 (contact@yourcompany.in)'}, timeout=5)
        d = r.json()
        return (float(d[0]['lat']), float(d[0]['lon'])) if d else (None, None)
    except Exception:
        return None, None
@app.route('/geo_photo/<int:pid>/site')
@login_required
def serve_site_photo(pid):
    tag = ProjectGeoTag.query.filter_by(project_id=pid).first_or_404()
    if not tag.site_photo_path or not os.path.isfile(tag.site_photo_path):
        abort(404)
    return send_file(tag.site_photo_path)


@app.route('/geo_photo/<int:pid>/site/download')
@login_required
def download_site_photo(pid):
    tag = ProjectGeoTag.query.filter_by(project_id=pid).first_or_404()
    if not tag.site_photo_path or not os.path.isfile(tag.site_photo_path):
        abort(404)
    proj = tag.project
    ext  = os.path.splitext(tag.site_photo_path)[1]
    return send_file(tag.site_photo_path, as_attachment=True,
                      download_name=f'{proj.project_code}_before_work_site_photo{ext}')


@app.route('/projects/<int:pid>/geo_tag/delete_site_photo', methods=['POST'])
@login_required
@roles_required('admin', 'onsite', 'coordinator', 'documents', 'office', 'documents_k', 'director')
def delete_site_photo(pid):
    tag = ProjectGeoTag.query.filter_by(project_id=pid).first()
    if tag and tag.site_photo_path:
        if os.path.isfile(tag.site_photo_path):
            try:
                os.remove(tag.site_photo_path)
            except OSError:
                pass
        tag.site_photo_path = None
        log_action(pid, 'Pre-work site photo removed')
        db.session.commit()
        flash('Site photo removed.', 'success')
    else:
        flash('No site photo to remove.', 'warning')
    return redirect(request.referrer or url_for('project_detail', pid=pid))
@app.route('/service_map')
@login_required
@roles_required('admin', 'onsite', 'coordinator', 'director','service')
def service_map():
    q         = _clean(request.args.get('q', ''), 100)
    radius_km = request.args.get('radius', 15, type=int)
    urgency   = request.args.get('urgency', 'all')
    if urgency not in ('all', 'overdue', 'due'):
        urgency = 'all'

    tags = (ProjectGeoTag.query.join(Project)
        .options(joinedload(ProjectGeoTag.project).joinedload(Project.payments),
                 joinedload(ProjectGeoTag.project).joinedload(Project.subsidy),
                 joinedload(ProjectGeoTag.project).joinedload(Project.service_records))
        .filter(ProjectGeoTag.latitude.isnot(None),
                Project.status != 'Cancelled',
                Project.work_category != 'Outside')            
        .all())
    tags = [t for t in tags if t.project.pending_amount <= 0 or t.project.status == 'Closed']

    def _next_visit_status(proj):
        recs = sorted(proj.service_records, key=lambda r: r.visit_number)
        nxt  = next((r for r in recs if r.status not in ('Completed', 'Skipped')), None)
        return nxt.status if nxt else None

    tag_urgency = {t.id: _next_visit_status(t.project) for t in tags}

    if urgency == 'overdue':
        tags = [t for t in tags if tag_urgency.get(t.id) == 'Overdue']
    elif urgency == 'due':
        tags = [t for t in tags if tag_urgency.get(t.id) in ('Overdue', 'Due')]

    tags_json = [{
        'lat':     t.latitude,
        'lng':     t.longitude,
        'code':    t.project.project_code,
        'name':    t.project.customer.name,
        'url':     url_for('project_detail', pid=t.project_id),
        'urgency': tag_urgency.get(t.id) or 'None',
    } for t in tags]

    center = (None, None)
    nearby = []
    if q:
        center = geocode_place(q + ', Kerala, India')
        if center[0]:
            for t in tags:
                dist = haversine_km(center[0], center[1], t.latitude, t.longitude)
                if dist <= radius_km:
                    nearby.append((t, round(dist, 1)))
            nearby.sort(key=lambda x: x[1])
        else:
            flash(f'Could not locate "{q}".', 'warning')

    return render_template('service_map.html', tags=tags, nearby=nearby,
        tags_json=tags_json, search_q=q, center=center, radius_km=radius_km,
        urgency=urgency, tag_urgency=tag_urgency)
def create_service_schedule(project):
    if project.work_category == 'Outside':          
        return
    if ServiceRecord.query.filter_by(project_id=project.id).first():
        return

    base = date.today()
    for visit_num in range(1, 11):
        months_ahead = visit_num * 6
        year_offset, month_offset = divmod(base.month - 1 + months_ahead, 12)
        sched = base.replace(year=base.year + year_offset, month=month_offset + 1)
        db.session.add(ServiceRecord(
            project_id     = project.id,
            visit_number   = visit_num,
            scheduled_date = sched,
            status         = 'Upcoming',
        ))

    # notify_onsite_team(project.id,
    #     f'Service schedule created for {project.project_code} — {project.customer.name}. '
    #     f'10 panel-cleaning visits over 5 years starting {base.strftime("%d %b %Y")}.', 'info')
    # log_action(project.id, 'Service schedule created (10 visits x 6 months)', new_val='Upcoming')

def _add_months(base_date, months):
    """Add N months to a date, safely handling month-end overflow (e.g. Jan 31 + 1mo)."""
    year_offset, month_offset = divmod(base_date.month - 1 + months, 12)
    year  = base_date.year + year_offset
    month = month_offset + 1
    last_day = calendar.monthrange(year, month)[1]
    day = min(base_date.day, last_day)
    return base_date.replace(year=year, month=month, day=day)
def _annotate_service_records(records):
    """Attach a `locked` flag and a display `state` to each ServiceRecord,
    based on whether the previous visit is completed. Templates should use
    this instead of recomputing the prev-visit chain themselves — keeping
    that logic in one place avoids the pill/row/body versions drifting."""
    annotated = []
    prev_done = True
    for rec in records:
        locked = not prev_done and rec.status not in ('Completed', 'Skipped')
        if locked:
            state = 'locked'
        elif rec.status == 'Completed':
            state = 'done'
        elif rec.status == 'Overdue':
            state = 'overdue'
        elif rec.status == 'Due':
            state = 'due'
        elif rec.status == 'Skipped':
            state = 'skipped'
        else:
            state = 'upcoming'
        annotated.append({'rec': rec, 'locked': locked, 'state': state})
        prev_done = rec.status == 'Completed'
    return annotated

def _reschedule_future_visits(rec, from_date):
    """After `rec` is completed on `from_date`, push every later, not-yet-completed/
    skipped visit for the same project to a fresh 6-month cadence starting from
    `from_date`, and recompute their status against today's date."""
    today    = date.today()
    due_soon = today + timedelta(days=30)

    future = (ServiceRecord.query
              .filter(ServiceRecord.project_id == rec.project_id,
                      ServiceRecord.visit_number > rec.visit_number,
                      ServiceRecord.status.notin_(['Completed', 'Skipped']))
              .order_by(ServiceRecord.visit_number)
              .all())

    changes = []
    for i, fut in enumerate(future, start=1):
        old_date = fut.scheduled_date
        new_date = _add_months(from_date, i * 6)
        if new_date == old_date:
            continue
        fut.scheduled_date = new_date
        if new_date < today:
            fut.status = 'Overdue'
        elif new_date <= due_soon:
            fut.status = 'Due'
        else:
            fut.status = 'Upcoming'
        changes.append(f'#{fut.visit_number}: {old_date} → {new_date}')

    if changes:
        log_action(rec.project_id,
            f'Service schedule shifted after visit #{rec.visit_number} completion: '
            + '; '.join(changes))
def refresh_service_statuses():
   
    today    = date.today()
    due_soon = today + timedelta(days=30)
 
    pending = ServiceRecord.query.filter(
        ServiceRecord.status.in_(['Upcoming', 'Due'])
    ).all()
 
    changed = False
    for rec in pending:
        if rec.scheduled_date < today:
            if rec.status != 'Overdue':
                rec.status = 'Overdue'
                changed = True
        elif rec.scheduled_date <= due_soon:
            if rec.status != 'Due':
                rec.status = 'Due'
                changed = True
 
    if changed:
        db.session.commit()

def auto_advance_stage(proj):
    if proj.status in ('Cancelled', 'OnHold', 'Completed', 'Closed'):
        return
    if proj.work_category == 'Outside':
        _auto_advance_outside_stage(proj)
        return

    db.session.expire(proj, ['documents', 'site_visits', 'onsite_progress',
                              'app_install', 'subsidy', 'assignments'])

    old_stage  = proj.stage
    old_status = proj.status
    doc_map    = {d.doc_type: d for d in proj.documents}

    def doc_done(*names):
        return all(
            doc_map.get(n) and doc_map[n].status in ('Received', 'Sent', 'Completed')
            for n in names
        )

    if proj.stage in ('Lead', 'Site Visit'):
        proj.stage  = 'Documentation'
        proj.status = 'InProgress'

    elif proj.stage == 'Documentation':
        if doc_done('Feasibility Receipt'):
            proj.stage  = 'Onsite Work'
            proj.status = 'InProgress'

    elif proj.stage == 'Onsite Work':
        op = proj.onsite_progress
        if op and op.electrical_status == 'Completed':
            proj.stage  = 'Connection'
            proj.status = 'InProgress'

    elif proj.stage == 'Connection':
        if doc_done('KSEB Connection'):
            proj.stage  = 'Payment'
            proj.status = 'InProgress'

    elif proj.stage == 'Payment':
    # For loan projects, only count up to contract amount — excess is tracked separately
        if proj.project_type == 'Loan':
            bank_total = sum(float(p.amount) for p in proj.payments if p.payment_source == 'Bank')
            customer_total = sum(float(p.amount) for p in proj.payments if p.payment_source == 'Customer')
            effective = min(bank_total, float(proj.total_receivable)) + customer_total
            fully_paid = proj.total_receivable > 0 and (effective + proj.total_waived) >= float(proj.total_receivable)
        else:
            fully_paid = proj.total_receivable > 0 and proj.pending_amount <= 0
        if fully_paid:
            if proj.project_subtype == 'DCR':
                proj.stage  = 'Subsidy'
                proj.status = 'InProgress'
            else:
                proj.stage  = 'Subsidy'  # non-DCR still goes here for warranty/app
                proj.status = 'InProgress'
                create_service_schedule(proj)

    elif proj.stage == 'Subsidy':
        fully_paid    = proj.total_receivable > 0 and proj.pending_amount <= 0
        warranty_done = doc_done('Warranty Card')
        app_done      = proj.app_install and proj.app_install.status == 'Completed'

        if proj.project_subtype == 'DCR':
            sub = proj.subsidy
            subsidy_done = sub and sub.status == 'Received'
        else:
            subsidy_done = True  # non-DCR skips subsidy requirement

        if fully_paid and warranty_done and app_done and subsidy_done:
            proj.stage  = 'Subsidy'   # stage stays Subsidy
            proj.status = 'Completed'

    if proj.status in ('Completed', 'Closed'):
        create_service_schedule(proj)

    if proj.stage != old_stage or proj.status != old_status:
        proj.staged_changed_at = datetime.utcnow()
        log_action(proj.id, f'Auto-advanced: {old_stage} → {proj.stage}',
                   old_val=old_status, new_val=proj.status)
        _notify_stage_transition(proj, old_stage, proj.stage)

def _auto_advance_outside_stage(proj):
    db.session.expire(proj, ['documents'])
    old_stage, old_status = proj.stage, proj.status
    doc_map = {d.doc_type: d for d in proj.documents}

    def doc_done(*names):
        return all(doc_map.get(n) and doc_map[n].status in ('Received','Sent','Completed') for n in names)

    if proj.stage in ('Lead', 'Site Visit'):
        proj.stage, proj.status = 'Documentation', 'InProgress'

    elif proj.stage == 'Documentation':
        
        if doc_done('Outside Work Completed'):
            proj.stage, proj.status = 'Payment', 'InProgress'

    elif proj.stage == 'Payment':
        pass  
              

    if proj.stage != old_stage or proj.status != old_status:
        proj.staged_changed_at = datetime.utcnow()
        log_action(proj.id, f'Auto-advanced (Outside): {old_stage} → {proj.stage}',
                   old_val=old_status, new_val=proj.status)
def _notify_stage_transition(proj, from_stage, to_stage):
    code = f'{proj.project_code} — {proj.customer.name}'

    if to_stage == 'Documentation':
        if proj.doc_staff_id:
            create_notification(proj.doc_staff_id, proj.id,
                f'{code}: Documentation stage started.', 'task')

    elif to_stage == 'Onsite Work':
        if proj.coordinator_id:
            create_notification(proj.coordinator_id, proj.id,
                f'{code} moved to Onsite Work stage.', 'info')

    elif to_stage == 'Connection':
        if proj.coordinator_id:
            create_notification(proj.coordinator_id, proj.id,
                f'{code}: Onsite work complete. KSEB connection step now active.', 'info')
        # if proj.doc_staff_id:
        #     create_notification(proj.doc_staff_id, proj.id,
        #         f'{code}: Electrical work done. Please update KSEB Connection document.', 'task')

    elif to_stage == 'Payment':
        if proj.coordinator_id:
            create_notification(proj.coordinator_id, proj.id,
                f'{code}: KSEB connection done. Ready for final payment collection.', 'info')
        for u in User.query.filter_by(role='payments', is_active=True).all():
            create_notification(u.id, proj.id,
                f'{code}: Entered Payment stage. '
                f'Collect remaining balance of ₹{proj.pending_amount:,.0f}.', 'task')
        if proj.doc_staff_id:
            create_notification(proj.doc_staff_id, proj.id,
                f'{code}: Project in payment stage. Please ensure Warranty Card '
                f'and App Installation documents are updated.', 'task')

    elif to_stage == 'Subsidy':
        if proj.project_subtype == 'DCR':
            for u in User.query.filter_by(role='payments', is_active=True).all():
                create_notification(u.id, proj.id,
                    f'{code}: Payment collected. Please redeem and receive the subsidy.', 'task')
        if proj.coordinator_id:
            create_notification(proj.coordinator_id, proj.id,
                f'{code}: Payment complete. Pending: warranty card, app install'
                + (', subsidy.' if proj.project_subtype == 'DCR' else '.'), 'info')

    # Project completed
    if proj.status == 'Completed' and from_stage != proj.stage:
        pass  # handled below

    if proj.status == 'Completed':
        if proj.coordinator_id:
            create_notification(proj.coordinator_id, proj.id,
                f'{code}: Project completed. All requirements met.', 'info')
        if proj.doc_staff_id:
            create_notification(proj.doc_staff_id, proj.id,
                f'{code}: Project marked completed.', 'info')

def create_notification(user_id, project_id, message, notif_type='info'):
    db.session.add(Notification(
        user_id=user_id, project_id=project_id,
        message=message[:255], notif_type=notif_type,
    ))
    send_push_notification(
        user_id, 'Power On Plus',
        message[:150],
        url=f'/projects/{project_id}' if project_id else '/dashboard',
    )
def send_push_notification(user_id, title, body, url='/dashboard', tag=None):
    subs = PushSubscription.query.filter_by(user_id=user_id).all()
    if not subs:
        print(f'[PUSH] No subscriptions for user {user_id}')
        return
    payload = _json.dumps({'title': title, 'body': body, 'url': url, 'tag': tag})
    vapid_private_key = os.environ.get('VAPID_PRIVATE_KEY') or os.environ.get('VAPID_PRIVATE_KEY_PATH')
    vapid_claims = {'sub': os.environ.get('VAPID_CLAIM_EMAIL', 'mailto:admin@example.com')}
    if not vapid_private_key:
        print('[PUSH] VAPID_PRIVATE_KEY not set — aborting')
        return
    for sub in subs:
        try:
            webpush(
                subscription_info={
                    'endpoint': sub.endpoint,
                    'keys': {'p256dh': sub.p256dh, 'auth': sub.auth},
                },
                data=payload,
                vapid_private_key=vapid_private_key,
                vapid_claims=vapid_claims,
            )
            print(f'[PUSH] Sent OK to sub {sub.id}')
        except WebPushException as ex:
            print(f'[PUSH] WebPushException: {ex}')
            if ex.response is not None:
                print(f'[PUSH] status={ex.response.status_code} body={ex.response.text}')
            if ex.response is not None and ex.response.status_code in (404, 410):
                db.session.delete(sub)
        except Exception as ex:
            print(f'[PUSH] Unexpected error: {repr(ex)}')
@app.route('/api/notifications/unread_count')
@login_required
def api_unread_count():
    count = Notification.query.filter_by(user_id=current_user.id, is_read=False).count()
    return jsonify({'count': count})
@app.route('/sw.js')
def service_worker():
    resp = make_response(send_from_directory(app.static_folder, 'sw.js'))
    resp.headers['Content-Type'] = 'application/javascript'
    resp.headers['Service-Worker-Allowed'] = '/'
    return resp


@app.route('/api/push/vapid_public_key')
@login_required
def push_public_key():
    return jsonify({'publicKey': os.environ.get('VAPID_PUBLIC_KEY', '')})


@app.route('/api/push/subscribe', methods=['POST'])
@login_required
@csrf.exempt
def push_subscribe():
    data = request.get_json(force=True, silent=True) or {}
    endpoint = data.get('endpoint')
    keys = data.get('keys', {})
    if not endpoint or not keys.get('p256dh') or not keys.get('auth'):
        return jsonify({'error': 'invalid subscription'}), 400

    sub = PushSubscription.query.filter_by(endpoint=endpoint).first()
    if sub:
        sub.user_id    = current_user.id
        sub.p256dh     = keys['p256dh']
        sub.auth       = keys['auth']
        sub.user_agent = request.headers.get('User-Agent', '')[:255]
    else:
        db.session.add(PushSubscription(
            user_id=current_user.id, endpoint=endpoint,
            p256dh=keys['p256dh'], auth=keys['auth'],
            user_agent=request.headers.get('User-Agent', '')[:255],
        ))
    db.session.commit()
    return jsonify({'ok': True})


@app.route('/api/push/unsubscribe', methods=['POST'])
@login_required
@csrf.exempt
def push_unsubscribe():
    data = request.get_json(force=True, silent=True) or {}
    endpoint = data.get('endpoint')
    if endpoint:
        PushSubscription.query.filter_by(endpoint=endpoint, user_id=current_user.id).delete()
        db.session.commit()
    return jsonify({'ok': True})
def _reconcile_excess_after_payment_change(proj):
    """Recomputes PaymentExcess (customer-side) and BankExcessReturn (bank-side)
    after a payment edit/delete. Only touches Pending/unsettled excess records —
    already-settled excess is left alone (it's historical, money already moved)."""

    receivable = float(proj.total_receivable)

    # ── Customer-side excess (PaymentExcess) ──────────────────────────────
    customer_total = sum(float(p.amount) for p in proj.payments if p.payment_source == 'Customer')
    bank_total     = sum(float(p.amount) for p in proj.payments if p.payment_source == 'Bank')

    if proj.project_type == 'Loan':
        # collected_amount = capped bank + all customer payments
        effective_collected = min(bank_total, receivable) + customer_total
    else:
        effective_collected = customer_total + bank_total

    proj.collected_amount = min(effective_collected, receivable) if receivable > 0 else effective_collected

    pending_pay_excess = [e for e in proj.payment_excesses if e.status == 'Pending']
    correct_pay_excess = max(0, round(effective_collected - receivable, 2)) if receivable > 0 else 0

    if proj.project_type == 'Loan':
        # For loan projects, overshoot from Bank is tracked separately via BankExcessReturn,
        # so PaymentExcess should only reflect Customer-side overshoot beyond what bank already covers.
        bank_capped   = min(bank_total, receivable)
        correct_pay_excess = max(0, round((bank_capped + customer_total) - receivable, 2))

    if pending_pay_excess:
        total_pending = sum(float(e.amount) for e in pending_pay_excess)
        if correct_pay_excess <= 0.01:
            # No excess remains — remove stale pending records entirely
            for e in pending_pay_excess:
                log_action(proj.id, f'Payment excess of ₹{float(e.amount):,.0f} removed '
                           '(reconciled after payment edit/delete)', old_val='Pending', new_val='Removed')
                db.session.delete(e)
        elif abs(correct_pay_excess - total_pending) > 0.01:
            # Adjust the most recent pending record to match; drop any extras
            pending_pay_excess.sort(key=lambda e: e.created_at, reverse=True)
            keep = pending_pay_excess[0]
            old_amt = float(keep.amount)
            keep.amount = correct_pay_excess
            keep.notes = (keep.notes or '') + f' [Adjusted from ₹{old_amt:,.0f} to ₹{correct_pay_excess:,.0f} after payment edit/delete]'
            for extra in pending_pay_excess[1:]:
                db.session.delete(extra)
            log_action(proj.id, f'Payment excess reconciled: ₹{old_amt:,.0f} → ₹{correct_pay_excess:,.0f}',
                       old_val=str(old_amt), new_val=str(correct_pay_excess))

    # ── Bank-side excess (BankExcessReturn) — Loan projects only ──────────
    if proj.project_type == 'Loan':
        contract_amt = float(proj.total_amount or 0)
        correct_bank_excess = max(0, round(bank_total - contract_amt, 2))
        exc = proj.bank_excess

        if exc and exc.returned:
            pass  # already settled — leave historical record untouched
        elif exc and not exc.returned:
            if correct_bank_excess <= 0.01:
                log_action(proj.id, f'Bank excess of ₹{float(exc.excess_amount):,.0f} removed '
                           '(reconciled after payment edit/delete)', old_val='Pending', new_val='Removed')
                db.session.delete(exc)
            elif abs(correct_bank_excess - float(exc.excess_amount)) > 0.01:
                old_amt = float(exc.excess_amount)
                exc.excess_amount = correct_bank_excess
                exc.notes = (exc.notes or '') + f' [Adjusted from ₹{old_amt:,.0f} to ₹{correct_bank_excess:,.0f} after payment edit/delete]'
                log_action(proj.id, f'Bank excess reconciled: ₹{old_amt:,.0f} → ₹{correct_bank_excess:,.0f}',
                           old_val=str(old_amt), new_val=str(correct_bank_excess))
        elif not exc and correct_bank_excess > 0.01:
            # A new excess emerged (e.g. an edit increased the bank payment)
            db.session.add(BankExcessReturn(
                project_id=proj.id, excess_amount=correct_bank_excess,
                notes='Auto-detected on payment edit/delete.',
                recorded_by=current_user.id,
            ))
            for u in User.query.filter_by(role='payments', is_active=True).all():
                create_notification(u.id, proj.id,
                    f'{proj.project_code} — {proj.customer.name}: Bank excess of ₹{correct_bank_excess:,.0f} '
                    f'detected after a payment was edited. Needs settlement.', 'task')
            log_action(proj.id, f'Bank excess detected: ₹{correct_bank_excess:,.0f}', new_val='Pending')
def record_stock_txn(stock_item_id, txn_type, quantity, source='Manual',
                      project_id=None, reference_id=None, notes=None, sale_type=None):
    item = StockItem.query.get(stock_item_id)
    if not item:
        return None
    delta = quantity if txn_type == 'In' else -quantity
    new_balance = float(item.current_qty or 0) + float(delta)
    db.session.add(StockTransaction(
        stock_item_id=stock_item_id, txn_type=txn_type, quantity=quantity,
        source=source, sale_type=sale_type, project_id=project_id,
        reference_id=reference_id, notes=notes, balance_after=new_balance,
        recorded_by=current_user.id,
    ))
    item.current_qty = new_balance
    
 
    if item.is_low:
        alert_users = User.query.filter(
            User.role.in_(['admin', 'stocks']), User.is_active == True
        ).all()
        for u in alert_users:
            create_notification(u.id, project_id,
                f'Low stock: {item.display_name} — {item.current_qty} {item.unit} left '
                f'(reorder level {item.reorder_level} {item.unit}).', 'warning')
    return item

def notify_onsite_team(project_id, message, notif_type='task'):
    for user in User.query.filter_by(role='onsite', is_active=True).all():
        create_notification(user.id, project_id, message, notif_type)

@app.route('/onsite_activity')
@login_required
@roles_required('admin', 'payments', 'director')
def onsite_activity():
    """Read-only day-by-day view of onsite work — for the Payments team.
    Shows phase start/completion events (structure/installation/electrical)
    and logged notes, per project, for the selected date."""
    selected_date = request.args.get('date')
    try:
        day = date.fromisoformat(selected_date) if selected_date else date.today()
    except ValueError:
        day = date.today()

    # ── Logged notes for the day ────────────────────────────────────────
    logs_today = (OnsiteLog.query
        .options(joinedload(OnsiteLog.project).joinedload(Project.customer))
        .filter(OnsiteLog.log_date == day)
        .order_by(OnsiteLog.created_at.desc())
        .all())

    # ── Phase start/completion events for the day ───────────────────────
    # Comes from OnsiteProgress's own start/end date fields — set when
    # structure/installation/electrical work is marked started or completed.
    phase_progress = (OnsiteProgress.query
        .options(joinedload(OnsiteProgress.project).joinedload(Project.customer))
        .filter(db.or_(
            OnsiteProgress.structure_start_date == day,
            OnsiteProgress.structure_end_date == day,
            OnsiteProgress.installation_start_date == day,
            OnsiteProgress.installation_end_date == day,
            OnsiteProgress.electrical_start_date == day,
            OnsiteProgress.electrical_end_date == day,
        ))
        .all())

    by_project = {}

    def _row(p):
        return by_project.setdefault(p.id, {'project': p, 'phase_events': [], 'logs': []})

    for op in phase_progress:
        events = []
        if op.structure_end_date == day and op.structure_work_status == 'Completed':
            events.append({'phase': 'Structure', 'event': 'Completed'})
        elif op.structure_start_date == day:
            events.append({'phase': 'Structure', 'event': 'Started'})
        if op.installation_end_date == day and op.installation_status == 'Completed':
            events.append({'phase': 'Installation', 'event': 'Completed'})
        elif op.installation_start_date == day:
            events.append({'phase': 'Installation', 'event': 'Started'})
        if op.electrical_end_date == day and op.electrical_status == 'Completed':
            events.append({'phase': 'Electrical', 'event': 'Completed'})
        elif op.electrical_start_date == day:
            events.append({'phase': 'Electrical', 'event': 'Started'})
        if events:
            _row(op.project)['phase_events'].extend(events)

    for l in logs_today:
        _row(l.project)['logs'].append(l)

    rows = sorted(by_project.values(), key=lambda r: r['project'].project_code)

    # ── Nearest prev/next dates with any activity (log or phase event) ──
    all_dates = set()
    for (d,) in db.session.query(OnsiteLog.log_date).distinct():
        if d: all_dates.add(d)
    for col in (OnsiteProgress.structure_start_date, OnsiteProgress.structure_end_date,
                OnsiteProgress.installation_start_date, OnsiteProgress.installation_end_date,
                OnsiteProgress.electrical_start_date, OnsiteProgress.electrical_end_date):
        for (d,) in db.session.query(col).filter(col.isnot(None)).distinct():
            if d: all_dates.add(d)

    earlier = sorted([d for d in all_dates if d < day], reverse=True)
    later   = sorted([d for d in all_dates if d > day])
    prev_activity_date = earlier[0] if earlier else None
    next_activity_date = later[0] if later else None
    recent_activity_dates = sorted(all_dates, reverse=True)[:10]

    return render_template('onsite_activity.html',
        rows=rows, day=day,
        prev_day=prev_activity_date or (day - timedelta(days=1)),
        next_day=next_activity_date or (day + timedelta(days=1)),
        has_prev=prev_activity_date is not None,
        has_next=next_activity_date is not None,
        recent_activity_dates=recent_activity_dates,
        today=date.today())
def next_project_code():
    numeric = []
    for (code,) in db.session.query(Project.project_code).all():
        try:
            numeric.append(int(code))
        except (ValueError, TypeError):
            pass
    
    base = max(max(numeric), 1780) if numeric else 1780  # ← floor at 1780
    existing = set(numeric)
    candidate = base + 1
    while candidate in existing:
        candidate += 1
    return str(candidate)


# ─────────────────────────────────────────────────────────────────────────────
# AUTH ROUTES
# ─────────────────────────────────────────────────────────────────────────────

@app.route('/')
def index():
    if current_user.is_authenticated:
        return redirect(url_for('dashboard'))
    return redirect(url_for('login'))


@app.route('/login', methods=['GET', 'POST'])
@limiter.limit('20 per minute; 5 per 10 seconds')   # ← brute-force throttle
def login():
    csrf_token = generate_csrf()
    if request.method == 'POST':
        username = _clean(request.form.get('username', ''), 80)
        password = request.form.get('password', '')

        # Generic error message — never reveal whether username exists
        _fail_msg = 'Invalid credentials.'

        u = User.query.filter_by(username=username).first()

        if not u:
            # log_login_attempt(username, False, 'unknown_user')
            flash(_fail_msg, 'danger')
            return redirect(url_for('login'))

        # Account locked?
        if u.is_locked():
            remaining = int((u.locked_until - datetime.utcnow()).total_seconds() / 60) + 1
            # log_lockout(username, remaining)
            flash(f'Account temporarily locked. Try again in {remaining} minute(s).', 'danger')
            return redirect(url_for('login'))

        # Wrong password
        if not u.check_password(password):
            u.record_failed_login()
            db.session.commit()
            # log_login_attempt(username, False, 'bad_password')
            if u.is_locked():
                # log_lockout(username, LOCKOUT_MINUTES)
                flash(f'Too many failed attempts. Account locked for {LOCKOUT_MINUTES} minutes.', 'danger')
            else:
                remaining_attempts = MAX_LOGIN_ATTEMPTS - u.failed_logins
                flash(f'{_fail_msg} {remaining_attempts} attempt(s) remaining.', 'danger')
            return redirect(url_for('login'))

        # Inactive account
        if u.status != 'active':
            # log_login_attempt(username, False, 'inactive')
            flash('Your account is not active. Contact admin.', 'danger')
            return redirect(url_for('login'))
        if u.is_deleted:
            # log_login_attempt(username, False, 'deleted')
            flash('This account no longer exists. Contact admin.', 'danger')
            return redirect(url_for('login'))

        # if u.status != 'active':
        #     flash('Your account is not active. Contact admin.', 'danger')
        #     return redirect(url_for('login'))
        # Success
        u.reset_login_attempts()
        db.session.commit()
                              
        login_user(u)
        session.permanent = True 
        # Regenerate session to prevent session fixation
        session.regenerate() if hasattr(session, 'regenerate') else None
        if u.role == 'stocks':
            return redirect(url_for('stock_dashboard'))
        return redirect(url_for('dashboard'))

    return render_template('login.html',csrf_token=csrf_token)


@app.route('/logout')
@login_required
def logout():
    logout_user()
    session.clear()
    flash('You have been logged out.', 'info')
    return redirect(url_for('login'))


# ── Change own password ───────────────────────────────────────────────────────
@app.route('/change_password', methods=['GET', 'POST'])
@login_required
@limiter.limit('10 per minute')
def change_password():
    if request.method == 'POST':
        current_pw = request.form.get('current_password', '')
        new_pw     = request.form.get('new_password', '')
        confirm_pw = request.form.get('confirm_password', '')
 
        if not current_user.check_password(current_pw):
            flash('Current password is incorrect.', 'danger')
            return redirect(url_for('change_password'))
 
        if new_pw != confirm_pw:
            flash('New passwords do not match.', 'danger')
            return redirect(url_for('change_password'))
 
        errors = _validate_password(new_pw)
        if errors:
            flash(f'Password must contain: {", ".join(errors)}.', 'danger')
            return redirect(url_for('change_password'))
 
        current_user.set_password(new_pw)
        current_user.reset_login_attempts()   # clear any lingering lockout
        db.session.flush()
 
        # Notify all admins that this user changed their own password
        admins = User.query.filter_by(role='admin', is_active=True).all()
        for admin in admins:
            if admin.id != current_user.id:   # don't notify yourself if you're admin
                db.session.add(Notification(
                    user_id=admin.id,
                    project_id=None,           # no project context
                    message=(
                        f'{current_user.full_name} ({current_user.username}) '
                        f'changed their own password.'
                    ),
                    notif_type='info',
                ))
        
        db.session.commit()
        # log_password_change(current_user.username, current_user.username)
        flash('Password changed successfully.', 'success')
        return redirect(url_for('dashboard'))
 
    return render_template('change_password.html')
@app.route('/admin/change_password', methods=['GET', 'POST'])
@login_required
@roles_required('admin')
@limiter.limit('20 per minute')
def admin_change_password():
    users = User.query.filter_by(is_active=True).order_by(User.role, User.full_name).all()
    selected_user = None
 
    if request.method == 'POST':
        target_id  = request.form.get('target_user_id', type=int)
        new_pw     = request.form.get('new_password', '')
        confirm_pw = request.form.get('confirm_password', '')
 
        if not target_id:
            flash('Please select a user.', 'danger')
            return render_template('admin_change_password.html', users=users, selected_user=None)
 
        target = User.query.get_or_404(target_id)
        selected_user = target
 
        if new_pw != confirm_pw:
            flash('Passwords do not match.', 'danger')
            return render_template('admin_change_password.html', users=users, selected_user=selected_user)
 
        errors = _validate_password(new_pw)
        if errors:
            flash(f'Password must contain: {", ".join(errors)}.', 'danger')
            return render_template('admin_change_password.html', users=users, selected_user=selected_user)
 
        target.set_password(new_pw)
        target.reset_login_attempts()   # clear any lockout on the target account
        db.session.flush()
        target_project = None
        if target.role == 'coordinator':
            target_project = (Project.query
                              .filter_by(coordinator_id=target.id)
                              .order_by(Project.updated_at.desc()).first())
        elif target.role == 'documents':
            target_project = (Project.query
                              .filter_by(doc_staff_id=target.id)
                              .order_by(Project.updated_at.desc()).first())
 
        if target_project:
            db.session.add(Notification(
                user_id=target.id,
                project_id=target_project.id,
                message=(
                    f'Your password was reset by admin ({current_user.full_name}). '
                    f'Please change it after logging in.'
                ),
                notif_type='warning',
            ))
        db.session.commit()
        # log_password_change(target.username, current_user.username)
        flash(
            f'Password for {target.full_name} ({target.username}) has been reset successfully. '
            f'They have been notified.',
            'success'
        )
        return redirect(url_for('manage_users'))
 
    # GET — allow pre-selecting a user via ?user_id=
    preselect_id = request.args.get('user_id', type=int)
    if preselect_id:
        selected_user = User.query.get(preselect_id)
 
    return render_template('admin_change_password.html', users=users, selected_user=selected_user)


# ─────────────────────────────────────────────────────────────────────────────
# DASHBOARD
# ─────────────────────────────────────────────────────────────────────────────

@app.route('/admin/document_stages')
@login_required
@roles_required('admin')
def manage_document_stages():
    stages = DocumentStage.query.order_by(DocumentStage.sort_order).all()
    return render_template('admin_document_stages.html', stages=stages)


@app.route('/admin/document_stages/new', methods=['POST'])
@login_required
@roles_required('admin')
def new_document_stage():
    last  = DocumentStage.query.order_by(DocumentStage.sort_order.desc()).first()
    stage = DocumentStage(
        name       = _clean(request.form['name'], 100),
        condition  = request.form.get('condition', 'always'),
        docs       = _clean(request.form['docs'], 1000),
        sort_order = (last.sort_order + 1) if last else 0,
        is_active  = True,
    )
    db.session.add(stage)
    db.session.commit()
    flash(f'Stage "{stage.name}" created.', 'success')
    return redirect(url_for('manage_document_stages'))


@app.route('/admin/document_stages/<int:sid>/edit', methods=['POST'])
@login_required
@roles_required('admin')
def edit_document_stage(sid):
    stage           = DocumentStage.query.get_or_404(sid)
    stage.name      = _clean(request.form['name'], 100)
    stage.condition = request.form.get('condition', stage.condition)
    stage.docs      = _clean(request.form['docs'], 1000)
    stage.is_active = 'is_active' in request.form
    db.session.commit()
    flash(f'Stage "{stage.name}" updated.', 'success')
    return redirect(url_for('manage_document_stages'))


@app.route('/admin/document_stages/<int:sid>/delete', methods=['POST'])
@login_required
@roles_required('admin')
def delete_document_stage(sid):
    stage           = DocumentStage.query.get_or_404(sid)
    stage.is_active = False
    db.session.commit()
    flash(f'Stage "{stage.name}" deactivated.', 'warning')
    return redirect(url_for('manage_document_stages'))


@app.route('/admin/document_stages/reorder', methods=['POST'])
@login_required
@roles_required('admin')
def reorder_document_stages():
    for i, sid in enumerate(request.form.getlist('order')):
        stage = DocumentStage.query.get(int(sid))
        if stage:
            stage.sort_order = i
    db.session.commit()
    return jsonify({'ok': True})


@app.route('/dashboard')
@login_required
def dashboard():
    if current_user.role == 'stocks':
        return redirect(url_for('stock_dashboard'))
    cutoff = datetime.utcnow() - timedelta(days=180)
    update_rows  = Project.query.filter(
        Project.status == 'InProgress',
        Project.created_at <= cutoff,
    ).update({'status':'Delayed'},synchronize_session=False)
    if update_rows:
        db.session.commit()

    role = current_user.role
    data = {}

    if role in ('admin','director'):
        data['total']     = Project.query.count()
        data['inprog']    = Project.query.filter_by(status='InProgress').count()
        data['completed'] = Project.query.filter(Project.status.in_(['Completed','Closed'])).count()
        data['onhold']    = Project.query.filter_by(status='OnHold').count()
        data['cancelled'] = Project.query.filter_by(status='Cancelled').count()
        data['delayed']   = Project.query.filter_by(status='Delayed').count()
        data['projects']  = Project.query.order_by(Project.updated_at.desc()).paginate(
            page=request.args.get('page', 1, type=int), per_page=15, error_out=False)
        active_ids = db.session.query(Project.id).filter(
            Project.status.notin_(['Cancelled', 'OnHold'])).subquery()
        data['collected'] = float(db.session.query(db.func.sum(Payment.amount)).filter(
            Payment.project_id.in_(active_ids)).scalar() or 0)
        data['total_amt'] = float(db.session.query(db.func.sum(Project.total_amount)).filter(
            Project.status.notin_(['Cancelled', 'OnHold'])).scalar() or 0)

    elif role == 'coordinator':
        my_projects    = Project.query.filter_by(coordinator_id=current_user.id).order_by(Project.updated_at.desc()).all()
        my_project_ids = [p.id for p in my_projects]
        total_value    = sum(float(p.total_amount or 0) for p in my_projects if p.status not in ['Cancelled','OnHold'])
        total_collected = sum(float(p.collected_amount or 0) for p in my_projects if p.status not in ['Cancelled','OnHold'])
        subsidy_list   = Subsidy.query.filter(Subsidy.project_id.in_(my_project_ids)).all() if my_project_ids else []
        data['projects']         = my_projects
        data['pending']          = Project.query.filter(Project.status.in_(['Lead','Created'])).count()
        data['site_visits'] = SiteVisit.query.filter(
    SiteVisit.project_id.in_(my_project_ids),
    SiteVisit.status == 'Scheduled'
).all() if my_project_ids else []
        data['delayed']          = Project.query.filter_by(status='Delayed').count()
        data['total_value']      = total_value
        data['total_collected']  = total_collected
        data['total_pending']    = total_value - total_collected
        data['subsidy_pending']  = sum(float(s.expected_amount or 0) - float(s.received_amount or 0) for s in subsidy_list)
        data['subsidy_received'] = sum(float(s.received_amount or 0) for s in subsidy_list)
        data['subsidy_list']     = subsidy_list

    elif role == 'documents':

        from sqlalchemy.orm import joinedload
        from sqlalchemy import func

   
        all_my_projects = Project.query.options(
        joinedload(Project.documents)
    ).filter_by(doc_staff_id=current_user.id).all()

    
        my_projects = [p for p in all_my_projects if p.status not in ['Cancelled', 'OnHold']]

        project_ids = [p.id for p in my_projects]

    
        doc_counts = []
        if project_ids:
            doc_counts = db.session.query(
            Document.project_id,
            func.sum(Document.status == 'Completed').label('done'),
            func.count(Document.id).label('total')
        ).filter(
            Document.project_id.in_(project_ids)
        ).group_by(Document.project_id).all()

    
        doc_map = {
        d.project_id: (int(d.done or 0), int(d.total or 0))
        for d in doc_counts
    }

    
        projects_with_counts = []
        for p in my_projects:
            done, total = doc_map.get(p.id, (0, 0))

            projects_with_counts.append({
            'project': p,
            'done_docs': done,
            'total_docs': total,
            'doc_pct': int(done / total * 100) if total > 0 else 0,
        })

    
        page = request.args.get('page', 1, type=int)
        per_page = 20
        total = len(projects_with_counts)
        start = (page - 1) * per_page

        data['projects'] = my_projects
        data['projects_with_counts'] = projects_with_counts[start:start+per_page]
        data['page'] = page
        data['per_page'] = per_page
        data['total_projects'] = total
        data['total_pages'] = (total + per_page - 1) // per_page

    
        data['queue'] = len(my_projects)

        data['new_projects'] = [
        p for p in my_projects
        if p.status == 'InProgress' and len(p.documents) == 0
    ]
        data['new_count'] = len(data['new_projects'])

        data['completed_projects'] = [
        p for p in my_projects
        if p.status in ['Completed', 'Closed']
    ]
        data['completed_count'] = len(data['completed_projects'])

    
        data['notifications'] = Notification.query.filter_by(
        user_id=current_user.id,
        is_read=False
    ).order_by(Notification.created_at.desc()).limit(20).all()
    elif role == 'documents_k':
        from sqlalchemy.orm import joinedload
        from sqlalchemy import func

        all_my_projects = Project.query.options(
        joinedload(Project.documents)
    ).filter_by(doc_staff_id=current_user.id).all()

        my_projects = [p for p in all_my_projects if p.status not in ['Cancelled', 'OnHold']]
        project_ids = [p.id for p in my_projects]

        doc_counts = []
        if project_ids:
            doc_counts = db.session.query(
            Document.project_id,
            func.sum(Document.status == 'Completed').label('done'),
            func.count(Document.id).label('total')
        ).filter(
            Document.project_id.in_(project_ids)
        ).group_by(Document.project_id).all()

        doc_map = {
        d.project_id: (int(d.done or 0), int(d.total or 0))
        for d in doc_counts
    }

        projects_with_counts = []
        for p in my_projects:
            done, total = doc_map.get(p.id, (0, 0))
            projects_with_counts.append({
            'project': p,
            'done_docs': done,
            'total_docs': total,
            'doc_pct': int(done / total * 100) if total > 0 else 0,
        })

        page = request.args.get('page', 1, type=int)
        per_page = 20
        total = len(projects_with_counts)
        start = (page - 1) * per_page

        data['projects'] = my_projects
        data['projects_with_counts'] = projects_with_counts[start:start+per_page]
        data['page'] = page
        data['per_page'] = per_page
        data['total_projects'] = total
        data['total_pages'] = (total + per_page - 1) // per_page
        data['queue'] = len(my_projects)
        data['new_projects'] = [
        p for p in my_projects
        if p.status == 'InProgress' and len(p.documents) == 0
    ]
        data['new_count'] = len(data['new_projects'])
        data['completed_projects'] = [
        p for p in my_projects if p.status in ['Completed', 'Closed']
    ]
        data['completed_count'] = len(data['completed_projects'])
        data['notifications'] = Notification.query.filter_by(
        user_id=current_user.id, is_read=False
    ).order_by(Notification.created_at.desc()).limit(20).all()
        
    # elif role == 'stocks':
    #     data['total_items'] = StockItem.query.filter_by(is_active=True).count()
    #     data['low_stock']   = StockItem.query.filter(
    #     StockItem.is_active == True,
    #     StockItem.current_qty <= StockItem.reorder_level
    # ).all()
    #     data['low_count']   = len(data['low_stock'])
    #     data['recent_txns'] = (StockTransaction.query
    #                        .order_by(StockTransaction.created_at.desc())
    #                        .limit(10).all())
    #     data['categories']  = sorted({
    #     c[0] for c in db.session.query(StockItem.category)
    #     .filter_by(is_active=True).distinct()
    # })
    
    elif role == 'office':
        from sqlalchemy.orm import joinedload
        from sqlalchemy import func

        all_projects = Project.query.options(
    joinedload(Project.documents)
).filter(
    Project.status.notin_(['Cancelled', 'OnHold']),
    Project.doc_staff_id == current_user.id
).all()

        project_ids = [p.id for p in all_projects]

        doc_counts = []
        if project_ids:
            doc_counts = db.session.query(
            Document.project_id,
            func.sum(Document.status == 'Completed').label('done'),
            func.count(Document.id).label('total')
            ).filter(
            Document.project_id.in_(project_ids)
            ).group_by(Document.project_id).all()

        doc_map = {
        d.project_id: (int(d.done or 0), int(d.total or 0))
        for d in doc_counts
        }

        projects_with_counts = []
        for p in all_projects:
            done, total = doc_map.get(p.id, (0, 0))
            projects_with_counts.append({
            'project': p,
            'done_docs': done,
            'total_docs': total,
            'doc_pct': int(done / total * 100) if total > 0 else 0,
        })

        page = request.args.get('page', 1, type=int)
        per_page = 20
        total = len(projects_with_counts)
        start = (page - 1) * per_page

        data['projects'] = all_projects
        data['projects_with_counts'] = projects_with_counts[start:start + per_page]
        data['page'] = page
        data['per_page'] = per_page
        data['total_projects'] = total
        data['total_pages'] = (total + per_page - 1) // per_page
        data['queue'] = len(all_projects)
        data['new_projects'] = [
        p for p in all_projects
        if p.status == 'InProgress' and len(p.documents) == 0
        ]
        data['new_count'] = len(data['new_projects'])
        data['completed_projects'] = [
        p for p in all_projects if p.status in ['Completed', 'Closed']
        ]
        data['completed_count'] = len(data['completed_projects'])
        data['notifications'] = Notification.query.filter_by(
        user_id=current_user.id, is_read=False
        ).order_by(Notification.created_at.desc()).limit(20).all()
    elif role == 'payments':
        active_ids = db.session.query(Project.id).filter(
            Project.status.notin_(['Cancelled', 'OnHold'])).subquery()
        data['total_collected'] = float(db.session.query(db.func.sum(Payment.amount)).filter(
            Payment.project_id.in_(active_ids)).scalar() or 0)
        total_amt               = float(db.session.query(db.func.sum(Project.total_amount)).filter(
            Project.status.notin_(['Cancelled', 'OnHold'])).scalar() or 0)
        data['total_pending']   = total_amt - data['total_collected']
        data['projects']        = Project.query.filter(
            Project.status.notin_(['Closed', 'Cancelled', 'OnHold'])).paginate(
            page=request.args.get('page', 1, type=int), per_page=20, error_out=False)

    elif role == 'onsite':
        from sqlalchemy.orm import joinedload
        data['projects'] = (Project.query
            .options(
                joinedload(Project.onsite_progress),
                joinedload(Project.materials),
                joinedload(Project.customer),
                joinedload(Project.payments),       
                joinedload(Project.coordinator),       
            )
            .filter(Project.status.in_(['InProgress','Delayed','Lead','Created']),
                Project.work_category != 'Outside')
            .order_by(Project.updated_at.desc())
            .all())
        data['workers'] = Worker.query.filter_by(is_active=True).all()
        data['tasks']   = Notification.query.filter_by(
            user_id=current_user.id, notif_type='task', is_read=False).order_by(
            Notification.created_at.desc()).all()

    elif role == 'service':
        pending  = AppInstallation.query.filter_by(status='Pending').all()
        scheduled = AppInstallation.query.filter_by(status='Scheduled').all()
        data['installs']        = pending
        data['scheduled']       = scheduled
        data['pending_count']   = len(pending)
        data['scheduled_count'] = len(scheduled)
        data['completed']       = AppInstallation.query.filter_by(status='Completed').count()

    return render_template('dashboard.html', data=data)
@app.route('/onsite')
@login_required
@roles_required('admin', 'onsite', 'director')
def onsite_board():

    from sqlalchemy.orm import joinedload

    projects = (Project.query
        .options(
            joinedload(Project.onsite_progress),
            joinedload(Project.materials),
            joinedload(Project.customer),
            joinedload(Project.payments),
            joinedload(Project.coordinator),
        )
        .filter(
            Project.status.in_(['InProgress', 'Delayed', 'Lead', 'Created']),
            Project.work_category != 'Outside'
        )
        .order_by(Project.updated_at.desc())
        .all())

    return render_template('onsite_dashboard.html', data={'projects': projects})

# ─────────────────────────────────────────────────────────────────────────────
# PROJECTS
# ─────────────────────────────────────────────────────────────────────────────
from sqlalchemy import case,cast,Integer
@app.route('/projects')
@login_required
def projects():
    status_filter   = request.args.get('status', '')
    search          = _clean(request.args.get('q', ''), 100)
    consumer_search = _clean(request.args.get('consumer_no', ''), 50)   
    page            = request.args.get('page', 1, type=int)
    q = (Project.query
         .join(Customer)
         .outerjoin(ConnectionDetails, ConnectionDetails.project_id == Project.id))
    if current_user.role == 'coordinator':
        q = q.filter(Project.coordinator_id == current_user.id)
    if current_user.role == 'documents_k':
        q = q.filter(Project.doc_staff_id == current_user.id)
    if status_filter:
        q = q.filter(Project.status == status_filter, Project.work_category != 'Outside')
    if search:
        q = q.filter(
            Customer.name.ilike(f'%{search}%') |
            Project.project_code.ilike(f'%{search}%')
        )
    if consumer_search:
        q = q.filter(ConnectionDetails.consumer_number.ilike(f'%{consumer_search}%'))
    if current_user.role == 'admin':
        q = q.order_by(Project.updated_at.desc())
    else:
        q = q.order_by(cast(Project.project_code, Integer).desc())

    pagination = q.paginate(page=page, per_page=50, error_out=False)
    project_ids_on_page = [p.id for p in pagination.items]

    pending_pay_excess_ids = set(
    r[0] for r in db.session.query(PaymentExcess.project_id)
    .filter(PaymentExcess.project_id.in_(project_ids_on_page),
            PaymentExcess.status == 'Pending')
    .distinct().all()
    ) if project_ids_on_page else set()

    pending_bank_excess_ids = set(
    r[0] for r in db.session.query(BankExcessReturn.project_id)
    .filter(BankExcessReturn.project_id.in_(project_ids_on_page),
            BankExcessReturn.returned == False)
    .distinct().all()
    ) if project_ids_on_page else set()

    pending_excess_ids = pending_pay_excess_ids | pending_bank_excess_ids
    return render_template('projects.html', projects=pagination.items, pagination=pagination,
                       status_filter=status_filter, search=search,
                       consumer_search=consumer_search,
                       pending_excess_ids=pending_excess_ids)   
@app.route('/projects/new', methods=['GET', 'POST'])
@login_required
@roles_required('coordinator','admin','documents','office','documents_k','director')
@limiter.limit('30 per minute')
def new_project():
    coordinators = User.query.filter(
    User.role.in_(['coordinator', 'director'])
).order_by(User.full_name).all()
    customers       = Customer.query.order_by(Customer.name).all()
    doc_staff       = User.query.filter_by(role='documents', is_active=True).all()
    office = User.query.filter_by(role='office').all()
    documents_k=User.query.filter_by(role='documents_k').all()
    suggested_code  = next_project_code()

    if request.method == 'POST':
        raw_structure = request.form.get('structure_capacity_kw', '').strip()
        structure_kw  = _safe_float(raw_structure) if raw_structure else None
        code = _clean(request.form.get('project_code', ''), 20)
        if not code:
            auto = next_project_code() or '1781'
            # Ensure auto-generated code is truly free
            while Project.query.filter_by(project_code=auto).first():
                auto = str(int(auto) + 1)
            code = auto
        if Project.query.filter_by(project_code=code).first():
            flash(f'MNRE number {code} is already registered.', 'danger')
            return render_template('new_project.html', customers=customers,
                                   doc_staff=doc_staff, suggested_code=suggested_code)
        cust_id = request.form.get('customer_id')
        if not cust_id:
            cust = Customer(
        name       = _clean(request.form.get('customer_name', ''), 120),
        phone      = _clean(request.form.get('phone', ''), 20) or None,
        email      = _clean(request.form.get('email', ''), 120) or None,
        house_name = _clean(request.form.get('house_name', ''), 120) or None,
        place      = _clean(request.form.get('place', ''), 120) or None,
        post       = _clean(request.form.get('post', ''), 120) or None,
        pincode    = _clean(request.form.get('pincode', ''), 10) or None,
        village    = _clean(request.form.get('village', ''), 120) or None,
        district   = _clean(request.form.get('district', ''), 80) or None,
        taluk      = _clean(request.form.get('taluk', ''), 120) or None,
        sub_co=request.form.get('sub_co','').strip() or None,
        )
            db.session.add(cust)
            db.session.flush()
            cust_id = cust.id

        # Resolve coordinator
        raw_coord_id    = request.form.get('coordinator_id') or ''
        coord_name_other = _clean(request.form.get('coordinator_name_other', ''), 120)

        if raw_coord_id == '__other__':
            resolved_coord_id   = None
            resolved_coord_name = coord_name_other or None
        else:
            resolved_coord_id   = int(raw_coord_id) if raw_coord_id else None
            resolved_coord_name = None

        notes_val = _clean(request.form.get('notes', ''), 2000)

        proj = Project(
    project_code         = code,
    customer_id          = cust_id,
    inverter_capacity_kw = _safe_float(request.form.get('inverter_capacity_kw')),
    panel_capacity_kw    = _safe_float(request.form.get('panel_capacity_kw')),
    structure_capacity_kw = structure_kw,
    project_type         = request.form['project_type'],
    status               = 'InProgress',
    stage                = 'Documentation',
    project_subtype      = request.form.get('project_subtype') or None,
    total_amount         = _safe_float(request.form.get('total_amount', 0)),
    coordinator_id       = resolved_coord_id,
    doc_staff_id         = request.form.get('doc_staff_id') or None,
    notes                = notes_val,
    roof_type            = request.form.get('roof_type') or None,
    roof_type_other      = _clean(request.form.get('roof_type_other', ''), 60) or None,
    inverter_type        = request.form.get('inverter_type') or None,
    coordinator_name     = resolved_coord_name,
    work_category         = request.form.get('work_category') if request.form.get('work_category') in ('Installation', 'Outside') else 'Installation',
        )
        db.session.add(proj)
        db.session.flush()
        log_action(proj.id, 'Project created', new_val='Created')
        geo_photo    = request.files.get('geo_photo')
        maps_url_raw = _clean(request.form.get('maps_url', ''), 500)
        has_photo    = False
        if (geo_photo and geo_photo.filename) or maps_url_raw:
            tag = ProjectGeoTag(project_id=proj.id, uploaded_by=current_user.id,
                                 uploaded_at=datetime.utcnow())
            if geo_photo and geo_photo.filename:
                ext = os.path.splitext(geo_photo.filename)[1].lower()
                if ext in ('.jpg', '.jpeg', '.png'):
                    os.makedirs(GEO_UPLOAD_DIR, exist_ok=True)
                    fname = f'{proj.project_code}_site_{int(datetime.utcnow().timestamp())}{ext}'
                    fpath = os.path.join(GEO_UPLOAD_DIR, fname)
                    geo_photo.save(fpath)
                    tag.site_photo_path = fpath          # ← changed from photo_path
                    tag.site_photo_uploaded_by = current_user.id
                    tag.site_photo_uploaded_at = datetime.utcnow()
                    has_photo = True
                else:
                    flash('Site photo must be JPG or PNG — photo was skipped.', 'warning')
            if maps_url_raw:
                tag.maps_url = maps_url_raw
                lat, lng = parse_gmaps_coords(maps_url_raw)
                if lat is not None:
                    tag.latitude, tag.longitude, tag.has_gps = lat, lng, True
                else:
                    flash('Maps link saved, but coordinates could not be read from it.', 'warning')
            db.session.add(tag)
            log_action(proj.id, 'Pre-work site photo/location added at project creation',
                       new_val='Photo' if has_photo else 'Location only')

        if proj.doc_staff_id:
            create_notification(
                proj.doc_staff_id, proj.id,
                f'You have been assigned to {proj.project_code}-{proj.customer.name} '
                f'({proj.project_type}, {proj.inverter_capacity_kw} kW).', 'task',
            )
        if proj.doc_staff_id:
            create_notification(
                proj.doc_staff_id, proj.id,
                f'You have been assigned to {proj.project_code}-{proj.customer.name} '
                f'({proj.project_type}, {proj.inverter_capacity_kw} kW).', 'task',
            )
        # ── Notify onsite team of new work ────────────────────────────────
        if proj.project_type == 'Cash':
            notify_onsite_team(proj.id,
                f'New cash work: {proj.project_code} — {proj.customer.name} '
                f'({proj.inverter_capacity_kw} kW). Assigned by {current_user.full_name}.', 'task')
        elif proj.project_type == 'Loan':
            notify_onsite_team(proj.id,
                f'New loan work: {proj.project_code} — {proj.customer.name} '
                f'({proj.inverter_capacity_kw} kW). Awaiting first bank payment before site work begins.', 'info')
        db.session.commit()
        flash(f'Project {proj.project_code} created successfully!', 'success')
        return redirect(url_for('project_detail', pid=proj.id))

    return render_template('new_project.html',coordinators=coordinators, customers=customers,
                           doc_staff=doc_staff, suggested_code=suggested_code,office=office,documents_k=documents_k)


@app.route('/projects/<int:pid>/edit', methods=['GET', 'POST'])
@login_required
@roles_required('admin', 'coordinator', 'documents','office','documents_k','director')
def edit_project(pid):
    proj = Project.query.get_or_404(pid)

    if current_user.role == 'coordinator' and proj.coordinator_id != current_user.id:
        flash('You can only edit your own projects.', 'danger')
        return redirect(url_for('project_detail', pid=pid))
    if current_user.role in ('documents','documents_k') and proj.doc_staff_id != current_user.id:
        flash('You can only edit projects assigned to you.', 'danger')
        return redirect(url_for('project_detail', pid=pid))
    if proj.status in ('Cancelled', 'Closed') and current_user.role != 'admin':
        flash('Cancelled or closed projects cannot be edited.', 'danger')
        return redirect(url_for('project_detail', pid=pid))

    doc_staff    = User.query.filter_by(role='documents',   is_active=True).all()
    coordinators = User.query.filter(
    User.role.in_(['coordinator', 'director']),
    User.is_active == True
).order_by(User.full_name).all()
    office = User.query.filter_by(role='office',is_active=True).all()
    documents_k=User.query.filter_by(role='documents_k',is_active=True).all()

    if request.method == 'POST':
        old_type     = proj.project_type
        old_subtype  = proj.project_subtype
        old_inverter_type = proj.inverter_type
        old_loan_sub = proj.loan_subtype
        old_amount   = float(proj.total_amount or 0)
        old_inverter_capacity_kw  = proj.inverter_capacity_kw
        old_panel_capacity_kw     = proj.panel_capacity_kw
        old_structure_capacity_kw = proj.structure_capacity_kw
        old_type     = proj.project_type
        old_subtype  = proj.project_subtype
        old_inverter_type = proj.inverter_type
        old_loan_sub = proj.loan_subtype
        old_amount   = float(proj.total_amount or 0)

        new_type     = request.form['project_type']
        new_subtype  = request.form.get('project_subtype') or None
        # new_loan_sub = request.form.get('loan_subtype') or None
        new_amount   = _safe_float(request.form.get('total_amount'))

        proj.inverter_capacity_kw = _safe_float(request.form.get('inverter_capacity_kw'))
        proj.panel_capacity_kw    = _safe_float(request.form.get('panel_capacity_kw'))
        raw_structure = request.form.get('structure_capacity_kw', '').strip()
        proj.structure_capacity_kw = _safe_float(raw_structure) if raw_structure else None 
        proj.project_type         = new_type
        proj.project_subtype      = new_subtype
        # proj.loan_subtype         = new_loan_sub
        proj.total_amount         = new_amount
        proj.notes                = _clean(request.form.get('notes', ''), 2000)
        proj.roof_type=request.form.get('roof_type') or None
        proj.roof_type_other = _clean(request.form.get('roof_type_other', ''), 60) or None
        proj.inverter_type = request.form.get('inverter_type') or None
        changes = []
        capacity_changes = []
        if old_inverter_capacity_kw != proj.inverter_capacity_kw:
            capacity_changes.append(
                f'Inverter capacity: {old_inverter_capacity_kw or 0} kW → {proj.inverter_capacity_kw or 0} kW'
            )
        if old_panel_capacity_kw != proj.panel_capacity_kw:
            capacity_changes.append(
            f'Panel capacity: {old_panel_capacity_kw or 0} kW → {proj.panel_capacity_kw or 0} kW'
            )
        if old_structure_capacity_kw != proj.structure_capacity_kw:
            capacity_changes.append(
            f'Structure capacity: {old_structure_capacity_kw or 0} kW → {proj.structure_capacity_kw or 0} kW'
            )

        if capacity_changes:
            onsite_users = User.query.filter_by(role='onsite', is_active=True).all()
            for u in onsite_users:
                create_notification(
            u.id, pid,
            f'{proj.project_code} — {proj.customer.name}: Capacity details updated '
            f'by {current_user.full_name}. {", ".join(capacity_changes)}.',
            'info'
        )

        if current_user.role == 'admin':
            new_created_at = request.form.get('created_at_override')
            if new_created_at:
                try:
                    proj.created_at=datetime.fromisoformat(new_created_at)
                except ValueError:
                    flash('Invalid data format for Created At.','danger')
            new_code = _clean(request.form.get('project_code', ''), 20)
            if new_code and new_code != proj.project_code:
                if Project.query.filter(Project.project_code == new_code, Project.id != pid).first():
                    flash('That MNRE number is already in use.', 'danger')
                    return redirect(url_for('edit_project', pid=pid))
                changes.append(f'MNRE: {proj.project_code} → {new_code}')
                proj.project_code = new_code

            # ── Work category override ──────────────────────────────────
            new_work_category = request.form.get('work_category')
            if new_work_category in ('Installation', 'Outside') and new_work_category != proj.work_category:
                changes.append(f'Work category: {proj.work_category} → {new_work_category}')
                proj.work_category = new_work_category
                if new_work_category == 'Outside':
                    # Outside-work projects don't have a service schedule
                    ServiceRecord.query.filter_by(project_id=pid).delete()
        if current_user.role in ('admin','documents','office','documents_k'):
            proj.customer.name       = _clean(request.form.get('customer_name', proj.customer.name), 120)
            proj.customer.phone      = _clean(request.form.get('customer_phone', ''), 20) or None
            proj.customer.email      = _clean(request.form.get('customer_email', ''), 120) or None
            proj.customer.house_name = _clean(request.form.get('customer_house_name', ''), 120) or None
            proj.customer.place      = _clean(request.form.get('customer_place', ''), 120) or None
            proj.customer.post       = _clean(request.form.get('customer_post', ''), 120) or None
            proj.customer.pincode    = _clean(request.form.get('customer_pincode', ''), 10) or None
            proj.customer.village    = _clean(request.form.get('customer_village', ''), 120) or None
            proj.customer.district   = _clean(request.form.get('customer_district', ''), 80) or None
            proj.customer.taluk      = _clean(request.form.get('customer_taluk', ''), 120) or None
            proj.customer.sub_co = request.form.get('customer_sub_co', '').strip() or None

            new_stage  = request.form.get('stage')
            new_status = request.form.get('status')
            if new_stage and new_stage != proj.stage:
                changes.append(f'Stage: {proj.stage} → {new_stage}')
                proj.stage = new_stage
                proj.staged_changed_at = datetime.utcnow()
            if new_status and new_status != proj.status:
                changes.append(f'Status: {proj.status} → {new_status}')
                proj.status = new_status
                # ── NEW: create service schedule if it just became Completed/Closed ──
                if proj.status in ('Completed', 'Closed') and proj.work_category != 'Outside':
                    create_service_schedule(proj)
            if 'coordinator_id' in request.form:
                raw_coord_id     = request.form.get('coordinator_id') or None
                coord_name_other = _clean(request.form.get('coordinator_name_other', ''), 120)

                if raw_coord_id == '__other__':
                    if proj.coordinator_id:
                        old_coord = proj.coordinator
                        create_notification(old_coord.id, pid,
                            f'You have been unassigned as coordinator from {proj.project_code} — {proj.customer.name}.', 'info')
                        changes.append(f'Coordinator: {old_coord.full_name} → {coord_name_other}')
                        proj.coordinator_id = None
                    proj.coordinator_name = coord_name_other or None

                elif raw_coord_id:
                    new_coord_id = int(raw_coord_id)
                    if proj.coordinator_id != new_coord_id:
                        old_coord = proj.coordinator
                        new_coord = User.query.get(new_coord_id)
                        if old_coord:
                            create_notification(old_coord.id, pid,
                            f'You have been unassigned as coordinator from {proj.project_code} — {proj.customer.name}.', 'info')
                        if new_coord:
                            create_notification(new_coord_id, pid,
                            f'You have been assigned as coordinator for {proj.project_code} — {proj.customer.name}.', 'task')
                            changes.append(
                            f'Coordinator: {old_coord.full_name if old_coord else "None"} → '
                            f'{new_coord.full_name if new_coord else "None"}'
                            )
                        proj.coordinator_id = new_coord_id
                        proj.coordinator_name = None
                elif raw_coord_id == '__clear__':          
                    if proj.coordinator_id:
                        old_coord = proj.coordinator
                        create_notification(old_coord.id, pid,
                            f'You have been unassigned as coordinator from {proj.project_code} — {proj.customer.name}.', 'info')
                        changes.append(f'Coordinator: {old_coord.full_name} → None')
                    proj.coordinator_id = None
                    proj.coordinator_name = None

        if new_amount > float(proj.collected_amount or 0):
            if proj.status == 'Completed' and proj.stage == 'Payment':
                proj.status = 'InProgress'
        if 'doc_staff_id' in request.form:
            new_staff_id = request.form.get('doc_staff_id') or None
            if new_staff_id:
                new_staff_id = int(new_staff_id)
                old_staff    = proj.doc_staff
                if old_staff and old_staff.id != new_staff_id:
                    create_notification(old_staff.id, pid,
                    f'You have been unassigned from {proj.project_code} — {proj.customer.name}.', 'info')
                if proj.doc_staff_id != new_staff_id:
                    create_notification(new_staff_id, pid,
                    f'You have been assigned to {proj.project_code} — {proj.customer.name} '
                    f'({proj.project_type}, {proj.inverter_capacity_kw} kW).', 'task')
                proj.doc_staff_id = new_staff_id
            else:
                proj.doc_staff_id = None

        if old_type != new_type:
            changes.append(f'Type: {old_type} → {new_type}')
        if old_subtype != new_subtype:
            changes.append(f'Subtype: {old_subtype or "None"} → {new_subtype or "None"}')
        if old_inverter_type != proj.inverter_type:
           changes.append(f'Inverter type: {old_inverter_type or "None"} → {proj.inverter_type or "None"}')
        # if old_loan_sub != new_loan_sub:
        #     changes.append(f'Loan type: {old_loan_sub or "None"} → {new_loan_sub or "None"}')
        if abs(old_amount - new_amount) > 0.01:
            changes.append(f'Amount: ₹{old_amount:,.0f} → ₹{new_amount:,.0f}')

        if current_user.role in ('coordinator', 'admin') and changes and proj.doc_staff_id:
            create_notification(proj.doc_staff_id, pid,
                f'{proj.project_code} — {proj.customer.name}: Project details edited '
                f'by {current_user.full_name}. Changes: {", ".join(changes)}.', 'info')
        if current_user.role in ('documents','office','documents_k') and changes and proj.coordinator_id:
            create_notification(proj.coordinator_id, pid,
                f'{proj.project_code} — {proj.customer.name}: Project details edited '
                f'by {current_user.full_name} (docs). Changes: {", ".join(changes)}.', 'info')

        if abs(old_amount - new_amount) > 0.01:
            pending        = new_amount - float(proj.collected_amount or 0)
            payments_users = User.query.filter_by(role='payments', is_active=True).all()
            for u in payments_users:
                msg = (
                    f'{proj.project_code} — {proj.customer.name}: Contract amount revised '
                    f'from ₹{old_amount:,.0f} to ₹{new_amount:,.0f} by {current_user.full_name}. '
                    + (f'Outstanding: ₹{pending:,.0f}.' if pending > 0
                       else f'Already collected ₹{float(proj.collected_amount or 0):,.0f}.')
                )
                create_notification(u.id, pid, msg, 'task' if pending > 0 else 'info')
        # ── Connection details (inline save) ──────────────────────────────────
        
        if 'connection_type' in request.form:
            cd = proj.connection_details or ConnectionDetails(project_id=pid)
            cd.connection_type = request.form.get('connection_type') or None
            cd.category = request.form.get('category') or None
            cd.consumer_number = _clean(request.form.get('consumer_number', ''), 50) or None
            cd.kseb_section = _clean(request.form.get('kseb_section', ''), 100) or None
            ownership_needed           = 'ownership_change_needed' in request.form
            cd.ownership_change_needed = ownership_needed
            cd.ownership_change_status = (
            request.form.get('ownership_change_status', 'Pending')
            if ownership_needed else 'Not Required'
            )
            load_needed              = 'load_clearance_needed' in request.form
            cd.load_clearance_needed = load_needed
            cd.load_clearance_status = (
                request.form.get('load_clearance_status', 'Pending')
            if load_needed else 'Not Required'
            )
            cd.notes      = _clean(request.form.get('cd_notes', ''), 500)
            cd.updated_by = current_user.id
            if not cd.id:
                db.session.add(cd)
        # ── Optional PRE-WORK site photo / location update ────────────────
        geo_photo    = request.files.get('geo_photo')
        maps_url_raw = _clean(request.form.get('maps_url', ''), 500)
        if (geo_photo and geo_photo.filename) or maps_url_raw:
            tag = proj.geo_tag or ProjectGeoTag(project_id=pid)
            if geo_photo and geo_photo.filename:
                ext = os.path.splitext(geo_photo.filename)[1].lower()
                if ext in ('.jpg', '.jpeg', '.png'):
                    if tag.site_photo_path and os.path.isfile(tag.site_photo_path):
                        try:
                            os.remove(tag.site_photo_path)
                        except OSError:
                            pass
                    os.makedirs(GEO_UPLOAD_DIR, exist_ok=True)
                    fname = f'{proj.project_code}_site_{int(datetime.utcnow().timestamp())}{ext}'
                    fpath = os.path.join(GEO_UPLOAD_DIR, fname)
                    geo_photo.save(fpath)
                    tag.site_photo_path = fpath           # ← changed from photo_path
                    tag.site_photo_uploaded_by = current_user.id
                    tag.site_photo_uploaded_at = datetime.utcnow()
                else:
                    flash('Site photo must be JPG or PNG — photo was skipped.', 'warning')
            if maps_url_raw:
                tag.maps_url = maps_url_raw
                lat, lng = parse_gmaps_coords(maps_url_raw)
                if lat is not None:
                    tag.latitude, tag.longitude, tag.has_gps = lat, lng, True
                else:
                    flash('Maps link saved, but coordinates could not be read from it.', 'warning')
            tag.uploaded_by = current_user.id
            tag.uploaded_at = datetime.utcnow()
            if not tag.id:
                db.session.add(tag)
            log_action(pid, 'Pre-work site photo/location updated')
        log_action(pid, 'Project edited: ' + (', '.join(changes) if changes else 'details updated'))
        db.session.commit()
        pending = new_amount - float(proj.collected_amount or 0)
        flash(f'Project updated. Outstanding balance: ₹{pending:,.0f}.' if pending > 0
              else 'Project details updated successfully.', 'success')
        return redirect(url_for('project_detail', pid=pid))

    return render_template('edit_project.html', proj=proj,
                           doc_staff=doc_staff, coordinators=coordinators,office=office,documents_k=documents_k)

@app.route('/projects/<int:pid>/geo_tag/delete_photo', methods=['POST'])
@login_required
@roles_required('admin', 'onsite', 'coordinator', 'documents', 'office', 'documents_k', 'director')
def delete_geo_photo(pid):
    tag = ProjectGeoTag.query.filter_by(project_id=pid).first()
    if tag and tag.photo_path:
        if os.path.isfile(tag.photo_path):
            try:
                os.remove(tag.photo_path)
            except OSError:
                pass
        tag.photo_path = None
        log_action(pid, 'Site photo removed')
        db.session.commit()
        flash('Site photo removed.', 'success')
    else:
        flash('No site photo to remove.', 'warning')
    return redirect(request.referrer or url_for('project_detail', pid=pid))


@app.route('/projects/<int:pid>/geo_tag/delete_location', methods=['POST'])
@login_required
@roles_required('admin', 'onsite', 'coordinator', 'documents', 'office', 'documents_k', 'director')
def delete_geo_location(pid):
    tag = ProjectGeoTag.query.filter_by(project_id=pid).first()
    if tag and tag.has_gps:
        tag.latitude = None
        tag.longitude = None
        tag.has_gps = False
        tag.maps_url = None
        log_action(pid, 'Site location removed')
        db.session.commit()
        flash('Location removed.', 'success')
    else:
        flash('No location to remove.', 'warning')
    return redirect(url_for('edit_project', pid=pid))
@app.route('/projects/<int:pid>/site_visit', methods=['POST'])
@login_required
@roles_required('admin', 'coordinator','director')
def add_site_visit(pid):
    proj = Project.query.get_or_404(pid)
    if proj.status in ('Cancelled', 'OnHold'):
        flash('Cannot schedule a site visit for this project.', 'danger')
        return redirect(url_for('project_detail', pid=pid))
    if proj.site_visits:
        flash('A site visit has already been scheduled for this project.', 'warning')
        return redirect(url_for('project_detail', pid=pid))
    visit = SiteVisit(
        project_id     = pid,
        scheduled_date = date.fromisoformat(request.form['scheduled_date'])
                         if request.form.get('scheduled_date') else date.today(),
        conducted_by   = current_user.id,
        status         = 'Scheduled',
    )
    db.session.add(visit)
    log_action(pid, 'Site visit scheduled', new_val='Scheduled')
    db.session.commit()
    flash('Site visit scheduled.', 'success')
    return redirect(url_for('project_detail', pid=pid))


@app.route('/projects/<int:pid>/site_visit/<int:vid>/complete', methods=['POST'])
@login_required
@roles_required('admin', 'coordinator','director')
def complete_site_visit(vid, pid):
    visit              = SiteVisit.query.get_or_404(vid)
    visit.status       = 'Completed'
    visit.visited_date = date.fromisoformat(request.form['visited_date']) \
                         if request.form.get('visited_date') else date.today()
    visit.observations = _clean(request.form.get('observations', ''), 2000)
    log_action(pid, 'Site visit completed', new_val='Completed')
    db.session.commit()
    flash('Site visit marked complete.', 'success')
    return redirect(url_for('project_detail', pid=pid))

@app.route('/projects/<int:pid>')
@login_required
def project_detail(pid):
    proj        = Project.query.get_or_404(pid)
    stages      = get_document_stages()
    logs        = ProjectLog.query.filter_by(project_id=pid).order_by(ProjectLog.created_at.desc()).all()
    workers     = Worker.query.filter_by(is_active=True).all()
    all_workers = Worker.query.filter_by(is_active=True).all()
    worker_rate = {str(w.id): float(w.rate_per_day or 0) for w in all_workers}
    return render_template('project_detail.html', proj=proj, logs=logs,
                           workers=workers, all_workers=all_workers,
                           worker_rate=worker_rate, today=date.today())


@app.route('/projects/<int:pid>/expenses', methods=['POST'])
@login_required
@roles_required('admin', 'documents','office','documents_k')
def update_expense(pid):
    proj         = Project.query.get_or_404(pid)
    expense_type = _clean(request.form.get('expense_type', ''), 50)
    if expense_type not in ('CD Payment', 'Meter', 'Load', 'Additional'):
        flash('Invalid expense type.', 'danger')
        return redirect(url_for('documents', pid=pid))
    amount       = _safe_float(request.form.get('amount'))
    paid_by      = request.form.get('paid_by', 'Customer')
    paid_date    = request.form.get('paid_date')
    notes        = _clean(request.form.get('notes', ''), 500)

    existing = ProjectExpense.query.filter_by(project_id=pid, expense_type=expense_type).first()
    if existing:
        existing.amount    = amount
        existing.paid_by   = paid_by
        existing.paid_date = date.fromisoformat(paid_date) if paid_date else existing.paid_date
        existing.notes     = notes
        existing.recorded_by = current_user.id
    else:
        db.session.add(ProjectExpense(
            project_id=pid, expense_type=expense_type, amount=amount,
            paid_by=paid_by, paid_date=date.fromisoformat(paid_date) if paid_date else None,
            notes=notes, recorded_by=current_user.id,
        ))

    log_action(pid, f'{expense_type} recorded: paid by {paid_by}, ₹{amount:,.0f}', new_val=paid_by)
    if paid_by == 'Company':
        for u in User.query.filter_by(role='payments', is_active=True).all():
            create_notification(u.id, pid,
                f'{proj.project_code} — {proj.customer.name}: {expense_type} of '
                f'₹{amount:,.0f} paid by company. To be recovered from customer.', 'task')
    db.session.commit()
    flash(f'{expense_type} updated.', 'success')
    return redirect(url_for('documents', pid=pid))
@app.route('/projects/<int:pid>/bank_excess/save', methods=['POST'])
@login_required
@roles_required('admin', 'payments')
def save_bank_excess(pid):
    proj = Project.query.get_or_404(pid)
    if proj.project_type != 'Loan':
        flash('Bank excess only applies to Loan projects.', 'danger')
        return redirect(url_for('project_detail', pid=pid))
 
    exc = proj.bank_excess or BankExcessReturn(project_id=pid)
    exc.excess_amount  = _safe_float(request.form.get('excess_amount'))
    exc.received_date  = date.fromisoformat(request.form['received_date']) if request.form.get('received_date') else None
    exc.notes          = _clean(request.form.get('notes', ''), 300) or None
    exc.recorded_by    = current_user.id
 
    if not exc.id:
        db.session.add(exc)
 
    log_action(pid, f'Bank excess recorded: ₹{float(exc.excess_amount):,.0f}',
               new_val=str(exc.excess_amount))
 
    for u in User.query.filter_by(role='payments', is_active=True).all():
        if u.id != current_user.id:
            create_notification(u.id, pid,
                f'{proj.project_code} — {proj.customer.name}: Bank excess of '
                f'₹{float(exc.excess_amount):,.0f} recorded. To be returned.', 'task')
 
    db.session.commit()
    flash(f'Bank excess of ₹{float(exc.excess_amount):,.0f} recorded.', 'success')
    return redirect(url_for('project_detail', pid=pid))
 
 
@app.route('/projects/<int:pid>/bank_excess/return', methods=['POST'])
@login_required
@roles_required('admin', 'payments')
def return_bank_excess(pid):
    exc = BankExcessReturn.query.filter_by(project_id=pid).first_or_404()
    action = request.form.get('action', 'Returned')
    if action not in ('Returned', 'Retained', 'Adjusted'):
        flash('Please choose a valid settlement option.', 'danger')
        return redirect(url_for('project_detail', pid=pid))

    exc.returned            = True
    exc.action               = action
    exc.returned_to          = request.form.get('returned_to') if action == 'Returned' else None
    exc.returned_date        = date.fromisoformat(request.form['returned_date']) if request.form.get('returned_date') else date.today()
    exc.returned_method      = request.form.get('returned_method') or None
    exc.returned_reference   = _clean(request.form.get('returned_reference', ''), 100) or None
    exc.returned_notes       = _clean(request.form.get('returned_notes', ''), 300) or None

    log_action(pid, f'Bank excess settled: ₹{float(exc.excess_amount):,.0f} — {exc.action_label}',
               new_val=action)
    db.session.commit()
    flash(f'Bank excess of ₹{float(exc.excess_amount):,.0f} marked as {exc.action_label}.', 'success')
    return redirect(url_for('project_detail', pid=pid))
@app.route('/payment_excess/<int:eid>/settle', methods=['POST'])
@login_required
@roles_required('admin', 'payments')
def settle_payment_excess(eid):
    exc = PaymentExcess.query.get_or_404(eid)
    action = request.form.get('action')
    if action not in ('Returned', 'Retained', 'Adjusted'):
        flash('Please choose a valid settlement option.', 'danger')
        return redirect(url_for('project_detail', pid=exc.project_id))

    exc.action               = action
    exc.returned_to          = request.form.get('returned_to') if action == 'Returned' else None
    exc.settlement_method    = request.form.get('settlement_method') or None
    exc.settlement_date      = date.fromisoformat(request.form['settlement_date']) if request.form.get('settlement_date') else date.today()
    exc.settlement_reference = _clean(request.form.get('settlement_reference', ''), 100) or None
    exc.settlement_notes     = _clean(request.form.get('settlement_notes', ''), 300) or None
    exc.status                = 'Settled'
    exc.settled_by           = current_user.id

    log_action(exc.project_id,
        f'Payment excess of ₹{float(exc.amount):,.0f} settled: {exc.action_label}',
        new_val=exc.action)
    db.session.commit()
    flash(f'Excess of ₹{float(exc.amount):,.0f} marked as {exc.action_label}.', 'success')
    return redirect(url_for('project_detail', pid=exc.project_id))

@app.route('/projects/<int:pid>/expenses/<int:eid>/mark_recovered', methods=['POST'])
@login_required
@roles_required('admin', 'payments')
def mark_expense_recovered(pid, eid):
    expense                    = ProjectExpense.query.get_or_404(eid)
    expense.recovered          = True
    expense.recovered_date     = date.fromisoformat(request.form['recovery_date']) if request.form.get('recovery_date') else date.today()
    expense.recovery_method    = request.form.get('recovery_method') or None
    expense.recovery_reference = _clean(request.form.get('recovery_reference', ''), 100) or None
    expense.recovery_notes     = _clean(request.form.get('recovery_notes', ''), 300) or None
    log_action(pid, f'{expense.expense_type} marked as recovered from customer')
    db.session.commit()
    flash(f'{expense.expense_type} marked as recovered.', 'success')
    return redirect(url_for('project_detail', pid=pid))


@app.route('/projects/<int:pid>/update_status', methods=['POST'])
@login_required
def update_status(pid):
    proj       = Project.query.get_or_404(pid)
    new_status = request.form.get('status')
    new_stage  = request.form.get('stage')
    old_status = proj.status

    if proj.status in ('Cancelled', 'OnHold'):
        flash('Cannot update status of a Cancelled or On Hold project.', 'danger')
        return redirect(url_for('project_detail', pid=pid))
    if new_stage:
        proj.stage = new_stage
        if not new_status:
            new_status = STAGE_STATUS_MAP.get(new_stage, proj.status)
    if new_status:
        proj.status = new_status
    if proj.status == 'Completed':
        if not proj.app_install or proj.app_install.status != 'Completed':
            if current_user.role != 'admin':
                flash('Project cannot be marked Completed until App Installation is done.', 'danger')
                proj.status = old_status

    log_action(pid, 'Status updated', old_val=old_status, new_val=proj.status)

    
    if proj.status in ('Completed', 'Closed') and proj.status != old_status and proj.work_category != 'Outside':
        create_service_schedule(proj)

    db.session.commit()
    flash('Project status updated.', 'success')
    return redirect(url_for('project_detail', pid=pid))


@app.route('/projects/<int:pid>/cancel', methods=['POST'])
@login_required
@roles_required('admin', 'coordinator', 'documents','office','documents_k','director')
def cancel_project(pid):
    proj = Project.query.get_or_404(pid)
    if current_user.role == 'coordinator' and proj.coordinator_id != current_user.id:
        flash('You can only cancel your own projects.', 'danger')
        return redirect(url_for('project_detail', pid=pid))
    if proj.status == 'Cancelled':
        flash('Project is already cancelled.', 'warning')
        return redirect(url_for('project_detail', pid=pid))
    if proj.status == 'Closed':
        flash('Closed projects cannot be cancelled.', 'danger')
        return redirect(url_for('project_detail', pid=pid))

    old_status  = proj.status
    proj.status = 'Cancelled'
    reason      = _clean(request.form.get('reason', ''), 500)
    log_action(pid, f'Project cancelled. Reason: {reason or "No reason provided"}',
               old_val=old_status, new_val='Cancelled')
    if proj.doc_staff_id:
        create_notification(proj.doc_staff_id, pid,
            f'Project {proj.project_code} — {proj.customer.name} has been cancelled.', 'warning')
    if proj.coordinator_id and proj.coordinator_id != current_user.id:
        create_notification(proj.coordinator_id, pid,
            f'Project {proj.project_code} — {proj.customer.name} has been cancelled by {current_user.full_name}.',
            'warning')
    db.session.commit()
    flash(f'Project {proj.project_code} has been cancelled.', 'success')
    return redirect(url_for('project_detail', pid=pid))


@app.route('/projects/<int:pid>/uncancel', methods=['POST'])
@login_required
@roles_required('admin')
def uncancel_project(pid):
    proj = Project.query.get_or_404(pid)
    if proj.status != 'Cancelled':
        flash('Project is not cancelled.', 'warning')
        return redirect(url_for('project_detail', pid=pid))
    reason      = _clean(request.form.get('reason', ''), 500)
    proj.status = 'InProgress' if proj.stage != 'Lead' else 'Lead'
    log_action(pid, f'Project uncancelled. Reason: {reason or "No reason provided"}',
               old_val='Cancelled', new_val=proj.status)
    if proj.coordinator_id:
        create_notification(proj.coordinator_id, pid,
            f'Project {proj.project_code} — {proj.customer.name} has been reinstated by {current_user.full_name}.', 'info')
    if proj.doc_staff_id:
        create_notification(proj.doc_staff_id, pid,
            f'Project {proj.project_code} — {proj.customer.name} has been reinstated.', 'info')
    db.session.commit()
    flash(f'Project {proj.project_code} has been reinstated.', 'success')
    return redirect(url_for('project_detail', pid=pid))


@app.route('/projects/<int:pid>/hold', methods=['POST'])
@login_required
@roles_required('admin', 'coordinator', 'documents','office','documents_k','director')
def hold_project(pid):
    proj = Project.query.get_or_404(pid)
    if current_user.role in ('documents','documents_k') and proj.doc_staff_id != current_user.id:
        flash('You can only put your assigned projects on hold.', 'danger')
        return redirect(url_for('project_detail', pid=pid))
    if proj.status in ('Cancelled', 'Closed', 'Completed'):
        flash(f'Cannot put a {proj.status} project on hold.', 'danger')
        return redirect(url_for('project_detail', pid=pid))

    old_status = proj.status
    reason     = _clean(request.form.get('reason', ''), 500)

    if proj.status == 'OnHold':
        proj.status = 'InProgress'
        log_action(pid, 'Project resumed from On Hold', old_val='OnHold', new_val='InProgress')
        if proj.coordinator_id:
            create_notification(proj.coordinator_id, pid,
                f'Project {proj.project_code} — {proj.customer.name} has been resumed.', 'info')
        flash(f'Project {proj.project_code} resumed.', 'success')
    else:
        proj.status = 'OnHold'
        log_action(pid, f'Project put on hold. Reason: {reason or "No reason provided"}',
                   old_val=old_status, new_val='OnHold')
        if proj.coordinator_id:
            create_notification(proj.coordinator_id, pid,
                f'Project {proj.project_code} — {proj.customer.name} has been put On Hold'
                + (f': {reason}' if reason else '.'), 'warning')
        flash(f'Project {proj.project_code} is now On Hold.', 'warning')

    db.session.commit()
    return redirect(url_for('project_detail', pid=pid))


@app.route('/projects/<int:pid>/assign_doc_staff', methods=['POST'])
@login_required
@roles_required('coordinator','director')
def assign_doc_staff(pid):
    proj         = Project.query.get_or_404(pid)
    new_staff_id = request.form.get('doc_staff_id')
    if not new_staff_id:
        flash('Please select a staff member.', 'danger')
        return redirect(url_for('project_detail', pid=pid))
    new_staff_id = int(new_staff_id)
    old_staff    = proj.doc_staff
    if old_staff and old_staff.id != new_staff_id:
        create_notification(old_staff.id, pid,
            f'You have been unassigned from {proj.project_code} — {proj.customer.name}.', 'info')
    proj.doc_staff_id = new_staff_id
    new_staff = User.query.get(new_staff_id)
    create_notification(new_staff_id, pid,
        f'You have been assigned to {proj.project_code} — {proj.customer.name} '
        f'({proj.project_type}, {proj.inverter_capacity_kw} kW).', 'task')
    log_action(pid, 'Doc staff assigned',
               old_val=old_staff.full_name if old_staff else None,
               new_val=new_staff.full_name if new_staff else None)
    db.session.commit()
    flash(f'{new_staff.full_name} assigned and notified.', 'success')
    return redirect(url_for('project_detail', pid=pid))


@app.route('/coordinator/analytics')
@login_required
@roles_required('coordinator','director')
def coordinator_analytics():
    from collections import Counter

    today          = date.today()
    my_projects    = Project.query.filter_by(coordinator_id=current_user.id).all()
    my_project_ids = [p.id for p in my_projects]

    this_month_count = sum(
        1 for p in my_projects
        if p.created_at.year == today.year and p.created_at.month == today.month
    )

    doc_staff_users = User.query.filter_by(role='documents', is_active=True).all()
    doc_analytics   = []
    for staff in doc_staff_users:
        assigned    = [p for p in my_projects if p.doc_staff_id == staff.id]
        completed   = [p for p in assigned if p.status in ['Completed', 'Closed']]
        inprog      = [p for p in assigned if p.status == 'InProgress']
        not_started = [p for p in assigned if p.status == 'InProgress' and len(p.documents) == 0]
        total_docs  = sum(len(get_expected_docs(p.project_type, p.project_subtype, p.loan_subtype)) for p in assigned)
        done_docs   = sum(get_doc_completion(p)[0] for p in assigned)
        doc_analytics.append({
            'staff': staff, 'assigned': len(assigned), 'completed': len(completed),
            'inprog': len(inprog), 'not_started': len(not_started),
            'total_docs': total_docs, 'done_docs': done_docs,
            'doc_pct': int(done_docs / total_docs * 100) if total_docs > 0 else 0,
        })

    stage_order  = ['Lead','Site Visit','Documentation','Onsite Work','Connection','Subsidy','Payment']
    stage_counts = Counter(p.stage for p in my_projects)
    stage_data   = [{'stage': s, 'count': stage_counts.get(s, 0)} for s in stage_order]

    payments_all = Payment.query.filter(
        Payment.project_id.in_(my_project_ids)).all() if my_project_ids else []

    week_labels, week_projects_data, week_payments_k = [], [], []
    for i in range(11, -1, -1):
        wstart = today - timedelta(weeks=i + 1) + timedelta(days=1)
        wend   = today - timedelta(weeks=i)
        week_labels.append(wstart.strftime('%d %b'))
        week_projects_data.append(sum(1 for p in my_projects if wstart <= p.created_at.date() <= wend))
        week_payments_k.append(round(sum(float(p.amount) for p in payments_all if wstart <= p.payment_date <= wend) / 1000, 1))

    days_in_month     = calendar.monthrange(today.year, today.month)[1]
    month_day_labels  = [str(d) for d in range(1, days_in_month + 1)]
    month_day_projects = [
        sum(1 for p in my_projects if p.created_at.date() == date(today.year, today.month, d))
        for d in range(1, days_in_month + 1)
    ]

    pay_month_labels, pay_month_values = [], []
    for i in range(5, -1, -1):
        month_date = today.replace(day=1)
        for _ in range(i):
            month_date = (month_date - timedelta(days=1)).replace(day=1)
        pay_month_labels.append(month_date.strftime('%b %Y'))
        pay_month_values.append(round(sum(
            float(p.amount) for p in payments_all
            if p.payment_date.year == month_date.year and p.payment_date.month == month_date.month
        ) / 1000, 1))

    trend_inprog, trend_completed, trend_delayed = [], [], []
    for i in range(11, -1, -1):
        snap = today - timedelta(weeks=i)
        trend_inprog.append(sum(1 for p in my_projects if p.status == 'InProgress' and p.created_at.date() <= snap))
        trend_completed.append(sum(1 for p in my_projects if p.status in ['Completed','Closed'] and p.created_at.date() <= snap))
        trend_delayed.append(sum(1 for p in my_projects if p.status == 'Delayed' and p.created_at.date() <= snap))

    chart_data = {
        'week_labels': week_labels, 'week_projects': week_projects_data,
        'week_payments_k': week_payments_k, 'month_day_labels': month_day_labels,
        'month_day_projects': month_day_projects, 'pay_month_labels': pay_month_labels,
        'pay_month_values': pay_month_values,
        'staff_names': [a['staff'].full_name.split()[0] for a in doc_analytics],
        'staff_done': [a['done_docs'] for a in doc_analytics],
        'staff_pending': [a['total_docs'] - a['done_docs'] for a in doc_analytics],
        'trend_inprog': trend_inprog, 'trend_completed': trend_completed,
        'trend_delayed': trend_delayed,
        'stage_labels': [s['stage'] for s in stage_data if s['count'] > 0],
        'stage_counts': [s['count'] for s in stage_data if s['count'] > 0],
    }

    return render_template('coordinator_analytics.html',
        my_projects=my_projects, doc_analytics=doc_analytics,
        unassigned_projects=[p for p in my_projects if not p.doc_staff_id],
        stage_data=stage_data, this_month_count=this_month_count, chart_data=chart_data)


# ─────────────────────────────────────────────────────────────────────────────
# PAYMENTS
# ─────────────────────────────────────────────────────────────────────────────

@app.route('/projects/<int:pid>/add_payment', methods=['POST'])
@login_required
@roles_required('admin', 'payments')
@limiter.limit('30 per minute')
def add_payment(pid):
    proj   = Project.query.get_or_404(pid)
    if proj.stage == 'Lead':
        flash('Payments cannot be recorded while the project is still in Lead stage.', 'danger')
        return redirect(url_for('project_detail', pid=pid))

    amount = _safe_float(request.form.get('amount'))
    if amount <= 0:
        flash('Please enter a valid payment amount greater than zero.', 'danger')
        return redirect(url_for('project_detail', pid=pid))
    source = request.form.get('payment_source', 'Customer')

    if proj.total_amount and float(proj.total_amount) > 0:
        is_bank_loan = (source == 'Bank' and proj.project_type == 'Loan')
        if not is_bank_loan:
            if float(proj.collected_amount or 0) >= float(proj.total_receivable):
                flash('This project is fully paid. No further payments can be recorded.', 'danger')
                return redirect(url_for('project_detail', pid=pid))
            # Overpayment is no longer blocked — any excess beyond total_receivable
            # is captured below as a PaymentExcess record for settlement.

    instalment = None
    if source == 'Bank':
        instalment = request.form.get('instalment')
        existing   = [p.instalment for p in proj.payments if p.payment_source == 'Bank']
        if instalment in existing:
            flash(f'{instalment} bank payment already recorded for this project.', 'danger')
            return redirect(url_for('project_detail', pid=pid))

    pay = Payment(
        project_id=pid, amount=amount,
        payment_type=request.form['payment_type'],
        payment_source=source, instalment=instalment,
        payment_date=date.fromisoformat(request.form['payment_date']),
        reference_no=_clean(request.form.get('reference_no', ''), 80) or None,
        received_by=current_user.id,
        notes=_clean(request.form.get('notes', ''), 500),
    )
    db.session.add(pay)

    # ── Apply payment, capping collected_amount at total_receivable.
    #    Anything beyond that becomes a settleable PaymentExcess. ─────────
    prior_collected = float(proj.collected_amount or 0)
    receivable      = float(proj.total_receivable)
    new_total_raw   = prior_collected + amount

    if receivable > 0 and new_total_raw > receivable:
        proj.collected_amount = receivable
        excess_amount = round(new_total_raw - receivable, 2)
        db.session.add(PaymentExcess(
            project_id=pid, amount=excess_amount, source=source,
            detected_date=pay.payment_date,
            notes=f'Auto-detected on {"Bank" if instalment else "Customer"} payment of '
                  f'₹{amount:,.0f} (ref: {pay.reference_no or "—"}).',
            recorded_by=current_user.id,
        ))
        for u in User.query.filter_by(role='payments', is_active=True).all():
            create_notification(u.id, pid,
                f'{proj.project_code} — {proj.customer.name}: Payment of ₹{excess_amount:,.0f} '
                f'received over the contract amount. Needs settlement.', 'task')
        log_action(pid, f'Excess payment detected: ₹{excess_amount:,.0f}', new_val='Pending')
    else:
        proj.collected_amount = new_total_raw

    log_action(pid, f"{'Bank' if instalment else 'Customer'} payment recorded: ₹{amount:,.0f}", new_val=str(amount))

    if source == 'Bank' and instalment == 'First':
        notify_onsite_team(pid,
            f'Loan work {proj.project_code} — {proj.customer.name} '
            f'({proj.inverter_capacity_kw} kW): First bank payment of ₹{amount:,.0f} received. ', 'task')

    auto_advance_stage(proj)
    db.session.commit()
    flash(f'Payment of ₹{amount:,.0f} recorded.', 'success')
    return redirect(url_for('project_detail', pid=pid))
@app.route('/projects/<int:pid>/payments/<int:pay_id>/edit', methods=['POST'])
@login_required
@roles_required('admin', 'payments')
def edit_payment(pid, pay_id):
    pay  = Payment.query.get_or_404(pay_id)
    proj = Project.query.get_or_404(pid)

    old_amount = float(pay.amount)
    new_amount = _safe_float(request.form.get('amount'))

    if new_amount <= 0:
        flash('Amount must be greater than 0.', 'danger')
        return redirect(url_for('project_detail', pid=pid))

    pay.amount       = new_amount
    pay.payment_type = request.form.get('payment_type', pay.payment_type)
    pay.payment_date = date.fromisoformat(request.form['payment_date']) if request.form.get('payment_date') else pay.payment_date
    pay.reference_no = _clean(request.form.get('reference_no', ''), 80) or None
    pay.notes        = _clean(request.form.get('notes', ''), 500)

    db.session.flush()  # ensure pay.amount is visible to proj.payments before reconciling

    log_action(pid, f'Payment edited: ₹{old_amount:,.0f} → ₹{new_amount:,.0f}',
               old_val=str(old_amount), new_val=str(new_amount))

    _reconcile_excess_after_payment_change(proj)
    auto_advance_stage(proj)
    db.session.commit()
    flash(f'Payment updated to ₹{new_amount:,.0f}.', 'success')
    return redirect(url_for('project_detail', pid=pid))


@app.route('/projects/<int:pid>/payments/<int:pay_id>/delete', methods=['POST'])
@login_required
@roles_required('admin', 'payments')
def delete_payment(pid, pay_id):
    pay  = Payment.query.get_or_404(pay_id)
    proj = Project.query.get_or_404(pid)

    amount = float(pay.amount)
    db.session.delete(pay)
    db.session.flush()  # ensure pay is gone from proj.payments before reconciling

    log_action(pid, f'Payment deleted: ₹{amount:,.0f} ({pay.payment_type})',
               old_val=str(amount), new_val='Deleted')

    _reconcile_excess_after_payment_change(proj)
    auto_advance_stage(proj)
    db.session.commit()
    flash(f'Payment of ₹{amount:,.0f} deleted.', 'warning')
    return redirect(url_for('project_detail', pid=pid))
@app.route('/projects/<int:pid>/loan_details/edit', methods=['POST'])
@login_required
@roles_required('admin')
def edit_loan_details_inline(pid):
    proj = Project.query.get_or_404(pid)
    ld   = proj.loan_detail or LoanDetail(project_id=pid)
    ld.bank_name   = _clean(request.form.get('bank_name', ''), 120) or None
    ld.loan_amount = _safe_float(request.form.get('loan_amount')) or None
    ld.notes       = _clean(request.form.get('notes', ''), 300) or None
    ld.updated_by  = current_user.id
    if not ld.id:
        db.session.add(ld)
    log_action(pid, 'Loan details edited inline',
               new_val=f'{ld.bank_name}, ₹{float(ld.loan_amount or 0):,.0f}')
    db.session.commit()
    flash('Loan details updated.', 'success')
    return redirect(url_for('project_detail', pid=pid))
@app.route('/projects/<int:pid>/bank_advance', methods=['POST'])
@login_required
@roles_required('admin', 'payments')
def save_bank_advance(pid):
    proj = Project.query.get_or_404(pid)
    if proj.project_type != 'Loan':
        flash('Bank advance only applies to Loan projects.', 'danger')
        return redirect(url_for('project_detail', pid=pid))

    adv = proj.bank_advance or CompanyBankAdvance(project_id=pid)
    adv.amount      = _safe_float(request.form.get('amount'))
    adv.paid_date   = date.fromisoformat(request.form['paid_date']) if request.form.get('paid_date') else None
    adv.notes       = _clean(request.form.get('notes', ''), 300) or None
    adv.recorded_by = current_user.id

    if not adv.id:
        db.session.add(adv)

    log_action(pid, f'Company bank advance recorded: ₹{float(adv.amount):,.0f}',
               new_val=str(adv.amount))

    for u in User.query.filter_by(role='payments', is_active=True).all():
        if u.id != current_user.id:
            create_notification(u.id, pid,
                f'{proj.project_code} — {proj.customer.name}: Company paid ₹{float(adv.amount):,.0f} '
                f'as initial bank advance. To be recovered when bank releases payment.', 'task')

    db.session.commit()
    flash(f'Bank advance of ₹{float(adv.amount):,.0f} recorded.', 'success')
    return redirect(url_for('project_detail', pid=pid))


@app.route('/projects/<int:pid>/bank_advance/recover', methods=['POST'])
@login_required
@roles_required('admin', 'payments')
def recover_bank_advance(pid):
    adv = CompanyBankAdvance.query.filter_by(project_id=pid).first_or_404()
    adv.recovered          = True
    adv.recovery_date      = date.fromisoformat(request.form['recovery_date']) if request.form.get('recovery_date') else date.today()
    adv.recovery_method    = request.form.get('recovery_method') or None
    adv.recovery_reference = _clean(request.form.get('recovery_reference', ''), 100) or None
    adv.recovery_notes     = _clean(request.form.get('recovery_notes', ''), 300) or None

    log_action(pid, f'Company bank advance recovered: ₹{float(adv.amount):,.0f}',
               new_val='Recovered')
    db.session.commit()
    flash(f'Bank advance of ₹{float(adv.amount):,.0f} marked as recovered.', 'success')
    return redirect(url_for('project_detail', pid=pid))
@app.route('/payments')
@login_required
@roles_required('admin', 'payments','director')
def payments_dashboard():
    page     = request.args.get('page', 1, type=int)
    pay_page = request.args.get('pay_page', 1, type=int)
    per_page = 15
    active_ids      = db.session.query(Project.id).filter(Project.status.notin_(['Cancelled','OnHold'])).subquery()
    total_collected = float(db.session.query(db.func.sum(Payment.amount)).filter(Payment.project_id.in_(active_ids)).scalar() or 0)
    total_value     = float(db.session.query(db.func.sum(Project.total_amount)).filter(Project.status.notin_(['Cancelled','OnHold'])).scalar() or 0)
    recent_payments = Payment.query.order_by(Payment.created_at.desc()).paginate(page=pay_page, per_page=per_page, error_out=False)
    pending_projs   = Project.query.filter(Project.status.notin_(['Closed','Cancelled','OnHold'])).order_by(Project.updated_at.desc()).paginate(page=page, per_page=per_page, error_out=False)
    return render_template('payments.html',
        total_collected=total_collected, total_pending=total_value - total_collected,
        total_value=total_value, recent_payments=recent_payments,
        pending_projs=pending_projs, page=page, pay_page=pay_page)
@app.route('/projects/<int:pid>/waive_balance', methods=['POST'])
@login_required
@roles_required('admin', 'payments')
def waive_balance(pid):
    proj = Project.query.get_or_404(pid)
    if proj.stage == 'Lead':
        flash('Cannot write off balance before payment stage.', 'danger')
        return redirect(url_for('project_detail', pid=pid))

    amount = _safe_float(request.form.get('amount'))
    reason = _clean(request.form.get('reason', ''), 500)

    if amount <= 0:
        flash('Please enter a valid write-off amount greater than zero.', 'danger')
        return redirect(url_for('project_detail', pid=pid))
    if not reason:
        flash('Please provide a reason for the write-off.', 'danger')
        return redirect(url_for('project_detail', pid=pid))
    if amount > proj.pending_amount + 0.01:
        flash(f'Cannot waive ₹{amount:,.0f} — only ₹{proj.pending_amount:,.0f} is pending.', 'danger')
        return redirect(url_for('project_detail', pid=pid))

    db.session.add(PaymentWaiver(
        project_id=pid, amount=amount, reason=reason,
        waived_by=current_user.id, waived_date=date.today(),
    ))
    log_action(pid, f'Balance written off: ₹{amount:,.0f}. Reason: {reason}', new_val=str(amount))

    for u in User.query.filter_by(role='payments', is_active=True).all():
        if u.id != current_user.id:
            create_notification(u.id, pid,
                f'{proj.project_code} — {proj.customer.name}: ₹{amount:,.0f} written off by '
                f'{current_user.full_name}. Reason: {reason}', 'info')

    db.session.flush()
    auto_advance_stage(proj)
    db.session.commit()
    flash(f'₹{amount:,.0f} written off. New pending balance: ₹{proj.pending_amount:,.0f}.', 'success')
    return redirect(url_for('project_detail', pid=pid))


@app.route('/projects/<int:pid>/waive_balance/<int:wid>/delete', methods=['POST'])
@login_required
@roles_required('admin')
def delete_waiver(pid, wid):
    w = PaymentWaiver.query.get_or_404(wid)
    amount = float(w.amount)
    db.session.delete(w)
    log_action(pid, f'Balance write-off reversed: ₹{amount:,.0f}', old_val=str(amount), new_val='Reversed')
    db.session.commit()
    flash(f'Write-off of ₹{amount:,.0f} reversed.', 'warning')
    return redirect(url_for('project_detail', pid=pid))

@app.route('/payments/pending_approvals')
@login_required
@roles_required('admin', 'onsite','director')
def pending_approvals():
    cards = JobCard.query.filter_by(status='PendingApproval').order_by(JobCard.closed_at.desc()).all()
    return render_template('pending_approvals.html', cards=cards)


# ─────────────────────────────────────────────────────────────────────────────
# DOCUMENTS
# ─────────────────────────────────────────────────────────────────────────────

@app.route('/projects/<int:pid>/documents', methods=['GET', 'POST'])
@login_required
def documents(pid):
    proj   = Project.query.get_or_404(pid)
    stages = get_document_stages()
    if current_user.role in ('documents','documents_k') and proj.doc_staff_id != current_user.id:
        flash('This project is not assigned to you.', 'danger')
        return redirect(url_for('dashboard'))

    if request.method == 'POST':
        done_before, expected = get_doc_completion(proj)
        was_complete = (expected > 0 and done_before == expected)
        if current_user.role != 'admin' and was_complete:
            flash('All documents are completed and locked.', 'danger')
            return redirect(url_for('documents', pid=pid))

        doc_type = request.form['doc_type']
        status   = request.form.get('status', 'Pending')
        received_date_str = request.form.get('received_date')
        received_date = (
        date.fromisoformat(received_date_str) if received_date_str else date.today()
        )
        existing = Document.query.filter_by(project_id=pid, doc_type=doc_type).first()
        if existing:
            existing.status = status
            if status != 'Pending':
                existing.received_date = received_date
            if request.form.get('notes') is not None:
                existing.notes = _clean(request.form.get('notes', ''), 500)
            log_action(pid, f'Document updated: {doc_type}', new_val=status)
        else:
            db.session.add(Document(
                project_id=pid, doc_type=doc_type, status=status,
                received_date=received_date if status != 'Pending' else None,
                notes=_clean(request.form.get('notes', ''), 500),
            ))
            log_action(pid, f'Document received: {doc_type}', new_val=status)

        if doc_type == 'Feasibility Receipt' and status in ('Received', 'Completed'):
            if not Notification.query.filter_by(project_id=pid, notif_type='task').filter(
                    Notification.message.like('%Structure work%')).first():
                notify_onsite_team(pid,
                    f'Feasibility done for {proj.project_code} — {proj.customer.name}. Start structure work.',
                    'task')
                log_action(pid, 'Onsite team notified: structure work', new_val='Notified')

        if doc_type == 'KSEB Connection' and status in ('Received', 'Completed'):
            if not AppInstallation.query.filter_by(project_id=pid).first():
                db.session.add(AppInstallation(project_id=pid, status='Pending', scheduled_date=date.today()))
                log_action(pid, 'KSEB connection complete → App Installation queued', new_val='Pending')
                for u in User.query.filter_by(role='service', is_active=True).all():
                    create_notification(u.id, pid,
                    f'KSEB connection done for {proj.project_code} — {proj.customer.name} '
                    f'({proj.inverter_capacity_kw} kW). Schedule app installation.', 'task')

    
        if current_user.role == 'admin' and was_complete and proj.doc_staff_id:
            create_notification(proj.doc_staff_id, pid,
                f'{proj.project_code} - {proj.customer.name}: Admin ({current_user.full_name}) '
                f'edited document "{doc_type}" → {status} after completion.', 'info')

        db.session.flush()
        auto_advance_stage(proj)
        db.session.commit()
        flash(f'{doc_type} — {status}.', 'success')
        return redirect(url_for('documents', pid=pid))

    return render_template('documents.html', proj=proj, stages=stages)

@app.route('/works_status')
@login_required
@roles_required('admin', 'documents','office','documents_k','director')
def works_status():
    from sqlalchemy.orm import joinedload

    projects = (Project.query
        .options(
            joinedload(Project.customer),
            joinedload(Project.documents),
            joinedload(Project.expenses),
            joinedload(Project.coordinator),
            joinedload(Project.doc_staff),
        )
        .filter(Project.status.notin_(['Cancelled', 'Closed', 'Completed']))
        .order_by(Project.updated_at.desc())
        .all())

    if current_user.role in ('documents','documents_k'):
        projects = [p for p in projects if p.doc_staff_id == current_user.id]

    KEY_DOCS = [
        'Feasibility Receipt',
        'MNRE',
        'CD Payment Receipt',
        'KSEB Connection',
        'Warranty Card',
    ]

    DONE_STATUSES = {'Received', 'Sent', 'Completed'}

    rows = []
    for p in projects:
        doc_map = {d.doc_type: d.status for d in p.documents}
        rows.append({
            'project':    p,
            'doc_map':    doc_map,
        })

    # Summary counts for cards at top
    def _done(doc_name):
        return sum(
            1 for r in rows
            if r['doc_map'].get(doc_name) in DONE_STATUSES
        )

    dcr_rows = [r for r in rows if r['project'].project_subtype == 'DCR']

    summary = {
        'total':    len(rows),
        'feas':     _done('Feasibility Receipt'),
        'mnre':     sum(1 for r in dcr_rows if r['doc_map'].get('MNRE') in DONE_STATUSES),
        'mnre_tot': len(dcr_rows),
        'cd':       _done('CD Payment Receipt'),
        'kseb':     _done('KSEB Connection'),
        'warr':     _done('Warranty Card'),
    }

    return render_template('works_status.html',
                           rows=rows,
                           key_docs=KEY_DOCS,
                           summary=summary,
                           done_statuses=DONE_STATUSES)

@app.route('/service_management')
@login_required
@roles_required('admin', 'onsite', 'coordinator','director','service')
def service_management():
    refresh_service_statuses()

    has_service = (db.session.query(Project.id)
        .join(ServiceRecord, ServiceRecord.project_id == Project.id)
        .filter(Project.status.notin_(['Cancelled']))
        .distinct()
        .subquery())

    projects = (Project.query
    .options(
        joinedload(Project.customer),
        joinedload(Project.coordinator),
        joinedload(Project.payments),
        joinedload(Project.subsidy),
    )
    .filter(Project.id.in_(has_service))
    .order_by(cast(Project.project_code, Integer))
    .all())

    # Exclude OnHold too, not just Cancelled — a held project's service
    # schedule shouldn't read as "actionable" even if pending_amount is 0.
    projects = [p for p in projects
            if p.status not in ('Cancelled', 'OnHold')
            and p.work_category != 'Outside'                  
            and (p.pending_amount <= 0 or p.status == 'Closed')]
    proj_data = []
    for p in projects:
        records   = sorted(p.service_records, key=lambda r: r.visit_number)
        annotated = _annotate_service_records(records)
        done    = sum(1 for r in records if r.status == 'Completed')
        over    = sum(1 for r in records if r.status == 'Overdue')
        due     = sum(1 for r in records if r.status == 'Due')
        skipped = sum(1 for r in records if r.status == 'Skipped')
        total   = len(records)
        pct     = int(done / total * 100) if total else 0
        next_v  = next((r for r in records if r.status not in ('Completed', 'Skipped')), None)
        next_locked = next(
            (a['locked'] for a in annotated if a['rec'] is next_v), False
        ) if next_v else False

        proj_data.append({
            'project':     p,
            'records':     records,
            'annotated':   annotated,
            'done':        done,
            'over':        over,
            'due':         due,
            'skipped':     skipped,
            'total':       total,
            'pct':         pct,
            'next':        next_v,
            'next_locked': next_locked,
        })

    all_records = [r for pd in proj_data for r in pd['records']]
    stats = {
        'projects':  len(proj_data),
        'overdue':   sum(1 for r in all_records if r.status == 'Overdue'),
        'due':       sum(1 for r in all_records if r.status == 'Due'),
        'completed': sum(1 for r in all_records if r.status == 'Completed'),
        'upcoming':  sum(1 for r in all_records if r.status == 'Upcoming'),
        'skipped':   sum(1 for r in all_records if r.status == 'Skipped'),
    }

    return render_template('service_management.html',
                           proj_data=proj_data,
                           stats=stats,
                           today=date.today())

@app.route('/projects/<int:pid>/connection_details', methods=['POST'])
@login_required
@roles_required('admin', 'documents','office','documents_k')
def update_connection_details(pid):
    proj = Project.query.get_or_404(pid)
    cd   = proj.connection_details or ConnectionDetails(project_id=pid)

    cd.connection_type = request.form.get('connection_type') or None
    cd.category=request.form.get('category') or None
    cd.consumer_number = _clean(request.form.get('consumer_number', ''), 50) or None
    cd.kseb_section = _clean(request.form.get('kseb_section', ''), 100) or None
    ownership_needed         = 'ownership_change_needed' in request.form
    cd.ownership_change_needed = ownership_needed
    cd.ownership_change_status = (
        request.form.get('ownership_change_status', 'Pending')
        if ownership_needed else 'Not Required'
    )

    load_needed              = 'load_clearance_needed' in request.form
    cd.load_clearance_needed = load_needed
    cd.load_clearance_status = (
        request.form.get('load_clearance_status', 'Pending')
        if load_needed else 'Not Required'
    )

    cd.notes      = _clean(request.form.get('notes', ''), 500)
    cd.updated_by = current_user.id

    if not cd.id:
        db.session.add(cd)

    log_action(pid, 'Connection details updated',
               new_val=f'{cd.connection_type}, OC:{cd.ownership_change_status}, LC:{cd.load_clearance_status}')

    # Notify coordinator if something is pending
    pending_items = []
    if ownership_needed and cd.ownership_change_status not in ('Completed', 'Not Required'):
        pending_items.append('Ownership Change')
    if load_needed and cd.load_clearance_status not in ('Completed', 'Not Required'):
        pending_items.append('Load Clearance')
    if pending_items and proj.coordinator_id:
        create_notification(proj.coordinator_id, pid,
            f'{proj.project_code} — {proj.customer.name}: Connection details updated. '
            f'Pending: {", ".join(pending_items)}.', 'info')

    db.session.commit()
    flash('Connection details updated.', 'success')
    return redirect(url_for('documents', pid=pid))
@app.route('/projects/<int:pid>/loan_details', methods=['POST'])
@login_required
@roles_required('admin', 'documents','office','documents_k')
def update_loan_details(pid):
    proj = Project.query.get_or_404(pid)
    if proj.project_type != 'Loan':
        flash('Loan details can only be added to Loan projects.', 'danger')
        return redirect(url_for('documents', pid=pid))
    ld = proj.loan_detail or LoanDetail(project_id=pid)
    ld.bank_name   = _clean(request.form.get('bank_name', ''), 120) or None
    ld.loan_amount = _safe_float(request.form.get('loan_amount')) or None
    ld.notes       = _clean(request.form.get('notes', ''), 300) or None
    ld.updated_by  = current_user.id
    if not ld.id:
        db.session.add(ld)
    log_action(pid, 'Loan details updated',
               new_val=f'{ld.bank_name}, ₹{float(ld.loan_amount or 0):,.0f}')
    db.session.commit()
    flash('Loan details updated.', 'success')
    return redirect(url_for('documents', pid=pid))
@app.route('/projects/<int:pid>/panel_details', methods=['POST'])
@login_required
@roles_required('admin', 'documents', 'onsite','office','documents_k','payments')
def update_panel_details(pid):
    proj = Project.query.get_or_404(pid)
    pd_  = proj.panel_details or PanelDetails(project_id=pid)

    pd_.panel_brand                = _clean(request.form.get('panel_brand', ''), 100) or None
    raw_num                        = request.form.get('num_panels', '')
    pd_.num_panels                 = int(raw_num) if raw_num.strip().isdigit() else None
    pd_.panel_serial_numbers       = _clean(request.form.get('panel_serial_numbers', ''), 2000) or None
    pd_.inverter_serial_number     = _clean(request.form.get('inverter_serial_number', ''), 100) or None
    pd_.net_meter_serial_number    = _clean(request.form.get('net_meter_serial_number', ''), 100) or None
    pd_.energy_meter_serial_number = _clean(request.form.get('energy_meter_serial_number', ''), 100) or None
    pd_.inverter_brand = _clean(request.form.get('inverter_brand', ''), 100) or None
    pd_.app_id          = _clean(request.form.get('app_id', ''), 100) or None          # ← NEW
    pd_.app_password    = _clean(request.form.get('app_password', ''), 100) or None    # ← NEW
    pd_.notes                      = _clean(request.form.get('notes', ''), 1000) or None
    pd_.updated_by                 = current_user.id

    if not pd_.id:
        db.session.add(pd_)

    log_action(pid, 'Panel details updated',
               new_val=f'{pd_.panel_brand}, {pd_.num_panels} panels')
    db.session.commit()
    flash('Panel details saved.', 'success')
    return redirect(url_for('project_detail', pid=pid))
@app.route('/projects/<int:pid>/documents/batch', methods=['POST'])
@login_required
def batch_documents(pid):
    proj = Project.query.get_or_404(pid)
    if current_user.role in ('documents','documents_k') and proj.doc_staff_id != current_user.id:
        flash('This project is not assigned to you.', 'danger')
        return redirect(url_for('dashboard'))
    

    doc_types = request.form.getlist('doc_types')
    status    = request.form.get('status', 'Received')

    for doc_type in doc_types:
        existing = Document.query.filter_by(project_id=pid, doc_type=doc_type).first()
        if existing:
            existing.status = status
            if status != 'Pending':
                existing.received_date = date.today()
        else:
            db.session.add(Document(
                project_id=pid, doc_type=doc_type, status=status,
                received_date=date.today() if status != 'Pending' else None,
            ))
        log_action(pid, f'Batch document update: {doc_type}', new_val=status)

    db.session.flush()
    auto_advance_stage(proj)
    db.session.commit()
    flash(f'{len(doc_types)} documents marked {status}.', 'success')
    return redirect(url_for('documents', pid=pid))
@app.route('/project/<int:pid>/panel-item/<int:item_id>/update', methods=['POST'])
@login_required
def update_panel_item(pid, item_id):
    item = PanelItem.query.get_or_404(item_id)
    item.brand      = request.form['brand']
    item.panel_type = request.form['panel_type']
    item.wattage    = int(request.form.get('wattage') or 0)
    item.quantity   = int(request.form['quantity'])
    db.session.commit()
    return redirect(url_for('onsite_progress', pid=pid))

@app.route('/project/<int:pid>/extra-material/<int:item_id>/update', methods=['POST'])
@login_required
def update_extra_material(pid, item_id):
    mat = ExtraMaterial.query.get_or_404(item_id)
    mat.description    = request.form['description']
    mat.quantity_label = request.form.get('quantity_label', '')
    db.session.commit()
    return redirect(url_for('onsite_progress', pid=pid))

# ─────────────────────────────────────────────────────────────────────────────
# KSEB
# ─────────────────────────────────────────────────────────────────────────────

@app.route('/projects/<int:pid>/kseb', methods=['GET', 'POST'])
@login_required
def kseb(pid):
    proj = Project.query.get_or_404(pid)
    task = proj.kseb_task or KSEBTask(project_id=pid)

    if request.method == 'POST':
        task.stamp_paper     = request.form.get('stamp_paper', task.stamp_paper)
        task.b_class_licence = request.form.get('b_class_licence', task.b_class_licence)
        task.file_sent       = 'file_sent' in request.form
        task.inspection_done = 'inspection_done' in request.form
        task.cd_payment_done = 'cd_payment_done' in request.form
        task.connection_done = 'connection_done' in request.form
        task.meter_available = 'meter_available' in request.form
        task.ae_completed    = 'ae_completed' in request.form
        task.notes           = _clean(request.form.get('notes', ''), 1000)
        if not task.id:
            db.session.add(task)
            db.session.flush()
        log_action(pid, 'KSEB tasks updated')

        if task.connection_done and not AppInstallation.query.filter_by(project_id=pid).first():
            db.session.add(AppInstallation(project_id=pid, status='Pending', scheduled_date=date.today()))
            log_action(pid, 'KSEB connection complete → App Installation queued', new_val='Pending')
            for u in User.query.filter_by(role='service', is_active=True).all():
                create_notification(u.id, pid,
                    f'KSEB connection done for {proj.project_code} — {proj.customer.name} '
                    f'({proj.inverter_capacity_kw} kW). Schedule app installation.', 'task')
        db.session.flush()
        auto_advance_stage(proj)
        db.session.commit()
        flash('KSEB tasks updated.', 'success')
        return redirect(url_for('kseb', pid=pid))

    return render_template('kseb.html', proj=proj, task=task)


# ─────────────────────────────────────────────────────────────────────────────
# WORKERS
# ─────────────────────────────────────────────────────────────────────────────

@app.route('/workers')
@login_required
@roles_required('admin', 'onsite', 'payments','director')
def workers():
    all_workers = Worker.query.options(
        selectinload(Worker.assignments).joinedload(WorkerAssignment.project),
        selectinload(Worker.job_cards).joinedload(JobCard.project),
        selectinload(Worker.job_cards).joinedload(JobCard.closer),
        selectinload(Worker.job_cards).joinedload(JobCard.approver),
        selectinload(Worker.advances).joinedload(WorkerAdvance.project),
        selectinload(Worker.ledger_entries).joinedload(WorkerLedger.recorder),
        selectinload(Worker.weekly_payments),
    ).order_by(Worker.is_active.desc(), Worker.name).all()
    return render_template('workers.html',
        workers=all_workers, today=date.today())
@app.route('/workers/new', methods=['GET', 'POST'])
@login_required
@roles_required('admin', 'onsite')
def new_worker():
    if request.method == 'POST':
        try:
            name  = _clean(request.form.get('name', ''), 100)
            phone = _clean(request.form.get('phone', ''), 20) or None
            skill = _clean(request.form.get('skill', ''), 80) or None
            rate  = _safe_float(request.form.get('rate_per_day', 0))

            if not name:
                flash('Worker name is required.', 'danger')
                return redirect(url_for('new_worker'))

            worker = Worker(name=name, phone=phone, skill=skill, rate_per_day=rate)
            db.session.add(worker)
            db.session.commit()
            # log_admin_action('CREATE_WORKER', target=name)
            flash(f'Worker {name} added successfully.', 'success')
            return redirect(url_for('workers'))
        except Exception as e:
            db.session.rollback()
            flash('An error occurred while adding the worker. Please try again.', 'danger')

    return render_template('new_worker.html')

@app.route('/workers/<int:worker_id>/edit', methods=['GET', 'POST'])
@login_required
@roles_required('admin', 'onsite')
def edit_worker(worker_id):
    worker = Worker.query.get_or_404(worker_id)

    if request.method == 'POST':
        try:
            name  = _clean(request.form.get('name', ''), 100)
            phone = _clean(request.form.get('phone', ''), 20) or None
            skill = _clean(request.form.get('skill', ''), 80) or None
            rate  = _safe_float(request.form.get('rate_per_day', 0))

            if not name:
                flash('Worker name is required.', 'danger')
                return redirect(url_for('edit_worker', worker_id=worker_id))

            old_rate = float(worker.rate_per_day or 0)
            worker.name         = name
            worker.phone        = phone
            worker.skill        = skill
            worker.rate_per_day = rate
            db.session.commit()

            # log_admin_action('EDIT_WORKER', target=name,
                            #  detail=f'rate ₹{old_rate:,.0f}→₹{rate:,.0f}' if abs(old_rate - rate) > 0.01 else '')
            flash(f'Worker {name} updated successfully.', 'success')
            return redirect(url_for('workers'))
        except Exception as e:
            db.session.rollback()
            flash('An error occurred while updating the worker. Please try again.', 'danger')

    return render_template('edit_worker.html', worker=worker)


@app.route('/workers/<int:worker_id>/delete', methods=['POST'])
@login_required
@roles_required('admin')
def delete_worker(worker_id):
    try:
        worker = Worker.query.get_or_404(worker_id)

        active = WorkerAssignment.query.filter_by(
            worker_id=worker_id
        ).filter(WorkerAssignment.status.in_(['Assigned', 'Active'])).first()

        if active:
            flash(
                f'Cannot delete {worker.name} — they have active assignments. '
                f'Complete or unassign them first.', 'danger'
            )
            return redirect(url_for('workers'))

        open_cards = JobCard.query.filter_by(
            worker_id=worker_id
        ).filter(JobCard.status.in_(['Open', 'PendingApproval', 'Approved'])).first()

        if open_cards:
            flash(
                f'Cannot delete {worker.name} — they have open or approved job cards.', 'danger'
            )
            return redirect(url_for('workers'))

        worker.is_active = False
        db.session.commit()
        # log_admin_action('DELETE_WORKER', target=worker.name)
        flash(f'Worker {worker.name} has been deactivated.', 'warning')

    except Exception as e:
        db.session.rollback()
        print("DELETE ERROR:",e)
        flash('An error occurred. Please try again.', 'danger')

    return redirect(url_for('workers'))

@app.route('/workers/<int:worker_id>/restore', methods=['POST'])
@login_required
@roles_required('admin')
def restore_worker(worker_id):
    try:
        worker = Worker.query.get_or_404(worker_id)
        worker.is_active = True
        db.session.commit()
        # log_admin_action('RESTORE_WORKER', target=worker.name)
        flash(f'Worker {worker.name} has been restored.', 'success')
    except Exception as e:
        db.session.rollback()
        flash('An error occurred. Please try again.', 'danger')
    return redirect(url_for('workers'))

@app.route('/projects/<int:pid>/assign_worker', methods=['POST'])
@login_required
@roles_required('admin', 'onsite')
def assign_worker(pid):
    start    = request.form.get('start_date')
    end      = request.form.get('end_date')
    days     = max((date.fromisoformat(end) - date.fromisoformat(start)).days + 1, 0) if start and end else 0
    phase    = request.form.get('work_phase', 'Structure')

    worker_id = int(request.form['worker_id'])
    wa = WorkerAssignment(project_id=pid, worker_id=worker_id,
        start_date=date.fromisoformat(start) if start else None,
        end_date=date.fromisoformat(end) if end else None,
        days_worked=days, work_phase=phase, status='Assigned')
    db.session.add(wa)

    assigned_worker = Worker.query.get(worker_id)
    db.session.add(JobCard(project_id=pid, worker_id=worker_id,
        work_phase=phase, rate_per_day=assigned_worker.rate_per_day, status='Open'))

    log_action(pid, f'Worker assigned: ID {assigned_worker.name} ({phase})', new_val='Assigned')
    db.session.commit()
    flash('Worker assigned.', 'success')
    return redirect(url_for('onsite_progress', pid=pid))


@app.route('/projects/<int:pid>/unassign_worker/<int:aid>', methods=['POST'])
@login_required
@roles_required('admin', 'onsite')
def unassign_worker(pid, aid):
    wa = WorkerAssignment.query.get_or_404(aid)
    if wa.status == 'Paid':
        flash('Cannot unassign a worker who has already been paid.', 'danger')
        return redirect(url_for('project_detail', pid=pid))
    for card in JobCard.query.filter_by(project_id=pid, worker_id=wa.worker_id,
            work_phase=wa.work_phase).filter(JobCard.status.in_(['Open','PendingApproval'])).all():
        card.status = 'Voided'
    worker_name = wa.worker.name
    phase       = wa.work_phase
    log_action(pid, f'Worker unassigned: {worker_name} — {phase}', old_val=wa.status)
    db.session.delete(wa)
    db.session.commit()
    flash(f'{worker_name} unassigned from {phase} phase.', 'success')
    return redirect(url_for('onsite_board', pid=pid))


@app.route('/projects/<int:pid>/update_assignment/<int:aid>', methods=['POST'])
@login_required
@roles_required('admin', 'onsite')
def update_assignment(pid, aid):
    wa        = WorkerAssignment.query.get_or_404(aid)
    new_phase = request.form.get('work_phase', wa.work_phase)
    wa.work_phase = new_phase

    new_status = request.form.get('status', wa.status)
    if new_status == 'Paid' and wa.status != 'Completed':
        flash('Worker must be marked Completed before payment can be recorded.', 'danger')
        return redirect(url_for('project_detail', pid=pid))
    wa.status = new_status

    if new_wid := request.form.get('worker_id'):
        wa.worker_id = int(new_wid)

    log_action(pid, f'Worker assignment updated: {wa.worker.name}', new_val=wa.status)
    db.session.commit()
    flash('Assignment updated.', 'success')
    return redirect(url_for('project_detail', pid=pid))


@app.route('/worker/payment/<int:pay_id>/delete', methods=['POST'])
@login_required
@roles_required('admin')
def delete_worker_payment(pay_id):
    pay    = WorkerWeeklyPayment.query.get_or_404(pay_id)
    worker = pay.worker
    reason = _clean(request.form.get('reason', ''), 500)
    for proj in pay.projects:
        for wa in proj.assignments:
            if wa.worker_id == worker.id and wa.status == 'Paid':
                wa.status = 'Completed'
        log_action(proj.id,
            f'Worker payment voided: {worker.name} — week {pay.week_start}. '
            f'Reason: {reason or "Not provided"}', old_val=str(pay.amount))
    db.session.delete(pay)
    db.session.commit()
    flash(f'Payment for {worker.name} (week {pay.week_start}) voided.', 'warning')
    return redirect(url_for('workers'))


@app.route('/projects/<int:pid>/onsite_progress', methods=['GET', 'POST'])
@login_required
@roles_required('admin', 'onsite')
def onsite_progress(pid):
    proj     = Project.query.get_or_404(pid)
    progress = proj.onsite_progress or OnsiteProgress(project_id=pid)
    all_workers=Worker.query.order_by(Worker.name).all()
    all_stock_items = StockItem.query.filter_by(is_active=True).order_by(
        StockItem.category, StockItem.name).all() 
    delivered_stock_txns = (StockTransaction.query               
        .filter_by(project_id=pid, source='OnsiteDelivery')
        .order_by(StockTransaction.created_at.desc())
        .all())
    if request.method == 'POST':
        
        new_struct  = request.form.get('structure_work_status', progress.structure_work_status)
        new_install = request.form.get('installation_status', '')
        new_elec    = request.form.get('electrical_status', '')

        old_struct  = progress.structure_work_status
        old_install = progress.installation_status or 'NotStarted'
        old_elec    = progress.electrical_status   or 'NotStarted'
        progress.important_notes=request.form.get('important_notes',progress.important_notes or '')
        progress.structure_work_status = new_struct
        progress.structure_notes       = _clean(request.form.get('structure_notes', ''), 500)

        progress.installation_status = new_install if new_install and new_install != 'None' else (progress.installation_status or 'NotStarted')
        progress.installation_notes  = _clean(request.form.get('installation_notes', ''), 500)

        progress.electrical_status = new_elec if new_elec and new_elec != 'None' else (progress.electrical_status or 'NotStarted')
        progress.electrical_notes  = _clean(request.form.get('electrical_notes', ''), 500)

        progress.updated_by = current_user.id 
        progress.materials_ordered        = request.form.get('materials_ordered') == 'on'
        progress.materials_ordered_date   = date.fromisoformat(request.form['materials_ordered_date']) if request.form.get('materials_ordered_date') else (progress.materials_ordered_date or date.today())
        progress.materials_delivered      = request.form.get('materials_delivered') == 'on'
        delivered_date_raw = (
            request.form.get('materials_delivered_date_override')
            or request.form.get('materials_delivered_date')
        )
        progress.materials_delivered_date = date.fromisoformat(delivered_date_raw) if delivered_date_raw else (progress.materials_delivered_date or date.today())

        for field, col in [
            ('structure_start_date','structure_start_date'), ('structure_end_date','structure_end_date'),
            ('installation_start_date','installation_start_date'), ('installation_end_date','installation_end_date'),
            ('electrical_start_date','electrical_start_date'), ('electrical_end_date','electrical_end_date'),
        ]:
            if val := request.form.get(field):
                setattr(progress, col, date.fromisoformat(val))

        if not progress.id:
            db.session.add(progress)

        changes = []
        if old_struct  != progress.structure_work_status: changes.append(f'Structure: {old_struct} → {progress.structure_work_status}')
        if old_install != progress.installation_status:   changes.append(f'Installation: {old_install} → {progress.installation_status}')
        if old_elec    != progress.electrical_status:     changes.append(f'Electrical: {old_elec} → {progress.electrical_status}')
        log_action(pid, ('Onsite: ' + ', '.join(changes)) if changes else 'Onsite dates/notes updated')

        if changes:
            for u in User.query.filter_by(role='payments', is_active=True).all():
                create_notification(u.id, pid,
                    f'{proj.project_code} — {proj.customer.name}: Onsite work update — '
                    + ', '.join(changes), 'info')

        db.session.flush()
        auto_advance_stage(proj)
        db.session.commit()
        flash('Onsite progress updated.', 'success')
        return redirect(url_for('onsite_progress', pid=pid))

    return render_template('onsite_progress.html', proj=proj, progress=progress,all_workers=all_workers,all_stock_items=all_stock_items,delivered_stock_txns=delivered_stock_txns,today=date.today())

@app.route('/projects/<int:pid>/onsite_log', methods=['POST'])
@login_required
@roles_required('admin', 'onsite')
def add_onsite_log(pid):
    note  = _clean(request.form.get('note', ''), 500)
    phase = request.form.get('work_phase', 'Structure')
    if not note:
        flash('Note cannot be empty.', 'danger')
        return redirect(url_for('onsite_progress', pid=pid))
    db.session.add(OnsiteLog(project_id=pid, log_date=date.today(),
        work_phase=phase, note=note, logged_by=current_user.id))
    log_action(pid, f'Onsite log added: {phase}', new_val=note[:60])
    db.session.commit()
    flash('Log entry added.', 'success')
    return redirect(url_for('onsite_progress', pid=pid))


@app.route('/job_card/<int:card_id>/close', methods=['POST'])
@login_required
@roles_required('admin', 'onsite')
def close_job_card(card_id):
    card = JobCard.query.get_or_404(card_id)
    if card.status != 'Open':
        flash('This job card is not open.', 'danger')
        return redirect(url_for('workers'))
    card.actual_days  = Decimal(request.form['actual_days'])
    card.final_amount = Decimal(request.form['final_amount'])
    card.description  = _clean(request.form.get('description', ''), 300) or None
    card.status       = 'Approved'
    card.closed_at    = datetime.utcnow()
    card.closed_by    = current_user.id
    card.approved_at = datetime.utcnow()
    card.approved_by = current_user.id
 
    last = WorkerLedger.query.filter_by(worker_id=card.worker_id).order_by(WorkerLedger.id.desc()).first()
    prev = float(last.balance_after) if last else 0.0
    db.session.add(WorkerLedger(
        worker_id=card.worker_id, entry_date=date.today(),
        entry_type='Earning', direction='Credit', amount=card.final_amount,
        reference_type='JobCard', reference_id=card.id,
        balance_after=prev + float(card.final_amount), recorded_by=current_user.id,
        notes=f'{card.work_phase} — {card.project.project_code}',
    ))
    log_action(card.project_id,
        f'Job card closed: {card.worker.name} — {card.work_phase} ({card.actual_days} days)',
        new_val=str(card.final_amount))
    db.session.commit()
    flash(f'Job card closed — ₹{card.final_amount:,.0f} added to {card.worker.name}\'s balance.', 'success')
    return redirect(url_for('workers'))


# @app.route('/job_card/<int:card_id>/approve', methods=['POST'])
# @login_required
# @roles_required('admin', 'onsite')
# def approve_job_card(card_id):
#     card = JobCard.query.get_or_404(card_id)
#     if card.status != 'PendingApproval':
#         flash('Card is not pending approval.', 'danger')
#         return redirect(url_for('workers'))
#     card.status      = 'Approved'
#     card.approved_at = datetime.utcnow()
#     card.approved_by = current_user.id

#     last = WorkerLedger.query.filter_by(worker_id=card.worker_id).order_by(WorkerLedger.id.desc()).first()
#     prev = float(last.balance_after) if last else 0.0
#     db.session.add(WorkerLedger(
#         worker_id=card.worker_id, entry_date=date.today(),
#         entry_type='Earning', direction='Credit', amount=card.final_amount,
#         reference_type='JobCard', reference_id=card.id,
#         balance_after=prev + float(card.final_amount), recorded_by=current_user.id,
#         notes=f'{card.work_phase} — {card.project.project_code}',
#     ))
#     log_action(card.project_id, f'Job card approved: {card.worker.name} — {card.work_phase}',
#                new_val=str(card.final_amount))
#     db.session.commit()
#     flash(f'Job card approved. ₹{card.final_amount:,.0f} added to {card.worker.name}\'s balance.', 'success')
#     return redirect(url_for('workers'))


@app.route('/job_card/<int:card_id>/void', methods=['POST'])
@login_required
@roles_required('admin', 'onsite')
def void_job_card(card_id):
    card = JobCard.query.get_or_404(card_id)
    if card.status == 'Paid':
        flash('Paid job cards cannot be voided.', 'danger')
        return redirect(url_for('workers'))
    old = card.status
    card.status = 'Voided'
    log_action(card.project_id, f'Job card voided: {card.worker.name} — {card.work_phase}',
               old_val=old, new_val='Voided')
    db.session.commit()
    flash(f'Job card for {card.worker.name} voided.', 'warning')
    return redirect(url_for('workers'))


@app.route('/worker/<int:worker_id>/advance', methods=['POST'])
@login_required
@roles_required('admin', 'onsite')
def give_advance(worker_id):
    worker = Worker.query.get_or_404(worker_id)
    amount = Decimal(request.form['amount'])
    if amount <= 0:
        flash('Amount must be greater than 0.', 'danger')
        return redirect(url_for('workers'))
    given_date = datetime.strptime(request.form['given_date'], '%Y-%m-%d').date()
    notes      = _clean(request.form.get('notes', ''), 300) or None
    advance = WorkerAdvance(worker_id=worker_id, amount=amount,
        given_date=given_date, given_by=current_user.id, notes=notes, status='Outstanding')
    db.session.add(advance)
    db.session.flush()

    last = WorkerLedger.query.filter_by(worker_id=worker_id).order_by(WorkerLedger.id.desc()).first()
    prev = float(last.balance_after) if last else 0.0
    db.session.add(WorkerLedger(
        worker_id=worker_id, entry_date=given_date, entry_type='Advance', direction='Debit',
        amount=amount, reference_type='Advance', reference_id=advance.id,
        balance_after=prev - float(amount), recorded_by=current_user.id,
        notes=notes or f'Advance given on {given_date}',
    ))
    db.session.commit()
    flash(f'Advance of ₹{amount:,.0f} recorded for {worker.name}.', 'success')
    return redirect(url_for('workers'))


@app.route('/worker/<int:worker_id>/settle', methods=['POST'])
@login_required
@roles_required('admin', 'onsite')
def settle_worker(worker_id):
    worker = Worker.query.get_or_404(worker_id)
    last   = WorkerLedger.query.filter_by(worker_id=worker_id).order_by(WorkerLedger.id.desc()).first()
    current_balance = float(last.balance_after) if last else 0.0
    if current_balance <= 0:
        flash('No balance to settle.', 'warning')
        return redirect(url_for('workers'))

    amount_paid = Decimal(request.form['amount'])
    notes       = _clean(request.form.get('notes', ''), 300) or None
    new_balance = current_balance - float(amount_paid)

    db.session.add(WorkerLedger(
        worker_id=worker_id, entry_date=date.today(), entry_type='Settlement',
        direction='Debit', amount=amount_paid, reference_type='Manual',
        balance_after=new_balance, recorded_by=current_user.id, notes=notes,
    ))

    for advance in WorkerAdvance.query.filter_by(worker_id=worker_id).filter(
            WorkerAdvance.status.in_(['Outstanding','PartiallyRecovered'])).all():
        still_owed = float(advance.amount) - float(advance.recovered_amount)
        if still_owed <= 0:
            continue
        recover = min(still_owed, float(amount_paid))
        advance.recovered_amount = float(advance.recovered_amount) + recover
        advance.status = 'Cleared' if float(advance.recovered_amount) >= float(advance.amount) else 'PartiallyRecovered'

    for jc in JobCard.query.filter_by(worker_id=worker_id, status='Approved').all():
        jc.status = 'Paid'
        log_action(jc.project_id, f'Worker paid (settlement): {worker.name} — {jc.work_phase}',
                   new_val=str(jc.final_amount))

    db.session.commit()
    flash(f'Settlement of ₹{amount_paid:,.0f} recorded for {worker.name}. Remaining balance: ₹{new_balance:,.0f}.', 'success')
    return redirect(url_for('workers'))


@app.route('/ledger/<int:entry_id>/void', methods=['POST'])
@login_required
@roles_required('admin')
def void_ledger_entry(entry_id):
    entry  = WorkerLedger.query.get_or_404(entry_id)
    worker = entry.worker
    last   = WorkerLedger.query.filter_by(worker_id=worker.id).order_by(WorkerLedger.id.desc()).first()
    prev   = float(last.balance_after) if last else 0.0
    reverse_dir = 'Debit' if entry.direction == 'Credit' else 'Credit'
    new_balance = prev - float(entry.amount) if entry.direction == 'Credit' else prev + float(entry.amount)
    db.session.add(WorkerLedger(
        worker_id=worker.id, entry_date=date.today(), entry_type=entry.entry_type,
        direction=reverse_dir, amount=entry.amount, reference_type='Reversal',
        reference_id=entry.id, balance_after=new_balance, recorded_by=current_user.id,
        notes=f'Reversal of entry #{entry.id}',
    ))
    db.session.commit()
    flash(f'Ledger entry #{entry.id} reversed. Balance updated to ₹{new_balance:,.0f}.', 'warning')
    return redirect(url_for('workers'))


@app.route('/advance/<int:advance_id>/recover', methods=['POST'])
@login_required
@roles_required('admin', 'onsite')
def recover_advance(advance_id):
    advance    = WorkerAdvance.query.get_or_404(advance_id)
    still_owed = float(advance.amount) - float(advance.recovered_amount)
    recover    = min(_safe_float(request.form.get('recover_amount'), still_owed), still_owed)
    advance.recovered_amount = float(advance.recovered_amount) + recover
    advance.status = 'Cleared' if float(advance.recovered_amount) >= float(advance.amount) else 'PartiallyRecovered'
    db.session.commit()
    flash(f'₹{recover:,.0f} marked as recovered from advance.', 'success')
    return redirect(url_for('workers'))
@app.route('/project/<int:pid>/panel-item/add', methods=['POST'])
@login_required
def add_panel_item(pid):
    proj = Project.query.get_or_404(pid)
    item = PanelItem(
        project_id=pid,
        brand=request.form['brand'],
        panel_type=request.form['panel_type'],
        wattage=int(request.form.get('wattage') or 0),
        quantity=int(request.form['quantity']),
    )
    db.session.add(item)
    db.session.flush()
    stock_item_id = request.form.get('stock_item_id', type=int)
    if stock_item_id:
        record_stock_txn(stock_item_id, 'Out', item.quantity, source='PanelItem',
                      project_id=pid, reference_id=item.id,
                      notes=f'Used on {proj.project_code}')
    db.session.add(item); db.session.commit()
    return redirect(url_for('onsite_progress', pid=pid))

@app.route('/project/<int:pid>/panel-item/<int:item_id>/delete', methods=['POST'])
@login_required
def delete_panel_item(pid, item_id):
    item = PanelItem.query.get_or_404(item_id)
    db.session.delete(item); db.session.commit()
    return redirect(url_for('onsite_progress', pid=pid))

@app.route('/project/<int:pid>/extra-material/add', methods=['POST'])
@login_required
def add_extra_material(pid):
    mat = ExtraMaterial(
        project_id=pid,
        description=request.form['description'],
        quantity_label=request.form.get('quantity_label',''),
    )
    db.session.add(mat); db.session.commit()
    return redirect(url_for('onsite_progress', pid=pid))

@app.route('/project/<int:pid>/extra-material/<int:item_id>/delete', methods=['POST'])
@login_required
def delete_extra_material(pid, item_id):
    mat = ExtraMaterial.query.get_or_404(item_id)
    db.session.delete(mat); db.session.commit()
    return redirect(url_for('onsite_progress', pid=pid))

@app.route('/projects/<int:pid>/materials/dispatch/<int:mid>', methods=['POST'])
@login_required
@roles_required('admin', 'onsite')
def dispatch_material(pid, mid):
    proj     = Project.query.get_or_404(pid)
    material = Material.query.get_or_404(mid)
    material.dispatch_status = 'Dispatched'
    material.dispatch_date   = date.today()
    log_action(pid, f'Material dispatched: {material.item_name}', new_val='Dispatched')
    stock_item_id = request.form.get('stock_item_id', type=int)
    if stock_item_id:
        record_stock_txn(stock_item_id, 'Out', float(material.quantity or 0), source='Material',
                      project_id=pid, reference_id=material.id,
                      notes=f'Dispatched on {proj.project_code}')
    db.session.commit()
    flash(f'{material.item_name} marked as dispatched.', 'success')
    return redirect(url_for('project_detail', pid=pid))

# ── Onsite: mark materials delivered AND deduct stock in one step ─────────────
@app.route('/projects/<int:pid>/onsite_progress/deliver_stock', methods=['POST'])
@login_required
@roles_required('admin', 'onsite','stocks')
def deliver_stock_to_site(pid):
    proj     = Project.query.get_or_404(pid)
    progress = proj.onsite_progress or OnsiteProgress(project_id=pid)

    item_ids = request.form.getlist('stock_item_id')
    qtys     = request.form.getlist('stock_qty')

    deducted = []
    for iid, qty_raw in zip(item_ids, qtys):
        qty = _safe_float(qty_raw)
        if not iid or qty <= 0:
            continue
        item = StockItem.query.get(int(iid))
        if not item:
            continue
        if qty > float(item.current_qty or 0):
            flash(f'Not enough stock for {item.display_name} — only '
                  f'{item.current_qty} {item.unit} available.', 'danger')
            return redirect(url_for('onsite_progress', pid=pid))
        record_stock_txn(item.id, 'Out', qty, source='OnsiteDelivery',
                          project_id=pid, notes=f'Delivered on site — {proj.project_code}')
        deducted.append(f'{item.display_name} ({qty} {item.unit})')

    progress.materials_delivered      = True
    progress.materials_delivered_date = (
        date.fromisoformat(request.form['delivered_date'])
        if request.form.get('delivered_date') else date.today()
    )
    if not progress.materials_ordered:
        progress.materials_ordered      = True
        progress.materials_ordered_date = progress.materials_ordered_date or progress.materials_delivered_date
    if not progress.id:
        db.session.add(progress)

    log_action(pid, 'Materials delivered to site' +
               (f': {", ".join(deducted)}' if deducted else ' (no stock items selected)'),
               new_val='Delivered')
    db.session.flush()
    auto_advance_stage(proj)
    db.session.commit()
    flash('Materials marked as delivered' +
          (' and deducted from stock.' if deducted else '.'), 'success')
    return redirect(url_for('onsite_progress', pid=pid))


# ── Stocks: manual purchase from a distributor (not tied to a project) ────────
@app.route('/stock/purchase', methods=['GET', 'POST'])
@login_required
@roles_required('admin', 'stocks', 'onsite')
@limiter.limit('30 per minute')
def stock_purchase():
    if request.method == 'GET':
        items = StockItem.query.filter_by(is_active=True).order_by(StockItem.category, StockItem.name).all()
        return render_template('stock_purchase_new.html', items=items)

    item_id  = request.form.get('stock_item_id', type=int)
    qty      = _safe_float(request.form.get('quantity'))
    vendor   = _clean(request.form.get('vendor', ''), 100)
    invoice  = _clean(request.form.get('invoice_no', ''), 60)
    notes_in = _clean(request.form.get('notes', ''), 200)

    if not item_id or qty <= 0:
        flash('Please select an item and enter a valid quantity.', 'danger')
        return redirect(url_for('stock_purchase'))

    item = StockItem.query.get_or_404(item_id)
    parts = [p for p in [f'Vendor: {vendor}' if vendor else '',
                          f'Invoice: {invoice}' if invoice else '',
                          notes_in] if p]
    record_stock_txn(item_id, 'In', qty, source='Purchase', notes=' | '.join(parts) or None)
    db.session.commit()
    flash(f'Purchase of {qty} {item.unit} "{item.display_name}" recorded'
          + (f' from {vendor}.' if vendor else '.'), 'success')
    return redirect(url_for('stock_dashboard'))
# ── Stocks: sale/dispatch to another distributor or buyer (stock out) ─────────
@app.route('/stock/sale', methods=['GET', 'POST'])
@login_required
@roles_required('admin', 'stocks','onsite')
@limiter.limit('30 per minute')
def stock_sale():
    if request.method == 'GET':
        items = StockItem.query.filter_by(is_active=True).order_by(StockItem.category, StockItem.name).all()
        return render_template('stock_sale_new.html', items=items)

    item_id   = request.form.get('stock_item_id', type=int)
    qty       = _safe_float(request.form.get('quantity'))
    sale_type = request.form.get('sale_type', '')
    buyer     = _clean(request.form.get('buyer', ''), 100)
    invoice   = _clean(request.form.get('invoice_no', ''), 60)
    notes_in  = _clean(request.form.get('notes', ''), 200)

    if not item_id or qty <= 0:
        flash('Please select an item and enter a valid quantity.', 'danger')
        return redirect(url_for('stock_sale'))

    if sale_type not in ('B2B', 'B2C'):
        flash('Please select a sale type (B2B or B2C).', 'danger')
        return redirect(url_for('stock_sale'))

    item = StockItem.query.get_or_404(item_id)
    if qty > float(item.current_qty or 0):
        flash(f'Cannot sell {qty} {item.unit} — only {item.current_qty} {item.unit} in stock.', 'danger')
        return redirect(url_for('stock_sale'))

    parts = [p for p in [f'Buyer: {buyer}' if buyer else '',
                          f'Invoice: {invoice}' if invoice else '',
                          notes_in] if p]
    record_stock_txn(item_id, 'Out', qty, source='Sale', sale_type=sale_type,
                      notes=' | '.join(parts) or None)
    db.session.commit()
    flash(f'{sale_type} sale of {qty} {item.unit} "{item.display_name}" recorded'
          + (f' to {buyer}.' if buyer else '.'), 'success')
    return redirect(url_for('stock_dashboard'))
@app.route('/product_replacements')
@login_required
@roles_required('admin', 'stocks','director')
def product_replacements():
    show_archived = request.args.get('archived') == '1'
    cat_search = _clean(request.args.get('cat_q', ''), 100)

    cat_q = ProductCategory.query.filter_by(is_active=True)
    if cat_search:
        cat_q = cat_q.filter(ProductCategory.name.ilike(f'%{cat_search}%'))
    categories = cat_q.order_by(ProductCategory.name).all()

    archived_categories = []
    if show_archived:
        arch_q = ProductCategory.query.filter_by(is_active=False)
        if cat_search:
            arch_q = arch_q.filter(ProductCategory.name.ilike(f'%{cat_search}%'))
        archived_categories = arch_q.order_by(ProductCategory.name).all()

    recent = (ProductReplacement.query
              .order_by(ProductReplacement.created_at.desc())
              .limit(10).all())
    pending_total = ProductReplacement.query.filter_by(status='Pending').count()
    return render_template('product_replacements.html',
        categories=categories, archived_categories=archived_categories,
        show_archived=show_archived, recent=recent, pending_total=pending_total,
        cat_search=cat_search)
@app.route('/product_replacements/category/new', methods=['POST'])
@login_required
@roles_required('admin', 'stocks', 'director')
def new_product_category():
    name = _clean(request.form.get('name', ''), 120)
    if not name:
        flash('Category name is required.', 'danger')
        return redirect(url_for('product_replacements'))
    if ProductCategory.query.filter(db.func.lower(ProductCategory.name) == name.lower()).first():
        flash(f'Category "{name}" already exists.', 'warning')
        return redirect(url_for('product_replacements'))
    cat = ProductCategory(name=name, created_by=current_user.id)
    db.session.add(cat)
    db.session.commit()
    flash(f'Category "{name}" created.', 'success')
    return redirect(url_for('product_replacement_category', cid=cat.id))
@app.route('/product_replacements/category/<int:cid>/edit', methods=['POST'])
@login_required
@roles_required('admin', 'stocks', 'director')
def edit_product_category(cid):
    cat = ProductCategory.query.get_or_404(cid)
    name = _clean(request.form.get('name', ''), 120)
    if not name:
        flash('Category name is required.', 'danger')
        return redirect(request.referrer or url_for('product_replacements'))
    dup = ProductCategory.query.filter(
        db.func.lower(ProductCategory.name) == name.lower(),
        ProductCategory.id != cid
    ).first()
    if dup:
        flash(f'Category "{name}" already exists.', 'warning')
        return redirect(request.referrer or url_for('product_replacements'))
    old_name = cat.name
    cat.name = name
    db.session.commit()
    flash(f'Category renamed from "{old_name}" to "{name}".', 'success')
    return redirect(request.referrer or url_for('product_replacements'))

@app.route('/product_replacements/category/<int:cid>/delete', methods=['POST'])
@login_required
@roles_required('admin','stocks')
def delete_product_category(cid):
    cat = ProductCategory.query.get_or_404(cid)
    cat.is_active = False
    db.session.commit()
    flash(f'Category "{cat.name}" archived.', 'warning')
    return redirect(url_for('product_replacements'))




@app.route('/product_replacements/category/<int:cid>')
@login_required
@roles_required('admin', 'stocks', 'director')
def product_replacement_category(cid):
    cat = ProductCategory.query.get_or_404(cid)
    search = _clean(request.args.get('q', ''), 100)
    status_filter = request.args.get('status', '')
    q = ProductReplacement.query.filter_by(category_id=cid)
    if search:
        q = q.filter(
            ProductReplacement.complaint_id.ilike(f'%{search}%') |
            ProductReplacement.order_id.ilike(f'%{search}%')
        )
    if status_filter in ('Pending', 'Cleared'):
        q = q.filter(ProductReplacement.status == status_filter)
    records = q.order_by(ProductReplacement.replacement_date.desc()).all()
    return render_template('product_replacement_category.html',
        cat=cat, records=records, search=search, status_filter=status_filter, today=date.today())

@app.route('/product_replacements/category/<int:cid>/add', methods=['POST'])
@login_required
@roles_required('admin', 'stocks', 'director')
@limiter.limit('30 per minute')
def add_product_replacement(cid):
    cat = ProductCategory.query.get_or_404(cid)

    complaint_id    = _clean(request.form.get('complaint_id', ''), 60)
    serial_number   = _clean(request.form.get('serial_number', ''), 100)
    purchaser_name  = _clean(request.form.get('purchaser_name', ''), 120)
    purchaser_phone = _clean(request.form.get('purchaser_phone', ''), 20) or None
    order_id        = _clean(request.form.get('order_id', ''), 20) or None
    new_serial      = _clean(request.form.get('new_serial_number', ''), 100) or None
    notes           = _clean(request.form.get('notes', ''), 1000) or None
    rep_date_raw    = request.form.get('replacement_date')

    if not complaint_id or not serial_number or not purchaser_name:
        flash('Complaint ID, serial number and purchaser name are required.', 'danger')
        return redirect(url_for('product_replacement_category', cid=cid))

    rec = ProductReplacement(
        category_id       = cid,
        complaint_id      = complaint_id,
        serial_number     = serial_number,
        replacement_date  = date.fromisoformat(rep_date_raw) if rep_date_raw else date.today(),
        purchaser_name    = purchaser_name,
        purchaser_phone   = purchaser_phone,
        order_id           = order_id,
        new_serial_number = new_serial,
        notes             = notes,
        recorded_by       = current_user.id,
    )
    db.session.add(rec)
    db.session.commit()
    flash(f'Replacement logged for {serial_number}.', 'success')
    return redirect(url_for('product_replacement_category', cid=cid))


@app.route('/product_replacements/<int:rid>/edit', methods=['POST'])
@login_required
@roles_required('admin', 'stocks', 'director')
def edit_product_replacement(rid):
    rec = ProductReplacement.query.get_or_404(rid)

    rec.complaint_id      = _clean(request.form.get('complaint_id', rec.complaint_id), 60)
    rec.serial_number     = _clean(request.form.get('serial_number', rec.serial_number), 100)
    rec.purchaser_name    = _clean(request.form.get('purchaser_name', rec.purchaser_name), 120)
    rec.purchaser_phone   = _clean(request.form.get('purchaser_phone', ''), 20) or None
    rec.order_id           = _clean(request.form.get('order_id', ''), 20) or None
    rec.new_serial_number = _clean(request.form.get('new_serial_number', ''), 100) or None
    rec.notes             = _clean(request.form.get('notes', ''), 1000) or None
    rep_date_raw = request.form.get('replacement_date')
    if rep_date_raw:
        rec.replacement_date = date.fromisoformat(rep_date_raw)

    db.session.commit()
    flash('Replacement record updated.', 'success')
    return redirect(url_for('product_replacement_category', cid=rec.category_id))
@app.route('/product_replacements/<int:rid>/clear', methods=['POST'])
@login_required
@roles_required('admin', 'stocks', 'director')
def mark_replacement_cleared(rid):
    rec = ProductReplacement.query.get_or_404(rid)
    if rec.status == 'Cleared':
        flash('This case is already marked cleared.', 'warning')
        return redirect(url_for('product_replacement_category', cid=rec.category_id))
    rec.status       = 'Cleared'
    rec.cleared_date = date.today()
    rec.cleared_by   = current_user.id
    db.session.commit()
    flash(f'Case for {rec.serial_number} marked as cleared.', 'success')
    return redirect(url_for('product_replacement_category', cid=rec.category_id))


@app.route('/product_replacements/<int:rid>/reopen', methods=['POST'])
@login_required
@roles_required('admin', 'stocks', 'director')
def reopen_replacement_case(rid):
    rec = ProductReplacement.query.get_or_404(rid)
    if rec.status == 'Pending':
        flash('This case is already pending.', 'warning')
        return redirect(url_for('product_replacement_category', cid=rec.category_id))
    rec.status       = 'Pending'
    rec.cleared_date = None
    rec.cleared_by   = None
    db.session.commit()
    flash(f'Case for {rec.serial_number} reopened.', 'warning')
    return redirect(url_for('product_replacement_category', cid=rec.category_id))
@app.route('/product_replacements/category/<int:cid>/restore', methods=['POST'])
@login_required
@roles_required('admin')
def restore_product_category(cid):
    cat = ProductCategory.query.get_or_404(cid)
    cat.is_active = True
    db.session.commit()
    flash(f'Category "{cat.name}" restored.', 'success')
    return redirect(url_for('product_replacements', archived=1))

@app.route('/product_replacements/<int:rid>/delete', methods=['POST'])
@login_required
@roles_required('admin','stocks')
def delete_product_replacement(rid):
    rec = ProductReplacement.query.get_or_404(rid)
    cid = rec.category_id
    db.session.delete(rec)
    db.session.commit()
    flash('Replacement record deleted.', 'warning')
    return redirect(url_for('product_replacement_category', cid=cid))
@app.route('/projects/<int:pid>/materials/update/<int:mid>', methods=['POST'])
@login_required
@roles_required('admin', 'onsite')
def update_material(pid, mid):
    material = Material.query.get_or_404(mid)
    material.item_name = _clean(request.form.get('item_name', material.item_name), 100)
    material.quantity  = _safe_float(request.form.get('quantity') or 0)
    if request.form.get('dispatch_date'):
        material.dispatch_date = date.fromisoformat(request.form['dispatch_date'])
    material.dispatch_status = request.form.get('dispatch_status', material.dispatch_status)
    if material.dispatch_status == 'Delivered' and not material.received_date:
        material.received_date = date.today()
    material.notes = _clean(request.form.get('notes', ''), 500)
    log_action(pid, f'Material updated: {material.item_name}', new_val=material.dispatch_status)
    db.session.commit()
    flash(f'{material.item_name} updated.', 'success')
    return redirect(url_for('onsite_progress', pid=pid))


@app.route('/projects/<int:pid>/materials/bulk_update', methods=['POST'])
@login_required
@roles_required('admin', 'onsite')
def bulk_update_materials(pid):
    proj     = Project.query.get_or_404(pid)
    dispatch_status = request.form.get('dispatch_status')
    dispatch_date   = request.form.get('dispatch_date')
    for m in proj.materials:
        m.dispatch_status = dispatch_status
        if dispatch_date:
            m.dispatch_date = date.fromisoformat(dispatch_date)
        if dispatch_status == 'Delivered' and not m.received_date:
            m.received_date = date.today()
    log_action(pid, f'All materials bulk updated to {dispatch_status}')
    db.session.commit()
    flash(f'All materials marked as {dispatch_status}.', 'success')
    return redirect(url_for('onsite_progress', pid=pid))


@app.route('/projects/<int:pid>/materials/add', methods=['POST'])
@login_required
@roles_required('admin', 'onsite')
def add_material(pid):
    proj     = Project.query.get_or_404(pid)
    material = Material(project_id=pid,
        item_name=_clean(request.form['item_name'], 100),
        quantity=_safe_float(request.form.get('quantity') or 0),
        dispatch_status='Pending',
        notes=_clean(request.form.get('notes', ''), 500))
    db.session.add(material)
    log_action(pid, f'Material added: {material.item_name}', new_val='Pending')
    db.session.commit()
    flash(f'{material.item_name} added.', 'success')
    return redirect(url_for('onsite_progress', pid=pid))

#stocks

@app.route('/stock')
@login_required
@roles_required('admin', 'stocks', 'onsite','director')
def stock_dashboard():
    selected_cat = request.args.get('category', '')
    categories = sorted({
        c[0] for c in db.session.query(StockItem.category)
        .filter_by(is_active=True).distinct()
    })
    q = StockItem.query.filter_by(is_active=True)
    if selected_cat:
        q = q.filter_by(category=selected_cat)
    items = q.order_by(StockItem.category, StockItem.brand, StockItem.name).all()
    low_stock = [i for i in items if i.is_low]
    recent_txns = (StockTransaction.query
                   .order_by(StockTransaction.created_at.desc())
                   .limit(15).all())
    recent_site_deliveries = (StockTransaction.query               # ← new
                   .filter_by(source='OnsiteDelivery')
                   .order_by(StockTransaction.created_at.desc())
                   .limit(10).all())
    return render_template('stock_dashboard.html', items=items, categories=categories,
                           selected_cat=selected_cat, low_stock=low_stock,
                           recent_txns=recent_txns,recent_site_deliveries=recent_site_deliveries)
 
 
@app.route('/stock/items')
@login_required
@roles_required('admin', 'stocks','onsite','director')
def manage_stock_items():
    items = StockItem.query.order_by(
        StockItem.is_active.desc(), StockItem.category, StockItem.brand).all()
    categories = sorted({i.category for i in items})
    brands     = sorted({i.brand for i in items if i.brand})
    units      = sorted({i.unit for i in items if i.unit})
    return render_template('stock_items.html', items=items, categories=categories,
                           brands=brands, units=units)
 
 
@app.route('/stock/items/new', methods=['GET', 'POST'])
@login_required
@roles_required('admin', 'stocks','onsite')
@limiter.limit('30 per minute')
def new_stock_item():
    if request.method == 'GET':
        categories = [r[0] for r in db.session.query(StockItem.category).distinct() if r[0]]
        brands     = [r[0] for r in db.session.query(StockItem.brand).distinct() if r[0]]
        units      = [r[0] for r in db.session.query(StockItem.unit).distinct() if r[0]]
        return render_template('stock_item_new.html',
                                categories=categories, brands=brands, units=units)

    category   = _clean(request.form.get('category', ''), 50)
    brand      = _clean(request.form.get('brand', ''), 60) or None
    subtype    = _clean(request.form.get('subtype', ''), 40) or None
    model_name = _clean(request.form.get('model_name', ''), 60) or None
    unit       = _clean(request.form.get('unit', ''), 20) or 'pcs'
    reorder_level = _safe_float(request.form.get('reorder_level', 0))

    raw_names = request.form.get('names', '')
    # dedupe within the pasted list itself, preserve order
    seen = set()
    names = []
    for line in raw_names.splitlines():
        n = _clean(line, 120)
        if n and n.lower() not in seen:
            seen.add(n.lower())
            names.append(n)

    if not category or not names:
        flash('Category and at least one item name are required.', 'danger')
        return redirect(url_for('new_stock_item'))

    created, skipped = [], []
    for name in names:
        dup = StockItem.query.filter_by(
            category=category, brand=brand, model_name=model_name, name=name,
        ).first()
        if dup:
            skipped.append(dup.display_name)
            continue

        item = StockItem(
            category      = category,
            brand         = brand,
            subtype       = subtype,
            model_name    = model_name,
            name          = name,
            unit          = unit,
            reorder_level = reorder_level,
            current_qty   = 0,
            created_by    = current_user.id,
        )
        db.session.add(item)
        created.append(name)

    db.session.commit()

    if created:
        flash(f'{len(created)} item(s) added to the catalog: {", ".join(created)}.', 'success')
    if skipped:
        flash(f'{len(skipped)} item(s) already existed and were skipped: {", ".join(skipped)}.', 'warning')
    if not created and skipped:
        return redirect(url_for('new_stock_item'))

    return redirect(url_for('manage_stock_items'))
@app.route('/stock/items/<int:iid>/edit', methods=['POST'])
@login_required
@roles_required('admin', 'stocks','onsite')
def edit_stock_item(iid):
    item = StockItem.query.get_or_404(iid)
    item.category      = _clean(request.form.get('category', item.category), 50)
    item.brand         = _clean(request.form.get('brand', ''), 60) or None
    item.subtype       = _clean(request.form.get('subtype', ''), 40) or None
    item.model_name    = _clean(request.form.get('model_name', ''), 60) or None
    item.name          = _clean(request.form.get('name', item.name), 120)
    item.unit          = _clean(request.form.get('unit', ''), 20) or 'pcs'
    item.reorder_level = _safe_float(request.form.get('reorder_level', item.reorder_level))
    db.session.commit()
    flash(f'"{item.display_name}" updated.', 'success')
    return redirect(url_for('manage_stock_items'))
 
 
@app.route('/stock/items/<int:iid>/deactivate', methods=['POST'])
@login_required
@roles_required('admin', 'stocks','onsite')
def deactivate_stock_item(iid):
    item = StockItem.query.get_or_404(iid)
    item.is_active = False
    db.session.commit()
    flash(f'"{item.display_name}" deactivated (kept for ledger history).', 'warning')
    return redirect(url_for('manage_stock_items'))
 
 
@app.route('/stock/items/<int:iid>/restore', methods=['POST'])
@login_required
@roles_required('admin', 'stocks','onsite')
def restore_stock_item(iid):
    item = StockItem.query.get_or_404(iid)
    item.is_active = True
    db.session.commit()
    flash(f'"{item.display_name}" restored.', 'success')
    return redirect(url_for('manage_stock_items'))
 
 
@app.route('/stock/adjust', methods=['POST'])
@login_required
@roles_required('admin', 'stocks', 'onsite')
@limiter.limit('30 per minute')
def stock_adjust():
    item_id  = request.form.get('stock_item_id', type=int)
    txn_type = request.form.get('txn_type', 'In')
    qty      = _safe_float(request.form.get('quantity'))
    notes    = _clean(request.form.get('notes', ''), 300)
 
    if not item_id or qty <= 0:
        flash('Please select an item and enter a valid quantity.', 'danger')
        return redirect(url_for('stock_dashboard'))
 
    item = StockItem.query.get_or_404(item_id)
    if txn_type == 'Out' and qty > float(item.current_qty or 0):
        flash(f'Cannot remove {qty} {item.unit} — only {item.current_qty} {item.unit} in stock.', 'danger')
        return redirect(url_for('stock_dashboard'))
 
    record_stock_txn(item_id, txn_type, qty, source='Manual', notes=notes)
    db.session.commit()
    flash(f'{txn_type} of {qty} {item.unit} recorded for "{item.display_name}".', 'success')
    return redirect(url_for('stock_dashboard'))
 
 
@app.route('/stock/items/<int:iid>/ledger')
@login_required
@roles_required('admin', 'stocks','onsite')
def stock_item_ledger(iid):
    item = StockItem.query.get_or_404(iid)
    txns = (StockTransaction.query.filter_by(stock_item_id=iid)
            .order_by(StockTransaction.created_at.desc()).all())
    return render_template('stock_ledger.html', item=item, txns=txns)
 
 
# ── Autocomplete data for the "add item" form (categories/brands/units
#    already in use — the dropdown grows automatically as new ones are added) ──
@app.route('/api/stock_meta')
@login_required
@roles_required('admin', 'stocks','onsite')
def api_stock_meta():
    categories = [c[0] for c in db.session.query(StockItem.category).distinct()]
    brands     = [b[0] for b in db.session.query(StockItem.brand).filter(StockItem.brand.isnot(None)).distinct()]
    units      = [u[0] for u in db.session.query(StockItem.unit).filter(StockItem.unit.isnot(None)).distinct()]
    return jsonify({
        'categories': sorted(categories),
        'brands':     sorted(brands),
        'units':      sorted(units),
    })

# ─────────────────────────────────────────────────────────────────────────────
# SUBSIDY
# ─────────────────────────────────────────────────────────────────────────────

@app.route('/projects/<int:pid>/subsidy', methods=['GET', 'POST'])
@login_required
def subsidy(pid):
    proj = Project.query.get_or_404(pid)
    sub  = proj.subsidy

    if request.method == 'POST':
        if current_user.role not in ['admin', 'payments', 'documents','office','documents_k']:
            flash('Subsidy can be updated only by payments or documents team', 'danger')
            return redirect(url_for('subsidy', pid=pid))

        if sub is None:
            sub = Subsidy(project_id=pid)
            db.session.add(sub)
            db.session.flush()

        exp        = _safe_float(request.form.get('expected_amount') or sub.expected_amount or 78000)
        new_status = request.form.get('status', sub.status)
        rec        = exp if new_status == 'Received' else _safe_float(request.form.get('received_amount') or sub.received_amount or 0)

        if rec > exp:
            flash(f'Received amount (₹{rec:,.0f}) cannot exceed expected amount (₹{exp:,.0f}).', 'danger')
            return redirect(url_for('subsidy', pid=pid))

        sub.status = new_status
        sub.expected_amount = exp
        sub.received_amount = rec
        sub.notes = _clean(request.form.get('notes', ''), 1000)
        if request.form.get('request_date'):
            sub.request_date = date.fromisoformat(request.form['request_date'])

        customer_share = float(sub.customer_share or 0)
        company_share  = float(sub.company_share  or 0)
        if new_status == 'Received':
            customer_share = _safe_float(request.form.get('customer_share'))
            company_share  = _safe_float(request.form.get('company_share'))
            if abs((customer_share + company_share) - rec) > 0.01:
                flash(f'Customer share (₹{customer_share:,.0f}) + Company share (₹{company_share:,.0f}) '
                      f'must equal received amount (₹{rec:,.0f}).', 'danger')
                return redirect(url_for('subsidy', pid=pid))
        sub.customer_share = customer_share
        sub.company_share  = company_share

        if sub.status == 'Redeemed':
            if proj.coordinator_id:
                create_notification(proj.coordinator_id, pid,
                    f'{proj.project_code} — {proj.customer.name}: Subsidy redeemed. Awaiting receipt.', 'info')
            log_action(pid, 'Subsidy updated: Redeemed', new_val=sub.status)
        elif sub.status == 'Received':
            if proj.coordinator_id:
                create_notification(proj.coordinator_id, pid,
                    f'{proj.project_code} — {proj.customer.name}: Subsidy amount received.', 'info')
            log_action(pid, 'Subsidy updated: Received', new_val=sub.status)

        db.session.flush()
        auto_advance_stage(proj)
        db.session.commit()
        flash('Subsidy record updated.', 'success')
        return redirect(url_for('project_detail', pid=pid) + '#subsidy')

    return render_template('project_detail.html', proj=proj, sub=sub)


# ─────────────────────────────────────────────────────────────────────────────
# APP INSTALLATION
# ─────────────────────────────────────────────────────────────────────────────

@app.route('/installations')
@login_required
@roles_required('admin', 'service', 'documents','office','documents_k','director')
def installations():
    today = date.today()
    pending = AppInstallation.query.filter_by(status='Pending').all()
    for i in pending:
        i.days_pending = (today - i.created_at.date()).days

    return render_template('installations.html',
        pending=pending,
        completed=AppInstallation.query.filter_by(status='Completed').all(),
        today=today)
@app.route('/installations/map')
@login_required
@roles_required('admin', 'service', 'documents', 'office', 'documents_k', 'director')
def app_install_map():
    q             = _clean(request.args.get('q', ''), 100)
    radius_km     = request.args.get('radius', 15, type=int)
    status_filter = request.args.get('status', 'Pending')
    if status_filter not in ('Pending', 'Completed', 'All'):
        status_filter = 'Pending'

    from sqlalchemy.orm import joinedload
    all_installs = (AppInstallation.query
    .join(Project)
    .options(...)
    .filter(Project.status != 'Cancelled',
            Project.work_category != 'Outside')                
    .order_by(AppInstallation.id.desc())
    .all())

    if status_filter == 'Pending':
        installs = [i for i in all_installs if i.status != 'Completed']
    elif status_filter == 'Completed':
        installs = [i for i in all_installs if i.status == 'Completed']
    else:
        installs = all_installs

    geo_installs = [i for i in installs if i.project.geo_tag and i.project.geo_tag.has_gps]
    no_geo       = [i for i in installs if not (i.project.geo_tag and i.project.geo_tag.has_gps)]

    tags_json = [{
        'lat':     i.project.geo_tag.latitude,
        'lng':     i.project.geo_tag.longitude,
        'code':    i.project.project_code,
        'name':    i.project.customer.name,
        'address': i.project.customer.full_address or '—',
        'status':  i.status,
        'url':     url_for('project_detail', pid=i.project_id),
    } for i in geo_installs]

    center = (None, None)
    nearby = []
    if q:
        center = geocode_place(q + ', Kerala, India')
        if center[0]:
            for i in geo_installs:
                dist = haversine_km(center[0], center[1],
                                     i.project.geo_tag.latitude, i.project.geo_tag.longitude)
                if dist <= radius_km:
                    nearby.append((i, round(dist, 1)))
            nearby.sort(key=lambda x: x[1])
        else:
            flash(f'Could not locate "{q}".', 'warning')

    return render_template('app_install_map.html',
        installs=geo_installs, no_geo=no_geo, nearby=nearby,
        tags_json=tags_json, search_q=q, center=center, radius_km=radius_km,
        status_filter=status_filter)
@app.route('/service')
@login_required
@roles_required('admin', 'onsite', 'coordinator')
def service_dashboard():
    refresh_service_statuses()
    today = date.today()

    overdue  = ServiceRecord.query.filter_by(status='Overdue').order_by(ServiceRecord.scheduled_date).all()
    due      = ServiceRecord.query.filter_by(status='Due').order_by(ServiceRecord.scheduled_date).all()

    # Only the next pending visit per project
    from sqlalchemy import func
    subq = (db.session.query(
                ServiceRecord.project_id,
                func.min(ServiceRecord.visit_number).label('min_visit')
            )
            .filter(ServiceRecord.status == 'Upcoming')
            .group_by(ServiceRecord.project_id)
            .subquery())

    upcoming = (ServiceRecord.query
                .join(subq, db.and_(
                    ServiceRecord.project_id == subq.c.project_id,
                    ServiceRecord.visit_number == subq.c.min_visit
                ))
                .order_by(ServiceRecord.scheduled_date)
                .limit(20).all())

    recent = (ServiceRecord.query
              .filter_by(status='Completed')
              .order_by(ServiceRecord.completed_date.desc())
              .limit(10).all())

    completed_year = ServiceRecord.query.filter(
        ServiceRecord.status == 'Completed',
        db.extract('year', ServiceRecord.completed_date) == today.year,
    ).count()
    total_active = ServiceRecord.query.filter(
        ServiceRecord.status.in_(['Upcoming', 'Due', 'Overdue'])
    ).count()

    return render_template('service_dashboard.html',
        overdue=overdue, due=due, upcoming=upcoming, recent=recent,
        total_due=len(overdue)+len(due), completed_year=completed_year,
        total_active=total_active, today=today)
 
@app.route('/projects/<int:pid>/service')
@login_required
def project_service(pid):
    refresh_service_statuses()
    proj = Project.query.get_or_404(pid)

    # NEW — this view previously had no status guard at all.
    if proj.status in ('Cancelled', 'OnHold'):
        flash('Service schedule is not active for a Cancelled or On Hold project.', 'warning')
        return redirect(url_for('project_detail', pid=pid))
    if proj.work_category == 'Outside':                       
        flash('Outside-work projects do not have a service schedule.', 'warning')
        return redirect(url_for('project_detail', pid=pid))
    records   = (ServiceRecord.query
               .filter_by(project_id=pid)
               .order_by(ServiceRecord.visit_number).all())
    annotated = _annotate_service_records(records)

    done    = sum(1 for r in records if r.status == 'Completed')
    over    = sum(1 for r in records if r.status == 'Overdue')
    due     = sum(1 for r in records if r.status == 'Due')
    skipped = sum(1 for r in records if r.status == 'Skipped')
    total   = len(records)
    pct     = int(done / total * 100) if total else 0
    next_v  = next((r for r in records if r.status not in ('Completed', 'Skipped')), None)
    next_locked = next(
        (a['locked'] for a in annotated if a['rec'] is next_v), False
    ) if next_v else False

    proj_data = [{
        'project':     proj,
        'records':     records,
        'annotated':   annotated,
        'done':        done,
        'over':        over,
        'due':         due,
        'skipped':     skipped,
        'total':       total,
        'pct':         pct,
        'next':        next_v,
        'next_locked': next_locked,
    }]

    stats = {
        'projects':  1,
        'overdue':   over,
        'due':       due,
        'completed': done,
        'upcoming':  sum(1 for r in records if r.status == 'Upcoming'),
        'skipped':   skipped,
    }

    return render_template('service_management.html',
                           proj_data=proj_data, stats=stats, today=date.today())
@app.route('/service/<int:sid>/complete', methods=['POST'])
@login_required
@roles_required('admin', 'onsite')
def complete_service(sid):
    rec = ServiceRecord.query.get_or_404(sid)

    if rec.status == 'Completed':
        flash('This service visit is already marked complete.', 'warning')
        return redirect(request.referrer or url_for('project_service', pid=rec.project_id))

    if rec.visit_number > 1 and current_user.role != 'admin':
        prev = ServiceRecord.query.filter_by(
            project_id=rec.project_id,
            visit_number=rec.visit_number - 1
        ).first()
        if not prev or prev.status != 'Completed':
            flash('Previous service visit must be completed first.', 'danger')
            return redirect(request.referrer or url_for('project_service', pid=rec.project_id))

    completed_date_str = request.form.get('completed_date', '')
    rec.completed_date = (date.fromisoformat(completed_date_str)
                          if completed_date_str else date.today())

    if rec.completed_date > date.today():
        flash('Completion date cannot be in the future.', 'danger')
        return redirect(request.referrer or url_for('project_service', pid=rec.project_id))

    rec.status         = 'Completed'
    rec.conducted_by   = current_user.id
    rec.panel_cleaning = 'panel_cleaning' in request.form
    rec.notes          = _clean(request.form.get('notes', ''), 1000)

    proj = rec.project
    backfilled_note = ' (backfilled by admin)' if (
        current_user.role == 'admin'
        and rec.completed_date < date.today() - timedelta(days=SERVICE_BACKFILL_FLAG_DAYS)
    ) else ''
    log_action(rec.project_id,
        f'Service visit #{rec.visit_number} completed{backfilled_note} — '
        f'panel cleaning: {"Yes" if rec.panel_cleaning else "No"}',
        new_val='Completed')

    _reschedule_future_visits(rec, rec.completed_date)

    if proj.coordinator_id:
        create_notification(proj.coordinator_id, rec.project_id,
            f'{proj.project_code} — {proj.customer.name}: Service visit #{rec.visit_number} '
            f'completed by {current_user.full_name}.', 'info')

    db.session.commit()
    flash(f'Service visit #{rec.visit_number} marked complete.', 'success')
    return redirect(request.referrer or url_for('project_service', pid=rec.project_id))
@app.route('/service/bulk_complete', methods=['POST'])
@login_required
@roles_required('admin', 'onsite')
def bulk_complete_service():
    ids                 = request.form.getlist('service_ids')
    completed_date_str  = request.form.get('completed_date', '')
    panel_cleaning      = 'panel_cleaning' in request.form
    notes               = _clean(request.form.get('notes', ''), 1000)

    if not ids:
        flash('No visits were selected.', 'warning')
        return redirect(url_for('service_management'))

    completed_date = date.fromisoformat(completed_date_str) if completed_date_str else date.today()
    if completed_date > date.today():
        flash('Completion date cannot be in the future.', 'danger')
        return redirect(url_for('service_management'))

    done, skipped_locked = 0, 0
    for sid in ids:
        rec = ServiceRecord.query.get(int(sid))
        if not rec or rec.status == 'Completed':
            continue
        if rec.visit_number > 1 and current_user.role != 'admin':
            prev = ServiceRecord.query.filter_by(
                project_id=rec.project_id, visit_number=rec.visit_number - 1).first()
            if not prev or prev.status != 'Completed':
                skipped_locked += 1
                continue

        rec.status         = 'Completed'
        rec.completed_date = completed_date
        rec.conducted_by   = current_user.id
        rec.panel_cleaning = panel_cleaning
        if notes:
            rec.notes = notes

        log_action(rec.project_id,
            f'Service visit #{rec.visit_number} completed via bulk action — '
            f'panel cleaning: {"Yes" if panel_cleaning else "No"}', new_val='Completed')
        _reschedule_future_visits(rec, completed_date)
        done += 1

    db.session.commit()
    msg = f'{done} visit(s) marked complete.'
    if skipped_locked:
        msg += f' {skipped_locked} skipped — previous visit not yet completed.'
    flash(msg, 'success' if done else 'warning')
    return redirect(url_for('service_management'))

@app.route('/service/<int:sid>/skip', methods=['POST'])
@login_required
@roles_required('admin')
def skip_service(sid):
    rec        = ServiceRecord.query.get_or_404(sid)
    old_status = rec.status
    reason     = _clean(request.form.get('reason', ''), 500)
    rec.status = 'Skipped'
    rec.notes  = reason
    log_action(rec.project_id,
        f'Service visit #{rec.visit_number} skipped. Reason: {reason or "None"}',
        old_val=old_status, new_val='Skipped')
    db.session.commit()
    flash(f'Service visit #{rec.visit_number} skipped.', 'warning')
    return redirect(request.referrer or url_for('project_service', pid=rec.project_id))
@app.route('/service/<int:sid>/unskip', methods=['POST'])
@login_required
@roles_required('admin')
def unskip_service(sid):
    rec = ServiceRecord.query.get_or_404(sid)
    if rec.status != 'Skipped':
        flash('This visit is not currently skipped.', 'warning')
        return redirect(request.referrer or url_for('project_service', pid=rec.project_id))

    old_reason = rec.notes
    rec.status = 'Upcoming' if rec.scheduled_date >= date.today() else 'Overdue'
    rec.notes  = None

    log_action(rec.project_id,
        f'Service visit #{rec.visit_number} un-skipped (was skipped: {old_reason or "no reason given"})',
        old_val='Skipped', new_val=rec.status)
    db.session.commit()
    flash(f'Visit #{rec.visit_number} restored to {rec.status}.', 'success')
    return redirect(request.referrer or url_for('project_service', pid=rec.project_id))

@app.route('/service/<int:sid>/reschedule', methods=['POST'])
@login_required
@roles_required('admin', 'onsite')
def reschedule_service(sid):
    rec           = ServiceRecord.query.get_or_404(sid)
    new_date_str  = request.form.get('new_date', '')
    if not new_date_str:
        flash('Please provide a new date.', 'danger')
        return redirect(request.referrer or url_for('project_service', pid=rec.project_id))
    old_date           = rec.scheduled_date
    rec.scheduled_date = date.fromisoformat(new_date_str)
    rec.status         = 'Upcoming'
    log_action(rec.project_id, f'Service visit #{rec.visit_number} rescheduled',
               old_val=str(old_date), new_val=str(rec.scheduled_date))
    db.session.commit()
    flash(f'Visit #{rec.visit_number} rescheduled to {rec.scheduled_date.strftime("%d %b %Y")}.', 'success')
    return redirect(request.referrer or url_for('project_service', pid=rec.project_id))
 
@app.route('/api/service_stats')
@login_required
def api_service_stats():
    refresh_service_statuses()
    return jsonify({
        'overdue':         ServiceRecord.query.filter_by(status='Overdue').count(),
        'due':             ServiceRecord.query.filter_by(status='Due').count(),
        'upcoming':        ServiceRecord.query.filter_by(status='Upcoming').count(),
        'completed_total': ServiceRecord.query.filter_by(status='Completed').count(),
    })

@app.route('/projects/<int:pid>/installation', methods=['POST'])
@login_required
@roles_required('admin', 'service')
def update_installation(pid):
    proj    = Project.query.get_or_404(pid)
    install = proj.app_install or AppInstallation(project_id=pid)
    install.status         = request.form.get('status', install.status)
    install.scheduled_date = date.fromisoformat(request.form['scheduled_date']) if request.form.get('scheduled_date') else install.scheduled_date
    install.completed_date = date.fromisoformat(request.form['completed_date']) if request.form.get('completed_date') else install.completed_date
    install.installed_by   = current_user.id
    install.notes          = _clean(request.form.get('notes', ''), 500)
    if install.status == 'Completed':
        log_action(pid, 'App installation completed', new_val='Completed')
        
    if not install.id:
        db.session.add(install)
    db.session.flush()
    auto_advance_stage(proj)
    db.session.commit()
    flash('Installation record updated.', 'success')
    return redirect(url_for('project_detail', pid=pid))


# ─────────────────────────────────────────────────────────────────────────────
# ADMIN — USER MANAGEMENT
# ─────────────────────────────────────────────────────────────────────────────

@app.route('/admin/users')
@login_required
@roles_required('admin')
def manage_users():
    active_users   = User.query.filter_by(is_deleted=False, is_active=True)\
                               .order_by(User.role, User.full_name).all()
    inactive_users = User.query.filter_by(is_deleted=False, is_active=False)\
                               .order_by(User.role, User.full_name).all()
    deleted_users  = User.query.filter_by(is_deleted=True)\
                               .order_by(User.role, User.full_name).all()
    return render_template('admin_users.html',
        users=active_users,
        inactive_users=inactive_users,
        deleted_users=deleted_users)
@app.route('/admin/users/<int:user_id>/restore', methods=['POST'])
@login_required
@roles_required('admin')
def restore_user(user_id):
    u = User.query.get_or_404(user_id)
    if not u.is_deleted:
        flash('User is not deleted.', 'warning')
        return redirect(url_for('manage_users'))
    u.is_deleted = False
    u.is_active  = True
    u.status     = 'active'
    db.session.commit()
    flash(f'User {u.username} restored successfully.', 'success')
    return redirect(url_for('manage_users'))

@app.route('/admin/users/new', methods=['GET', 'POST'])
@login_required
@roles_required('admin')
def new_user():
    if request.method == 'POST':
        username  = _clean(request.form.get('username', ''), 80).lower()
        full_name = _clean(request.form.get('full_name', ''), 120)
        role      = request.form.get('role', '')
        password  = request.form.get('password', '')
        raw_email = _clean(request.form.get('email', ''), 120).lower()
        raw_phone = _clean(request.form.get('phone', ''), 20)
 
        # ── Normalise phone: keep digits only, strip leading +91 / 0 ────────
        phone_digits = ''.join(c for c in raw_phone if c.isdigit())
        if phone_digits.startswith('91') and len(phone_digits) == 12:
            phone_digits = phone_digits[2:]          # strip country code
        if phone_digits.startswith('0') and len(phone_digits) == 11:
            phone_digits = phone_digits[1:]
        phone_clean = phone_digits if len(phone_digits) == 10 else ''
 
        # ── At least one contact method required ────────────────────────────
        if not raw_email and not phone_clean:
            flash('Please provide at least an email address or a phone number.', 'danger')
            return render_template('new_user.html')
 
        # ── Validate email format when provided ─────────────────────────────
        if raw_email and not re.match(r'^[^@\s]+@[^@\s]+\.[^@\s]+$', raw_email):
            flash('Please enter a valid email address.', 'danger')
            return render_template('new_user.html')
 
        # ── Validate phone when provided ─────────────────────────────────────
        if raw_phone and not phone_clean:
            flash('Please enter a valid 10-digit phone number.', 'danger')
            return render_template('new_user.html')
 
        # ── Password strength ────────────────────────────────────────────────
        errors = _validate_password(password)
        if errors:
            flash(f'Password must contain: {", ".join(errors)}.', 'danger')
            return render_template('new_user.html')
 
        # ── Duplicate username ───────────────────────────────────────────────
        if User.query.filter_by(username=username).first():
            flash('Username already taken.', 'danger')
            return render_template('new_user.html')
 
        # ── Duplicate email (only when a real email is provided) ─────────────
        # if raw_email and User.query.filter(
        #         db.func.lower(User.email) == raw_email,
        #         ~User.email.like('%@noemail.local')).first():
        #     flash('Email address is already registered.', 'danger')
        #     return render_template('new_user.html')
 
        # ── Duplicate phone ──────────────────────────────────────────────────
        if phone_clean and User.query.filter_by(phone=phone_clean).first():
            flash('Phone number is already registered.', 'danger')
            return render_template('new_user.html')
 
        # ── Build the email stored in DB ─────────────────────────────────────
        # When no email is given, use a placeholder so User.email is never NULL.
        # The login route looks up by phone first in that case.
        stored_email = raw_email if raw_email else f'{username}@noemail.local'
 
        u = User(
            username  = username,
            email     = stored_email,
            phone     = phone_clean or None,
            full_name = full_name,
            role      = role,
        )
        u.set_password(password)
        db.session.add(u)
        db.session.commit()
        # log_admin_action('CREATE_USER', target=u.username, detail=u.role)
 
        contact_info = raw_email if raw_email else f'+91 {phone_clean}'
        flash(
            f'User {u.username} created. ',
            'success'
        )
        return redirect(url_for('manage_users'))
 
    return render_template('new_user.html')

@app.route('/admin/users/<int:user_id>/edit', methods=['GET', 'POST'])
@login_required
@roles_required('admin')
def edit_user(user_id):
    u = User.query.get_or_404(user_id)
 
    if request.method == 'POST':
        full_name  = _clean(request.form.get('full_name', ''), 120)
        username   = _clean(request.form.get('username',  ''), 80).lower()
        role       = request.form.get('role', u.role)
        new_pw     = request.form.get('password', '')
        raw_email  = _clean(request.form.get('email', ''), 120).lower()
        raw_phone  = _clean(request.form.get('phone', ''), 20)
 
        # ── Normalise phone ───────────────────────────────────────────────
        phone_digits = ''.join(c for c in raw_phone if c.isdigit())
        if phone_digits.startswith('91') and len(phone_digits) == 12:
            phone_digits = phone_digits[2:]
        if phone_digits.startswith('0') and len(phone_digits) == 11:
            phone_digits = phone_digits[1:]
        phone_clean = phone_digits if len(phone_digits) == 10 else ''
 
        # ── At least one contact method ───────────────────────────────────
        if not raw_email and not phone_clean:
            flash('Please provide at least an email address or a phone number.', 'danger')
            return render_template('edit_user.html', user=u)
 
        # ── Validate email format ─────────────────────────────────────────
        if raw_email and not re.match(r'^[^@\s]+@[^@\s]+\.[^@\s]+$', raw_email):
            flash('Please enter a valid email address.', 'danger')
            return render_template('edit_user.html', user=u)
 
        # ── Validate phone ────────────────────────────────────────────────
        if raw_phone and not phone_clean:
            flash('Please enter a valid 10-digit phone number.', 'danger')
            return render_template('edit_user.html', user=u)
 
        # ── Duplicate username (exclude self) ─────────────────────────────
        if User.query.filter(User.username == username, User.id != user_id).first():
            flash('Username already taken.', 'danger')
            return render_template('edit_user.html', user=u)
 
        # ── Duplicate email (exclude self and @noemail.local rows) ─────────
        if raw_email and User.query.filter(
                db.func.lower(User.email) == raw_email,
                User.id != user_id,
                ~User.email.like('%@noemail.local')).first():
            flash('Email address is already registered.', 'danger')
            return render_template('edit_user.html', user=u)
 
        # ── Duplicate phone (exclude self) ────────────────────────────────
        if phone_clean and User.query.filter(
                User.phone == phone_clean,
                User.id != user_id).first():
            flash('Phone number is already registered.', 'danger')
            return render_template('edit_user.html', user=u)
 
        # ── Password strength (only when a new password is supplied) ──────
        if new_pw:
            errors = _validate_password(new_pw)
            if errors:
                flash(f'Password must contain: {", ".join(errors)}.', 'danger')
                return render_template('edit_user.html', user=u)
            u.set_password(new_pw)
 
        # ── Build stored email ────────────────────────────────────────────
        stored_email = raw_email if raw_email else f'{username}@noemail.local'
 
        u.full_name = full_name
        u.username  = username
        u.email     = stored_email
        u.phone     = phone_clean or None
        u.role      = role
 
        db.session.commit()
        # log_admin_action('EDIT_USER', target=u.username)
 
        contact_info = raw_email if raw_email else f'+91 {phone_clean}'
        flash(
            f'User {u.username} updated. '
            f'OTP login contact: {contact_info}.',
            'success'
        )
        return redirect(url_for('manage_users'))
 
    return render_template('edit_user.html', user=u)

@app.route('/admin/users/<int:user_id>/delete', methods=['POST'])
@login_required
@roles_required('admin')
def delete_user(user_id):
    u = User.query.get_or_404(user_id)
    if u.id == current_user.id:
        flash('You cannot delete your own account.', 'danger')
        return redirect(url_for('manage_users'))
    u.is_deleted = True
    u.is_active  = False
    u.status     = 'inactive'
    db.session.commit()
    # log_admin_action('DELETE_USER', target=u.username)
    flash(f'User {u.username} deleted. You can restore them from the Deleted Users section.', 'success')
    return redirect(url_for('manage_users'))

@app.route('/admin/users/<int:user_id>/status', methods=['POST'])
@login_required
@roles_required('admin')
def change_user_status(user_id):
    u = User.query.get_or_404(user_id)
    if u.id == current_user.id:
        flash('You cannot change your own status.', 'danger')
        return redirect(url_for('manage_users'))
    if u.is_deleted:
        flash('Cannot change status of a deleted user. Restore them first.', 'danger')
        return redirect(url_for('manage_users'))
    new_status = request.form.get('status')
    if new_status not in ('active', 'inactive'):
        flash('Invalid status.', 'danger')
        return redirect(url_for('manage_users'))
    u.status    = new_status
    u.is_active = (new_status == 'active')
    db.session.commit()
    # log_admin_action('USER_STATUS', target=u.username, detail=new_status)
    flash(f'{u.username} marked as {new_status}.', 'success')
    return redirect(url_for('manage_users'))


@app.route('/admin/analytics')
@login_required
@roles_required('admin','director')
def admin_analytics():
    from collections import defaultdict, Counter

    today    = date.today()
    all_projs = Project.query.all()
    all_pays  = Payment.query.all()

    def _monthly_iter(n):
        result = []
        for i in range(n - 1, -1, -1):
            ref = today.replace(day=1)
            for _ in range(i):
                ref = (ref - timedelta(days=1)).replace(day=1)
            result.append(ref)
        return result

    monthly_refs = _monthly_iter(12)

    weekly_projects  = defaultdict(int)
    weekly_payments  = defaultdict(float)
    for i in range(7, -1, -1):
        ws    = today - timedelta(weeks=i)
        we    = today - timedelta(weeks=i - 1)
        label = ws.strftime('%d %b')
        weekly_projects[label] = sum(1 for p in all_projs if ws <= p.created_at.date() < we)
        weekly_payments[label] = sum(float(p.amount) for p in all_pays if ws <= p.payment_date < we)

    monthly_projects = {r.strftime('%b %Y'): sum(1 for p in all_projs if p.created_at.year == r.year and p.created_at.month == r.month) for r in monthly_refs}
    monthly_payments = {r.strftime('%b %Y'): sum(float(p.amount) for p in all_pays if p.payment_date.year == r.year and p.payment_date.month == r.month) for r in monthly_refs}

    yearly_projects, yearly_payments = defaultdict(int), defaultdict(float)
    for i in range(4, -1, -1):
        yr = today.year - i
        yearly_projects[str(yr)] = sum(1 for p in all_projs if p.created_at.year == yr)
        yearly_payments[str(yr)] = sum(float(p.amount) for p in all_pays if p.payment_date.year == yr)

    status_counts = Counter(p.status for p in all_projs)
    type_counts   = Counter(p.project_type for p in all_projs)

    doc_staff_users = User.query.filter_by(role='documents', is_active=True).all()
    doc_staff_stats = []
    for staff in doc_staff_users:
        assigned  = [p for p in all_projs if p.doc_staff_id == staff.id]
        completed = [p for p in assigned if p.status in ['Completed','Closed']]
        inprog    = [p for p in assigned if p.status == 'InProgress']
        not_started = [p for p in inprog if len(p.documents) == 0]
        total_docs  = sum(len(get_expected_docs(p.project_type, p.project_subtype, p.loan_subtype)) for p in assigned)
        done_docs   = sum(get_doc_completion(p)[0] for p in assigned)
        doc_staff_stats.append({
            'name': staff.full_name, 'assigned': len(assigned), 'completed': len(completed),
            'inprog': len(inprog), 'not_started': len(not_started),
            'total_docs': total_docs, 'done_docs': done_docs,
            'doc_pct': int(done_docs / total_docs * 100) if total_docs > 0 else 0,
        })

    coordinators = User.query.filter_by(role='coordinator', is_active=True).all()
    coord_stats  = []
    for c in coordinators:
        cp = [p for p in all_projs if p.coordinator_id == c.id]
        coord_stats.append({
            'name': c.full_name.split()[0], 'total': len(cp),
            'completed': sum(1 for p in cp if p.status in ['Completed','Closed']),
            'delayed':   sum(1 for p in cp if p.status == 'Delayed'),
            'collected': sum(float(p.collected_amount or 0) for p in cp),
        })

    chart_data = {
        'weekly_labels':    list(weekly_projects.keys()),
        'weekly_projects':  list(weekly_projects.values()),
        'weekly_payments':  list(weekly_payments.values()),
        'monthly_labels':   list(monthly_projects.keys()),
        'monthly_projects': list(monthly_projects.values()),
        'monthly_payments': list(monthly_payments.values()),
        'yearly_labels':    list(yearly_projects.keys()),
        'yearly_projects':  list(yearly_projects.values()),
        'yearly_payments':  list(yearly_payments.values()),
        'status_labels':    list(status_counts.keys()),
        'status_counts':    list(status_counts.values()),
        'type_labels':      list(type_counts.keys()),
        'type_counts':      list(type_counts.values()),
        'coord_names':      [c['name']      for c in coord_stats],
        'coord_total':      [c['total']     for c in coord_stats],
        'coord_completed':  [c['completed'] for c in coord_stats],
        'coord_delayed':    [c['delayed']   for c in coord_stats],
        'coord_collected':  [c['collected'] for c in coord_stats],
        'staff_names':      [s['name'].split()[0]               for s in doc_staff_stats],
        'staff_completed':  [s['completed']                     for s in doc_staff_stats],
        'staff_inprog':     [s['inprog'] - s['not_started']     for s in doc_staff_stats],
        'staff_notstarted': [s['not_started']                   for s in doc_staff_stats],
        'staff_pending':    [s['assigned'] - s['completed']     for s in doc_staff_stats],
    }

    total_collected = sum(float(p.amount) for p in all_pays if p.project.status not in ['Cancelled','OnHold'])
    total_value     = sum(float(p.total_amount or 0) for p in all_projs if p.status not in ['Cancelled','OnHold'])

    return render_template('admin_analytics.html',
        total_projects=len(all_projs), total_collected=total_collected,
        total_value=total_value, total_pending=total_value - total_collected,
        coord_stats=coord_stats, chart_data=chart_data, doc_staff_stats=doc_staff_stats)


# ─────────────────────────────────────────────────────────────────────────────
# API — JSON ENDPOINTS
# ─────────────────────────────────────────────────────────────────────────────

@app.route('/api/projects')
@login_required
def api_projects():
    projects_list = Project.query.join(Customer).all()
    return jsonify([{
        'id': p.id, 'code': p.project_code, 'customer': p.customer.name,
        'status': p.status, 'stage': p.stage,
        'kw': float(p.inverter_capacity_kw), 'type': p.project_type,
        'collected': float(p.collected_amount or 0), 'total': float(p.total_amount or 0),
        'payment_pct': p.payment_pct, 'days_open': p.days_open,
    } for p in projects_list])


@app.route('/api/dashboard_stats')
@login_required
def api_dashboard_stats():
    return jsonify({
        'total':     Project.query.count(),
        'inprog':    Project.query.filter_by(status='InProgress').count(),
        'completed': Project.query.filter(Project.status.in_(['Completed','Closed'])).count(),
        'delayed':   Project.query.filter_by(status='Delayed').count(),
        'collected': float(db.session.query(db.func.sum(Payment.amount)).scalar() or 0),
    })


@app.route('/api/notifications')
@login_required
@limiter.limit('60 per minute')
def api_notifications():
    notifs = Notification.query.options(
        joinedload(Notification.project),
        noload(Notification.user),
    ).filter_by(user_id=current_user.id).order_by(
        Notification.created_at.desc()).limit(20).all()
    return jsonify([{
        'id': n.id, 'message': n.message, 'type': n.notif_type,
        'project_id': n.project_id, 'code': n.project.project_code if n.project else '',
        'created_at': n.created_at.strftime('%d %b %H:%M'),
        'is_read': n.is_read,
        'action_url': n.action_url or (f'/projects/{n.project_id}' if n.project_id else '/dashboard'),
    } for n in notifs])


@app.route('/api/notifications/read/<int:nid>', methods=['POST'])
@login_required
@csrf.exempt
def mark_notification_read(nid):
    n = Notification.query.get_or_404(nid)
    if n.user_id != current_user.id:
        return jsonify({'error': 'forbidden'}), 403
    n.is_read = True
    db.session.commit()
    return jsonify({'ok': True})


@app.route('/api/notifications/read_all', methods=['POST'])
@login_required
@csrf.exempt
def mark_all_read():
    Notification.query.filter_by(user_id=current_user.id, is_read=False).update({'is_read': True})
    db.session.commit()
    return jsonify({'ok': True})


# ─────────────────────────────────────────────────────────────────────────────
# TEMPLATE FILTERS
# ─────────────────────────────────────────────────────────────────────────────
@app.context_processor
def inject_file_helpers():
    return dict(file_exists=lambda p: bool(p) and os.path.isfile(p))
@app.template_filter('inr')
def inr_filter(value):
    try:
        return f'₹{float(value):,.0f}'
    except Exception:
        return '₹0'


@app.template_filter('date_fmt')
def date_fmt(value):
    if not value:
        return '—'
    if isinstance(value, (datetime, date)):
        return value.strftime('%d/%m/%Y')
    return str(value)


# ─────────────────────────────────────────────────────────────────────────────
# EXCEL REPORT GENERATION  (unchanged from original — omitted for brevity)
# ─────────────────────────────────────────────────────────────────────────────

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

C_HEADER_BG = '1A3C5E'; C_HEADER_FG = 'FFFFFF'
C_SUBHDR_BG = '2E6DA4'; C_SUBHDR_FG = 'FFFFFF'
C_ACCENT_BG = 'D6E4F0'; C_ALT_BG    = 'F2F7FB'
C_TOTAL_BG  = 'FFF3CD'; C_TOTAL_FG  = '856404'
C_BORDER    = 'BFCBD6'
C_GREEN_BG  = 'D4EDDA'; C_GREEN_FG  = '155724'
C_RED_BG    = 'F8D7DA'; C_RED_FG    = '721C24'
C_AMBER_BG  = 'FFF3CD'; C_AMBER_FG  = '856404'
STATUS_COLORS = {
    'Completed': (C_GREEN_BG, C_GREEN_FG), 'Closed': (C_GREEN_BG, C_GREEN_FG),
    'InProgress': ('D1ECF1', '0C5460'),    'Delayed': (C_RED_BG,  C_RED_FG),
    'OnHold': (C_AMBER_BG, C_AMBER_FG),   'Cancelled': ('E2E3E5', '383D41'),
    'Lead': ('E2E3E5', '383D41'),
}

def _fill(h): return PatternFill('solid', start_color=h, fgColor=h)
def _font(bold=False, color='000000', size=10, italic=False):
    return Font(name='Arial', bold=bold, color=color, size=size, italic=italic)
def _border():
    s = Side(style='thin', color=C_BORDER)
    return Border(left=s, right=s, top=s, bottom=s)
def _center(): return Alignment(horizontal='center', vertical='center', wrap_text=True)
def _left():   return Alignment(horizontal='left',   vertical='center', wrap_text=True)
def _right():  return Alignment(horizontal='right',  vertical='center')
def _inr(v):
    try: return float(v or 0)
    except: return 0.0

def _style_header_cell(cell, text, bg=C_HEADER_BG, fg=C_HEADER_FG, size=10, center=True):
    cell.value = text; cell.font = _font(bold=True, color=fg, size=size)
    cell.fill  = _fill(bg); cell.border = _border()
    cell.alignment = _center() if center else _left()

def _style_data_cell(cell, value, bg='FFFFFF', fg='000000', bold=False,
                     align='left', number_fmt=None):
    cell.value = value; cell.font = _font(bold=bold, color=fg)
    cell.fill  = _fill(bg); cell.border = _border()
    cell.alignment = _center() if align == 'center' else (_right() if align == 'right' else _left())
    if number_fmt: cell.number_format = number_fmt

def _page_setup(ws):
    ws.sheet_view.showGridLines = False
    ws.page_setup.orientation   = 'landscape'
    ws.page_setup.paperSize     = ws.PAPERSIZE_A4
    ws.page_setup.fitToPage     = True
    ws.page_setup.fitToWidth    = 1
    ws.page_setup.fitToHeight   = 0
    ws.print_options.horizontalCentered = True


def build_coordinator_monthly_report(coordinator, all_projects, year, month, output_dir='/tmp'):
    month_name  = calendar.month_name[month]
    month_start = date(year, month, 1)
    month_end   = date(year, month, calendar.monthrange(year, month)[1])
    month_projects     = [p for p in all_projects
                           if month_start <= p.created_at.date() <= month_end
                           and p.status not in ('Cancelled',)]
    created_this_month = [p for p in all_projects if p.created_at.year == year and p.created_at.month == month]
    wb = Workbook()
    ws = wb.active; ws.title = 'Summary'; _page_setup(ws); ws.freeze_panes = 'A6'

    for row_h, (r, h) in enumerate([(28,'Power On Plus Solar Solutions'),(22,f'Monthly Work Report — {month_name} {year}'),(18,f'Coordinator: {coordinator.full_name}'),(8,'')], 1):
        ws.merge_cells(f'A{row_h}:H{row_h}')
        c = ws[f'A{row_h}']; c.value = h if h else None
        if row_h == 1:   c.font = _font(bold=True, color=C_HEADER_FG, size=14); c.fill = _fill(C_HEADER_BG)
        elif row_h == 2: c.font = _font(bold=True, color=C_HEADER_FG, size=11); c.fill = _fill(C_SUBHDR_BG)
        elif row_h == 3: c.font = _font(italic=True, color='333333');            c.fill = _fill(C_ALT_BG)
        if row_h < 4: c.alignment = _center()
        ws.row_dimensions[row_h].height = row_h

    ws.row_dimensions[1].height = 28; ws.row_dimensions[2].height = 22
    ws.row_dimensions[3].height = 18; ws.row_dimensions[4].height = 8
    ws.row_dimensions[5].height = 20

    for col, h in enumerate(['New This Month','Active Projects','Completed','Delayed',
                              'Total Value (₹)','Collected (₹)','Pending (₹)','Collection %'], 1):
        _style_header_cell(ws.cell(5, col), h, bg=C_SUBHDR_BG)

    active    = [p for p in month_projects if p.status in ('InProgress','Delayed','Lead','OnHold')]
    completed = [p for p in month_projects if p.status in ('Completed','Closed')]
    delayed   = [p for p in month_projects if p.status == 'Delayed']
    total_val = sum(_inr(p.total_amount) for p in month_projects)
    collected = sum(_inr(p.collected_amount) for p in month_projects)
    pending   = total_val - collected
    pct       = (collected / total_val * 100) if total_val else 0

    ws.row_dimensions[6].height = 20
    for col, (val, fmt) in enumerate(zip(
        [len(created_this_month), len(active), len(completed), len(delayed),
         total_val, collected, pending, pct / 100],
        [None,None,None,None,'₹#,##0','₹#,##0','₹#,##0','0.0%']), 1):
        _style_data_cell(ws.cell(6, col), val, align='center', bold=True, number_fmt=fmt)

    ws.row_dimensions[7].height = 8; ws.row_dimensions[8].height = 20
    for col, h in enumerate(['MNRE No.','Customer','Place','Type','Subtype','Status',
                              'Contract (₹)','Collected (₹)','Pending (₹)','Doc Staff','Created'], 1):
        _style_header_cell(ws.cell(8, col), h)

    row = 9
    for i, p in enumerate(sorted(month_projects, key=lambda x: x.created_at, reverse=True)):
        bg = C_ALT_BG if i % 2 == 0 else 'FFFFFF'
        s_bg, s_fg = STATUS_COLORS.get(p.status, ('FFFFFF','000000'))
        pend_val = _inr(p.total_amount) - _inr(p.collected_amount)
        ws.row_dimensions[row].height = 18
        for col, (val, fmt, aln) in enumerate(zip(
            [p.project_code, p.customer.name,p.customer.place or '—', p.project_type, p.project_subtype or '—',
              p.status, _inr(p.total_amount), _inr(p.collected_amount), pend_val,
             p.doc_staff.full_name if p.doc_staff else '—', p.created_at.strftime('%d %b %Y')],
            [None,None,None,None,None,None,'₹#,##0','₹#,##0','₹#,##0',None,None,None],
            ['center','left','center','center','center','center','right','right','right','left','center','center']), 1):
            cell = ws.cell(row, col)
            _style_data_cell(cell, val, bg=s_bg if col==6 else bg, fg=s_fg if col==6 else '000000',
                             align=aln, number_fmt=fmt)
        row += 1

    ws.row_dimensions[row].height = 20
    for col, (val, fmt, aln) in enumerate(zip(
    ['TOTAL', f'{len(month_projects)} projects','','','','',total_val,collected,pending,'',''],
    [None,None,None,None,None,None,'₹#,##0','₹#,##0','₹#,##0',None,None],
    ['center','left','','','','','right','right','right','','']), 1):
        cell = ws.cell(row, col)
        if val == '': cell.fill = _fill(C_TOTAL_BG); cell.border = _border(); continue
        _style_data_cell(cell, val, bg=C_TOTAL_BG, fg=C_TOTAL_FG, bold=True, align=aln or 'center', number_fmt=fmt)

    for i, w in enumerate([12,24,8,10,16,12,14,14,14,18,13], 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws2 = wb.create_sheet('New This Month'); _page_setup(ws2)
    _build_project_sheet(ws2, created_this_month, f'New Projects — {month_name} {year}', coordinator.full_name)
    ws3 = wb.create_sheet('By Stage'); _page_setup(ws3)
    _build_stage_sheet(ws3, month_projects, month_name, year, coordinator.full_name)
    ws4 = wb.create_sheet('Payments'); _page_setup(ws4)
    _build_payments_sheet(ws4, month_projects, month_name, year, coordinator.full_name, month_start, month_end)

    fname = f'Report_{coordinator.username}_{year}_{month:02d}.xlsx'.replace(' ', '_')
    path  = os.path.join(output_dir, fname)
    wb.save(path)
    return path


def _build_project_sheet(ws, projects, title, coord_name):
    ws.merge_cells('A1:J1'); c = ws['A1']; c.value = title
    c.font = _font(bold=True, color=C_HEADER_FG, size=12); c.fill = _fill(C_HEADER_BG); c.alignment = _center()
    ws.merge_cells('A2:J2'); c = ws['A2']; c.value = f'Coordinator: {coord_name}'
    c.font = _font(italic=True, color='444444'); c.fill = _fill(C_ALT_BG); c.alignment = _center()
    ws.row_dimensions[3].height = 8
    for col, h in enumerate(['MNRE No.','Customer','Place','Type','Subtype','Status',
                              'Contract (₹)','Collected (₹)','Pending (₹)','Doc Staff'], 1):
        _style_header_cell(ws.cell(4, col), h)
    for i, p in enumerate(projects):
        row = i + 5; bg = C_ALT_BG if i % 2 == 0 else 'FFFFFF'
        s_bg, s_fg = STATUS_COLORS.get(p.status, ('FFFFFF','000000'))
        pend = _inr(p.total_amount) - _inr(p.collected_amount)
        for col, (val, fmt, aln) in enumerate(zip(
    [p.project_code, p.customer.name, p.project_type, p.project_subtype or '—',
     p.customer.place or '—',
     p.status, _inr(p.total_amount), _inr(p.collected_amount), pend,
     p.doc_staff.full_name if p.doc_staff else '—'],
    [None,None,None,None,None,None,'₹#,##0','₹#,##0','₹#,##0',None],
    ['center','left','center','center','left','center','right','right','right','left']), 1):
            _style_data_cell(ws.cell(row, col), val,
                bg=s_bg if col==6 else bg, fg=s_fg if col==6 else '000000', align=aln, number_fmt=fmt)
        ws.row_dimensions[row].height = 17
    for i, w in enumerate([12,24,8,10,16,12,14,14,14,18], 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def _build_stage_sheet(ws, projects, month_name, year, coord_name):
    stages = ['Lead','Site Visit','Documentation','Onsite Work','Connection','Subsidy','Payment']
    ws.merge_cells('A1:D1'); c = ws['A1']; c.value = f'Stage Breakdown — {month_name} {year}'
    c.font = _font(bold=True, color=C_HEADER_FG, size=12); c.fill = _fill(C_HEADER_BG); c.alignment = _center()
    ws.merge_cells('A2:D2'); c = ws['A2']; c.value = f'Coordinator: {coord_name}'
    c.font = _font(italic=True, color='444444'); c.fill = _fill(C_ALT_BG); c.alignment = _center()
    ws.row_dimensions[3].height = 8
    for col, h in enumerate(['Stage','Count','Total Value (₹)','Collected (₹)'], 1):
        _style_header_cell(ws.cell(4, col), h)
    row = 5
    for i, stage in enumerate(stages):
        sp = [p for p in projects if p.stage == stage]
        bg = C_ALT_BG if i % 2 == 0 else 'FFFFFF'
        for col, (val, fmt, aln) in enumerate(zip(
            [stage, len(sp), sum(_inr(p.total_amount) for p in sp), sum(_inr(p.collected_amount) for p in sp)],
            [None,None,'₹#,##0','₹#,##0'], ['left','center','right','right']), 1):
            _style_data_cell(ws.cell(row, col), val, bg=bg, align=aln, number_fmt=fmt)
        ws.row_dimensions[row].height = 17; row += 1
    for col, (val, fmt, aln) in enumerate(zip(
        ['TOTAL', len(projects), sum(_inr(p.total_amount) for p in projects), sum(_inr(p.collected_amount) for p in projects)],
        [None,None,'₹#,##0','₹#,##0'], ['center','center','right','right']), 1):
        _style_data_cell(ws.cell(row, col), val, bg=C_TOTAL_BG, fg=C_TOTAL_FG, bold=True, align=aln, number_fmt=fmt)
    ws.row_dimensions[row].height = 20
    ws.column_dimensions['A'].width = 20; ws.column_dimensions['B'].width = 10
    ws.column_dimensions['C'].width = 18; ws.column_dimensions['D'].width = 18


def _build_payments_sheet(ws, projects, month_name, year, coord_name, month_start, month_end):
    ws.merge_cells('A1:G1'); c = ws['A1']; c.value = f'Payments — {month_name} {year}'
    c.font = _font(bold=True, color=C_HEADER_FG, size=12); c.fill = _fill(C_HEADER_BG); c.alignment = _center()
    ws.merge_cells('A2:G2'); c = ws['A2']; c.value = f'Coordinator: {coord_name}'
    c.font = _font(italic=True, color='444444'); c.fill = _fill(C_ALT_BG); c.alignment = _center()
    ws.row_dimensions[3].height = 8
    for col, h in enumerate(['MNRE No.','Customer','Date','Amount (₹)','Type','Source','Reference'], 1):
        _style_header_cell(ws.cell(4, col), h)
    all_pays = sorted(
        [(p, pay) for p in projects for pay in p.payments if month_start <= pay.payment_date <= month_end],
        key=lambda x: x[1].payment_date, reverse=True)
    for i, (p, pay) in enumerate(all_pays):
        row = i + 5; bg = C_ALT_BG if i % 2 == 0 else 'FFFFFF'
        for col, (val, fmt, aln) in enumerate(zip(
            [p.project_code, p.customer.name, pay.payment_date.strftime('%d %b %Y'),
             _inr(pay.amount), pay.payment_type, pay.payment_source, pay.reference_no or '—'],
            [None,None,None,'₹#,##0',None,None,None],
            ['center','left','center','right','center','center','center']), 1):
            _style_data_cell(ws.cell(row, col), val, bg=bg, align=aln, number_fmt=fmt)
        ws.row_dimensions[row].height = 17
    if all_pays:
        tr = len(all_pays) + 5; total_amt = sum(_inr(pay.amount) for _, pay in all_pays)
        for col in range(1, 8):
            cell = ws.cell(tr, col)
            if col == 1:   _style_data_cell(cell, 'TOTAL', bg=C_TOTAL_BG, fg=C_TOTAL_FG, bold=True, align='center')
            elif col == 2: _style_data_cell(cell, f'{len(all_pays)} payments', bg=C_TOTAL_BG, fg=C_TOTAL_FG, bold=True)
            elif col == 4: _style_data_cell(cell, total_amt, bg=C_TOTAL_BG, fg=C_TOTAL_FG, bold=True, align='right', number_fmt='₹#,##0')
            else:          cell.fill = _fill(C_TOTAL_BG); cell.border = _border()
        ws.row_dimensions[tr].height = 20
    for i, w in enumerate([12,24,14,16,12,12,18], 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def build_docstaff_monthly_report(staff, all_projects, year, month, output_dir='/tmp'):
    month_name  = calendar.month_name[month]
    month_start = date(year, month, 1)
    month_end   = date(year, month, calendar.monthrange(year, month)[1])
    month_projects     = [p for p in all_projects
                           if month_start <= p.created_at.date() <= month_end
                           and p.status not in ('Cancelled',)]
    created_this_month = [p for p in all_projects if p.created_at.year == year and p.created_at.month == month]
    wb = Workbook(); ws = wb.active; ws.title = 'Summary'; _page_setup(ws); ws.freeze_panes = 'A6'
 
    for r, (txt, bg_c, fg_c, sz) in enumerate([
        ('Power On Plus Solar Solutions', C_HEADER_BG, C_HEADER_FG, 14),
        (f'Documents Staff Monthly Report — {month_name} {year}', C_SUBHDR_BG, C_HEADER_FG, 11),
        (f'Staff: {staff.full_name}', C_ALT_BG, '333333', 10),
        ('', 'FFFFFF', '000000', 10),
    ], 1):
        ws.merge_cells(f'A{r}:H{r}')
        c = ws[f'A{r}']; c.value = txt or None
        c.font = _font(bold=(r<3), color=fg_c, size=sz, italic=(r==3))
        c.fill = _fill(bg_c); c.alignment = _center()
    ws.row_dimensions[1].height = 28; ws.row_dimensions[2].height = 22
    ws.row_dimensions[3].height = 18; ws.row_dimensions[4].height = 8; ws.row_dimensions[5].height = 20
 
    def _doc_done(project, doc_name):
        dm = {d.doc_type: d for d in project.documents}
        return dm.get(doc_name) and dm[doc_name].status in ('Received','Sent','Completed')
 
    completed_p    = [p for p in month_projects if p.status in ('Completed','Closed')]
    inprog_p       = [p for p in month_projects if p.status == 'InProgress']
    delayed_p      = [p for p in month_projects if p.status == 'Delayed']
    feas_done      = sum(1 for p in month_projects if _doc_done(p, 'Feasibility Receipt'))
    cd_done        = sum(1 for p in month_projects if _doc_done(p, 'CD Payment Receipt'))
    conn_done      = sum(1 for p in month_projects if _doc_done(p, 'KSEB Connection'))
    pay_done       = sum(1 for p in month_projects if _doc_done(p, 'Payment Completion'))
    total_work_amt = sum(_inr(p.total_amount) for p in month_projects)
 
    # ── Summary KPI row ──────────────────────────────────────────────────────
    for col, h in enumerate(['Total','New This Month','Completed','In Progress',
                              'Work Amount (₹)','Feasibility','CD Payment','Connection',
                              'Payment Compl.','Delayed'], 1):
        _style_header_cell(ws.cell(5, col), h, bg=C_SUBHDR_BG)
 
    ws.row_dimensions[6].height = 20
    for col, (val, fmt) in enumerate(zip(
        [len(month_projects), len(created_this_month), len(completed_p), len(inprog_p),
         total_work_amt, feas_done, cd_done, conn_done, pay_done, len(delayed_p)],
        [None,None,None,None,'₹#,##0',None,None,None,None,None]), 1):
        _style_data_cell(ws.cell(6, col), val, align='center', bold=True, number_fmt=fmt)
 
    ws.row_dimensions[7].height = 8; ws.row_dimensions[8].height = 20
 
    # ── Detail table headers (aligned 1:1 with vals below) ──────────────────
    detail_headers = ['MNRE No.','Customer','Type','Subtype','Status','Work Amount (₹)',
                       'MNRE','Feasibility','CD Payment','Connection','Payment Compl.',
                       'Coordinator','Created']
    for col, h in enumerate(detail_headers, 1):
        _style_header_cell(ws.cell(8, col), h)
 
    row = 9
    for i, p in enumerate(sorted(month_projects, key=lambda x: x.created_at, reverse=True)):
        bg = C_ALT_BG if i % 2 == 0 else 'FFFFFF'
        s_bg, s_fg = STATUS_COLORS.get(p.status, ('FFFFFF','000000'))
        def _tick(dn): return '✓' if _doc_done(p, dn) else '✗'
        ws.row_dimensions[row].height = 18
        vals   = [p.project_code, p.customer.name, p.project_type, p.project_subtype or '—',
                  p.status, _inr(p.total_amount),
                  _tick('MNRE'), _tick('Feasibility Receipt'), _tick('CD Payment Receipt'),
                  _tick('KSEB Connection'), _tick('Payment Completion'),
                  p.coordinator.full_name if p.coordinator else '—',
                  p.created_at.strftime('%d %b %Y')]
        fmts   = [None,None,None,None,None,'₹#,##0',None,None,None,None,None,None,None]
        aligns = ['center','left','center','center','center','right',
                  'center','center','center','center','center','left','center']
        for col, (val, fmt, aln) in enumerate(zip(vals, fmts, aligns), 1):
            cell = ws.cell(row, col)
            tick_cols = (7, 8, 9, 10, 11)   # MNRE, Feasibility, CD Payment, Connection, Payment Compl.
            c_bg = s_bg if col == 5 else (C_GREEN_BG if val == '✓' else C_RED_BG) if col in tick_cols else bg
            c_fg = s_fg if col == 5 else (C_GREEN_FG if val == '✓' else C_RED_FG) if col in tick_cols else '000000'
            _style_data_cell(cell, val, bg=c_bg, fg=c_fg, align=aln, number_fmt=fmt)
        row += 1
 
    mnre_done = sum(1 for p in month_projects if _doc_done(p, 'MNRE'))
    ws.row_dimensions[row].height = 20
    for col in range(1, len(detail_headers) + 1):
        cell = ws.cell(row, col)
        if col == 1:  _style_data_cell(cell, 'TOTAL', bg=C_TOTAL_BG, fg=C_TOTAL_FG, bold=True, align='center')
        elif col == 2: _style_data_cell(cell, f'{len(month_projects)} projects', bg=C_TOTAL_BG, fg=C_TOTAL_FG, bold=True)
        elif col == 6: _style_data_cell(cell, total_work_amt, bg=C_TOTAL_BG, fg=C_TOTAL_FG, bold=True, align='right', number_fmt='₹#,##0')
        elif col == 7: _style_data_cell(cell, mnre_done, bg=C_TOTAL_BG, fg=C_TOTAL_FG, bold=True, align='center')
        elif col == 8: _style_data_cell(cell, feas_done, bg=C_TOTAL_BG, fg=C_TOTAL_FG, bold=True, align='center')
        elif col == 9: _style_data_cell(cell, cd_done,   bg=C_TOTAL_BG, fg=C_TOTAL_FG, bold=True, align='center')
        elif col == 10:_style_data_cell(cell, conn_done, bg=C_TOTAL_BG, fg=C_TOTAL_FG, bold=True, align='center')
        elif col == 11:_style_data_cell(cell, pay_done,  bg=C_TOTAL_BG, fg=C_TOTAL_FG, bold=True, align='center')
        else:          cell.fill = _fill(C_TOTAL_BG); cell.border = _border()
 
    col_widths = [12,24,8,10,12,16,8,10,10,10,10,20,13]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
 
    ws2 = wb.create_sheet('Document Status'); _page_setup(ws2)
    ws2.merge_cells('A1:F1'); c = ws2['A1']; c.value = f'Document Status Detail — {month_name} {year}'
    c.font = _font(bold=True, color=C_HEADER_FG, size=12); c.fill = _fill(C_HEADER_BG); c.alignment = _center()
    ws2.merge_cells('A2:F2'); c = ws2['A2']; c.value = f'Staff: {staff.full_name}'
    c.font = _font(italic=True, color='444444'); c.fill = _fill(C_ALT_BG); c.alignment = _center()
    ws2.row_dimensions[3].height = 8
    for col, h in enumerate(['MNRE No.','Customer','Document','Status','Received Date','Stage'], 1):
        _style_header_cell(ws2.cell(4, col), h)
    doc_row = 5
    for i, p in enumerate(sorted(month_projects, key=lambda x: x.created_at, reverse=True)):
        expected_docs = get_expected_docs(p.project_type, p.project_subtype, p.loan_subtype)
        doc_map = {d.doc_type: d for d in p.documents}
        bg = C_ALT_BG if i % 2 == 0 else 'FFFFFF'
        for doc_name in expected_docs:
            doc_rec = doc_map.get(doc_name)
            status  = doc_rec.status if doc_rec else 'Pending'
            rec_date = doc_rec.received_date.strftime('%d %b %Y') if doc_rec and doc_rec.received_date else '—'
            d_bg, d_fg = (C_GREEN_BG, C_GREEN_FG) if status in ('Received','Completed','Sent') else (C_RED_BG, C_RED_FG)
            for col, (val, aln) in enumerate(zip(
                [p.project_code, p.customer.name, doc_name, status, rec_date, p.stage],
                ['center','left','left','center','center','center']), 1):
                _style_data_cell(ws2.cell(doc_row, col), val,
                    bg=d_bg if col==4 else bg, fg=d_fg if col==4 else '000000', align=aln)
            ws2.row_dimensions[doc_row].height = 16; doc_row += 1
    ws2.column_dimensions['A'].width = 12; ws2.column_dimensions['B'].width = 24
    ws2.column_dimensions['C'].width = 28; ws2.column_dimensions['D'].width = 12
    ws2.column_dimensions['E'].width = 14; ws2.column_dimensions['F'].width = 16
 
    ws3 = wb.create_sheet('By Stage'); _page_setup(ws3)
    _build_stage_sheet(ws3, month_projects, month_name, year, staff.full_name)
    ws4 = wb.create_sheet('New This Month'); _page_setup(ws4)
    _build_project_sheet(ws4, created_this_month, f'New Projects — {month_name} {year}', staff.full_name)
 
    fname = f'DocsReport_{staff.username}_{year}_{month:02d}.xlsx'.replace(' ', '_')
    path  = os.path.join(output_dir, fname)
    wb.save(path)
    return path
 
# ─────────────────────────────────────────────────────────────────────────────
# REPORT DOWNLOAD ROUTES
# ─────────────────────────────────────────────────────────────────────────────

# ── coordinator_reports route ─────────────────────────────────────────────────
@app.route('/admin/coordinator_reports')
@login_required
@roles_required('admin', 'director','payments','office')
def coordinator_reports():
    coordinators = User.query.filter(
        User.role.in_(['coordinator', 'director']),
        User.is_active == True
    ).order_by(User.full_name).all()
    today = date.today()

    # Collect free-text coordinator names (projects with no coordinator_id)
    other_names = (
        db.session.query(Project.coordinator_name)
        .filter(Project.coordinator_id == None, Project.coordinator_name != None, Project.coordinator_name != '')
        .distinct()
        .all()
    )
    seen = {}
    for (name,) in other_names:
        key = name.strip().lower()
        if key not in seen:
            seen[key] = name.strip()
    other_coord_names = sorted(seen.values(), key=lambda n: n.lower())

    return render_template('coordinator_reports.html',
                           coordinators=coordinators,
                           other_coord_names=other_coord_names,
                           current_year=today.year,
                           current_month=today.month)
@app.route('/admin/coordinator_reports/download')
@login_required
@roles_required('admin','director','payments','office')
def download_coordinator_report():
    coord_id   = request.args.get('coordinator_id', type=int)
    coord_name = request.args.get('coordinator_name', '').strip()
    year       = request.args.get('year',  type=int)
    month      = request.args.get('month', type=int)

    if not (coord_id or coord_name) or not all([year, month]) or not (1 <= month <= 12):
        flash('Invalid report parameters.', 'danger')
        return redirect(url_for('coordinator_reports'))

    label, coordinator, projects = _resolve_coord_projects(coord_id, coord_name)
    path = build_coordinator_monthly_report(coordinator, projects, year, month, tempfile.gettempdir())
    month_name = calendar.month_name[month]
    return send_file(path, as_attachment=True,
        download_name=f'Report_{coordinator.username}_{month_name}_{year}.xlsx',
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

def _resolve_coord_projects(coord_id, coord_name):
    """
    Returns (label_name, projects_queryset_list) for either a User-based
    coordinator (coord_id) or a free-text coordinator_name (coord_name).
    """
    if coord_id:
        coordinator = User.query.get_or_404(coord_id)
        projects = Project.query.filter_by(coordinator_id=coord_id).all()
        return coordinator.full_name, coordinator, projects
    else:
        # Case-insensitive match on coordinator_name
        projects = Project.query.filter(
            Project.coordinator_id == None,
            db.func.lower(Project.coordinator_name) == coord_name.strip().lower()
        ).all()
        # Fake a simple object for templates that expect coordinator.full_name
        class _FakeCoord:
            def __init__(self, name):
                self.full_name = name
                self.username  = name.replace(' ', '_').lower()
        return coord_name.strip(), _FakeCoord(coord_name.strip()), projects

@app.route('/admin/docstaff_reports')
@login_required
@roles_required('admin','director','office')
def docstaff_reports():
    today = date.today()
    return render_template('docstaff_reports.html',
        staff_list=User.query.filter(
            User.role.in_(['documents', 'office','documents_k']),
            User.is_active == True
        ).order_by(User.full_name).all(),
        current_year=today.year,
        current_month=today.month)

@app.route('/admin/docstaff_reports/download')
@login_required
@roles_required('admin','director','office')
def download_docstaff_report():
    staff_id = request.args.get('staff_id', type=int)
    year     = request.args.get('year',     type=int)
    month    = request.args.get('month',    type=int)
    if not all([staff_id, year, month]) or not (1 <= month <= 12):
        flash('Invalid report parameters.', 'danger')
        return redirect(url_for('docstaff_reports'))
    staff = User.query.get_or_404(staff_id)
    if staff.role not in ('documents', 'office','documents_k'):
        flash('Selected user is not a documents or office staff member.', 'danger')
        return redirect(url_for('docstaff_reports'))
    projects   = Project.query.filter_by(doc_staff_id=staff_id).all()
    path       = build_docstaff_monthly_report(staff, projects, year, month, tempfile.gettempdir())
    month_name = calendar.month_name[month]
    return send_file(path, as_attachment=True,
        download_name=f'DocsReport_{staff.username}_{month_name}_{year}.xlsx',
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

# ─────────────────────────────────────────────────────────────────────────────
# REPORT HELPERS  — paste anywhere above the routes, e.g. near _inr()
# ─────────────────────────────────────────────────────────────────────────────

STATUS_COLORS_HTML = {
    'Completed':  ('#d4edda', '#155724'),
    'Closed':     ('#d4edda', '#155724'),
    'InProgress': ('#d1ecf1', '#0c5460'),
    'Delayed':    ('#f8d7da', '#721c24'),
    'OnHold':     ('#fff3cd', '#856404'),
    'Cancelled':  ('#e2e3e5', '#383d41'),
    'Lead':       ('#e2e3e5', '#383d41'),
}


def _inr_fmt(v):
    try:
        return f'₹{float(v or 0):,.0f}'
    except Exception:
        return '₹0'


def _doc_done(project, doc_name):
    dm = {d.doc_type: d for d in project.documents}
    return bool(dm.get(doc_name) and dm[doc_name].status in ('Received', 'Sent', 'Completed'))


def _project_to_dict_coord(p):
    bg, fg = STATUS_COLORS_HTML.get(p.status, ('#fff', '#000'))
    return {
        'code':      p.project_code,
        'customer':  p.customer.name,
        'place':     p.customer.place or '—',
        'inverter_kw': p.inverter_capacity_kw,
        'panel_kw':    p.panel_capacity_kw,
        'type':      p.project_type,
        'subtype':   p.project_subtype or '—',
        'roof':      'Clay Tile (Oodu)' if p.roof_type == 'Clay Tile' else (p.roof_type or '—'),
        'status':    p.status,
        'status_bg': bg,
        'status_fg': fg,
        'contract':  _inr_fmt(p.total_amount),
        'collected': _inr_fmt(p.collected_amount),
        'pending':   _inr_fmt(float(p.total_amount or 0) - float(p.collected_amount or 0)),
        'doc_staff': p.doc_staff.full_name if p.doc_staff else '—',
        'created':   p.created_at.strftime('%d %b %Y'),
    }


def _project_to_dict_docstaff(p):
    bg, fg = STATUS_COLORS_HTML.get(p.status, ('#fff', '#000'))
    cd_ = p.connection_details
    load_s = (cd_.load_clearance_status if cd_ and cd_.load_clearance_needed else 'N/A')
    ow_s   = (cd_.ownership_change_status if cd_ and cd_.ownership_change_needed else 'N/A')
    return {
        'code':        p.project_code,
        'customer':    p.customer.name,
        'place':       p.customer.place or '—',
        'inverter_kw': p.inverter_capacity_kw,
        'panel_kw':    p.panel_capacity_kw,
        'coordinator': p.coordinator.full_name if p.coordinator else (p.coordinator_name or '—'),
        'roof':        'Clay Tile (Oodu)' if p.roof_type == 'Clay Tile' else (p.roof_type or '—'),
        'loan':        p.project_type,
        'load_ow':     f'L:{load_s[:6]} OW:{ow_s[:6]}',
        'type':        p.project_type,
        'subtype':     p.project_subtype or '—',
        'work_amount': _inr_fmt(p.total_amount),
        'status':      p.status,
        'status_bg':   bg,
        'status_fg':   fg,
        'feas':        _doc_done(p, 'Feasibility Receipt'),
        'cd':          _doc_done(p, 'CD Payment Receipt'),
        'conn':        _doc_done(p, 'KSEB Connection'),
        'mnre':        _doc_done(p, 'MNRE'),
        'created':     p.created_at.strftime('%d %b %Y'),
    }


# ── Template filter used in print templates ───────────────────────────────────
@app.template_filter('format_inr')
def format_inr_filter(value):
    try:
        return f'₹{int(float(value)):,}'
    except Exception:
        return '₹0'


# ─────────────────────────────────────────────────────────────────────────────
# ALL-WORKS EXCEL BUILDERS
# ─────────────────────────────────────────────────────────────────────────────

def build_allworks_coordinator_report(coordinator, projects, output_dir='/tmp'):
    """Single flat sheet — all coordinator projects, no month filter."""
    wb = Workbook()
    ws = wb.active
    ws.title = 'All Projects'
    _page_setup(ws)
    ws.freeze_panes = 'A6'

    # Title block
    titles = [
        ('Power On Plus Solar Solutions',          C_HEADER_BG, C_HEADER_FG, 14, False, True),
        (f'All Works — {coordinator.full_name}',   C_SUBHDR_BG, C_HEADER_FG, 11, False, True),
        (f'Generated: {date.today().strftime("%d %b %Y")}', C_ALT_BG, '444444', 10, True, True),
        ('', 'FFFFFF', '000000', 8, False, False),
    ]
    for r, (txt, bg_c, fg_c, sz, italic, center) in enumerate(titles, 1):
        ws.merge_cells(f'A{r}:L{r}')
        c = ws[f'A{r}']
        c.value = txt or None
        c.font  = _font(bold=(not italic and bool(txt)), color=fg_c, size=sz, italic=italic)
        c.fill  = _fill(bg_c)
        if center:
            c.alignment = _center()
    ws.row_dimensions[1].height = 28
    ws.row_dimensions[2].height = 22
    ws.row_dimensions[3].height = 18
    ws.row_dimensions[4].height = 8

    # Column headers
    ws.row_dimensions[5].height = 20
    headers = ['MNRE No.', 'Customer', 'Type', 'Subtype', 'Place', 'Status',
           'Contract (₹)', 'Collected (₹)', 'Pending (₹)', 'Doc Staff', 'Created']
    for col, h in enumerate(headers, 1):
        _style_header_cell(ws.cell(5, col), h)

    # Data
    sorted_projects = sorted(projects, key=lambda x: x.created_at, reverse=True)
    row = 6
    for i, p in enumerate(sorted_projects):
        bg     = C_ALT_BG if i % 2 == 0 else 'FFFFFF'
        s_bg, s_fg = STATUS_COLORS.get(p.status, ('FFFFFF', '000000'))
        pend   = float(p.total_amount or 0) - float(p.collected_amount or 0)
        ws.row_dimensions[row].height = 17
        vals = [p.project_code, p.customer.name, p.project_type, p.project_subtype or '—',
            p.customer.place or '—',
            p.status,
            float(p.total_amount or 0), float(p.collected_amount or 0), pend,
            p.doc_staff.full_name if p.doc_staff else '—',
            p.created_at.strftime('%d %b %Y')]
        fmts   = [None, None, None, None, None, None, '₹#,##0', '₹#,##0', '₹#,##0', None, None]
        aligns = ['center', 'left', 'center', 'center', 'left', 'center',
          'right', 'right', 'right', 'left', 'center']
        for col, (val, fmt, aln) in enumerate(zip(vals, fmts, aligns), 1):
            _style_data_cell(ws.cell(row, col), val,
                             bg=s_bg if col == 6 else bg,
                             fg=s_fg if col == 6 else '000000',
                             align=aln, number_fmt=fmt)
        row += 1

    # Totals
    total_val  = sum(float(p.total_amount    or 0) for p in sorted_projects)
    total_coll = sum(float(p.collected_amount or 0) for p in sorted_projects)
    ws.row_dimensions[row].height = 20
    for col in range(1, 12):
        cell = ws.cell(row, col)
        if col == 1:
            _style_data_cell(cell, 'TOTAL', bg=C_TOTAL_BG, fg=C_TOTAL_FG, bold=True, align='center')
        elif col == 2:
            _style_data_cell(cell, f'{len(sorted_projects)} projects',
                             bg=C_TOTAL_BG, fg=C_TOTAL_FG, bold=True)
        elif col == 7:
            _style_data_cell(cell, total_val,  bg=C_TOTAL_BG, fg=C_TOTAL_FG,
                             bold=True, align='right', number_fmt='₹#,##0')
        elif col == 8:
            _style_data_cell(cell, total_coll, bg=C_TOTAL_BG, fg=C_TOTAL_FG,
                             bold=True, align='right', number_fmt='₹#,##0')
        elif col == 9:
            _style_data_cell(cell, total_val - total_coll, bg=C_TOTAL_BG, fg=C_TOTAL_FG,
                             bold=True, align='right', number_fmt='₹#,##0')
        else:
            cell.fill = _fill(C_TOTAL_BG)
            cell.border = _border()

    col_widths = [12, 24, 8, 10, 16, 12, 14, 14, 14, 18, 13]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    path = os.path.join(output_dir, f'AllWorks_{coordinator.username}.xlsx')
    wb.save(path)
    return path


def build_allworks_docstaff_report(staff, projects, output_dir='/tmp'):
    """Single flat sheet — all projects for a doc staff member, no month filter."""
    wb = Workbook()
    ws = wb.active
    ws.title = 'All Projects'
    _page_setup(ws)
    ws.freeze_panes = 'A6'
 
    titles = [
        ('Power On Plus Solar Solutions',        C_HEADER_BG, C_HEADER_FG, 14, False, True),
        (f'All Works — {staff.full_name}',       C_SUBHDR_BG, C_HEADER_FG, 11, False, True),
        (f'Generated: {date.today().strftime("%d %b %Y")}', C_ALT_BG, '444444', 10, True, True),
        ('', 'FFFFFF', '000000', 8, False, False),
    ]
    for r, (txt, bg_c, fg_c, sz, italic, center) in enumerate(titles, 1):
        ws.merge_cells(f'A{r}:L{r}')
        c = ws[f'A{r}']
        c.value = txt or None
        c.font  = _font(bold=(not italic and bool(txt)), color=fg_c, size=sz, italic=italic)
        c.fill  = _fill(bg_c)
        if center:
            c.alignment = _center()
    ws.row_dimensions[1].height = 28
    ws.row_dimensions[2].height = 22
    ws.row_dimensions[3].height = 18
    ws.row_dimensions[4].height = 8
 
    ws.row_dimensions[5].height = 20
    headers = ['MNRE No.', 'Customer', 'Type', 'Subtype', 'Status', 'Work Amount (₹)',
               'Feasibility', 'CD Payment', 'MNRE', 'KSEB Conn.', 'Coordinator', 'Created']
    for col, h in enumerate(headers, 1):
        _style_header_cell(ws.cell(5, col), h)
 
    sorted_projects = sorted(projects, key=lambda x: x.created_at, reverse=True)
    row = 6
    for i, p in enumerate(sorted_projects):
        bg     = C_ALT_BG if i % 2 == 0 else 'FFFFFF'
        s_bg, s_fg = STATUS_COLORS.get(p.status, ('FFFFFF', '000000'))
        feas   = _doc_done(p, 'Feasibility Receipt')
        cd     = _doc_done(p, 'CD Payment Receipt')
        conn   = _doc_done(p, 'KSEB Connection')
        mnre   = _doc_done(p, 'MNRE')
        ws.row_dimensions[row].height = 17
        vals = [p.project_code, p.customer.name, p.project_type, p.project_subtype or '—',
                p.status, _inr(p.total_amount),
                '✓' if feas else '✗',
                '✓' if cd else '✗',
                '✓' if mnre else '✗',
                '✓' if conn else '✗',
                p.coordinator.full_name if p.coordinator else '—',
                p.created_at.strftime('%d %b %Y')]
        fmts   = [None,None,None,None,None,'₹#,##0',None,None,None,None,None,None]
        aligns = ['center','left','center','center','center','right',
                  'center','center','center','center','left','center']
        for col, (val, fmt, aln) in enumerate(zip(vals, fmts, aligns), 1):
            if col == 5:
                c_bg, c_fg = s_bg, s_fg
            elif col in (7, 8, 9, 10):
                c_bg = C_GREEN_BG if val == '✓' else C_RED_BG
                c_fg = C_GREEN_FG if val == '✓' else C_RED_FG
            else:
                c_bg, c_fg = bg, '000000'
            _style_data_cell(ws.cell(row, col), val, bg=c_bg, fg=c_fg, align=aln, number_fmt=fmt)
        row += 1
 
    feas_done      = sum(1 for p in sorted_projects if _doc_done(p, 'Feasibility Receipt'))
    cd_done        = sum(1 for p in sorted_projects if _doc_done(p, 'CD Payment Receipt'))
    conn_done      = sum(1 for p in sorted_projects if _doc_done(p, 'KSEB Connection'))
    mnre_done      = sum(1 for p in sorted_projects if _doc_done(p, 'MNRE'))
    total_work_amt = sum(_inr(p.total_amount) for p in sorted_projects)
    ws.row_dimensions[row].height = 20
    for col in range(1, len(headers) + 1):
        cell = ws.cell(row, col)
        if col == 1:
            _style_data_cell(cell, 'TOTAL', bg=C_TOTAL_BG, fg=C_TOTAL_FG, bold=True, align='center')
        elif col == 2:
            _style_data_cell(cell, f'{len(sorted_projects)} projects',
                             bg=C_TOTAL_BG, fg=C_TOTAL_FG, bold=True)
        elif col == 6:
            _style_data_cell(cell, total_work_amt, bg=C_TOTAL_BG, fg=C_TOTAL_FG, bold=True, align='right', number_fmt='₹#,##0')
        elif col == 7:
            _style_data_cell(cell, feas_done, bg=C_TOTAL_BG, fg=C_TOTAL_FG, bold=True, align='center')
        elif col == 8:
            _style_data_cell(cell, cd_done, bg=C_TOTAL_BG, fg=C_TOTAL_FG, bold=True, align='center')
        elif col == 9:
            _style_data_cell(cell, mnre_done, bg=C_TOTAL_BG, fg=C_TOTAL_FG, bold=True, align='center')
        elif col == 10:
            _style_data_cell(cell, conn_done, bg=C_TOTAL_BG, fg=C_TOTAL_FG, bold=True, align='center')
        else:
            cell.fill = _fill(C_TOTAL_BG)
            cell.border = _border()
 
    col_widths = [12, 24, 8, 10, 12, 16, 10, 10, 10, 10, 20, 13]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
 
    path = os.path.join(output_dir, f'AllWorks_Docs_{staff.username}.xlsx')
    wb.save(path)
    return path


# ─────────────────────────────────────────────────────────────────────────────
# NEW ROUTES — paste after the existing download_docstaff_report route
# ─────────────────────────────────────────────────────────────────────────────

# ── Coordinator: all-works Excel ─────────────────────────────────────────────
@app.route('/admin/coordinator_reports/download_all')
@login_required
@roles_required('admin','director','payments','office')
def download_coordinator_report_all():
    coord_id   = request.args.get('coordinator_id', type=int)
    coord_name = request.args.get('coordinator_name', '').strip()

    if not (coord_id or coord_name):
        flash('Please select a coordinator.', 'danger')
        return redirect(url_for('coordinator_reports'))

    label, coordinator, projects = _resolve_coord_projects(coord_id, coord_name)
    path = build_allworks_coordinator_report(coordinator, projects, tempfile.gettempdir())
    return send_file(path, as_attachment=True,
        download_name=f'AllWorks_{coordinator.username}.xlsx',
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


# ── Doc staff: all-works Excel ───────────────────────────────────────────────
@app.route('/admin/docstaff_reports/download_all')
@login_required
@roles_required('admin','director','office')
def download_docstaff_report_all():
    staff_id = request.args.get('staff_id', type=int)
    if not staff_id:
        flash('Please select a staff member.', 'danger')
        return redirect(url_for('docstaff_reports'))
    staff = User.query.get_or_404(staff_id)
    if staff.role not in ('documents', 'office','documents_k'):
        flash('Selected user is not a documents or office staff member.', 'danger')
        return redirect(url_for('docstaff_reports'))
    projects = Project.query.filter_by(doc_staff_id=staff_id).all()
    path = build_allworks_docstaff_report(staff, projects, tempfile.gettempdir())
    return send_file(path, as_attachment=True,
        download_name=f'AllWorks_Docs_{staff.username}.xlsx',
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


# ── Coordinator: JSON preview (shared by monthly + all-works) ─────────────────
@app.route('/admin/coordinator_reports/preview_data')
@login_required
@roles_required('admin','director','payments','office')
def coordinator_report_preview_data():
    coord_id   = request.args.get('coordinator_id', type=int)
    coord_name = request.args.get('coordinator_name', '').strip()
    month      = request.args.get('month', type=int)
    year       = request.args.get('year',  type=int)

    if not (coord_id or coord_name):
        return jsonify({'error': 'Missing coordinator'}), 400

    label, coordinator, all_projects = _resolve_coord_projects(coord_id, coord_name)

    if month and year:
        month_start = date(year, month, 1)
        month_end = date(year, month, calendar.monthrange(year, month)[1])
        projects  = [p for p in all_projects
                     if month_start <= p.created_at.date() <= month_end and p.status != 'Cancelled']
    else:
        projects = [p for p in all_projects if p.status != 'Cancelled']

    total_val = sum(float(p.total_amount    or 0) for p in projects)
    collected = sum(float(p.collected_amount or 0) for p in projects)
    active    = [p for p in projects if p.status in ('InProgress', 'Delayed', 'Lead', 'OnHold')]
    completed = [p for p in projects if p.status in ('Completed', 'Closed')]

    return jsonify({
        'kpis': {
            'Total':     len(projects),
            'Active':    len(active),
            'Completed': len(completed),
            'Value':     _inr_fmt(total_val),
            'Collected': _inr_fmt(collected),
            'Pending':   _inr_fmt(total_val - collected),
        },
        'projects': [_project_to_dict_coord(p)
                     for p in sorted(projects, key=lambda x: x.created_at, reverse=True)],
    })


# ── Doc staff: JSON preview (shared by monthly + all-works) ──────────────────
@app.route('/admin/docstaff_reports/preview_data')
@login_required
@roles_required('admin','director','office')
def docstaff_report_preview_data():
    staff_id = request.args.get('staff_id', type=int)
    month    = request.args.get('month', type=int)
    year     = request.args.get('year',  type=int)
    if not staff_id:
        return jsonify({'error': 'Missing staff_id'}), 400
 
    staff        = User.query.get_or_404(staff_id)
    all_projects = Project.query.filter_by(doc_staff_id=staff_id).all()
 
    if month and year:
        month_start = date(year, month, 1)
        month_end = date(year, month, calendar.monthrange(year, month)[1])
        projects  = [p for p in all_projects
                     if month_start <= p.created_at.date() <= month_end and p.status != 'Cancelled']
    else:
        projects = [p for p in all_projects if p.status != 'Cancelled']
    feas_done       = sum(1 for p in projects if _doc_done(p, 'Feasibility Receipt'))
    cd_done         = sum(1 for p in projects if _doc_done(p, 'CD Payment Receipt'))
    conn_done       = sum(1 for p in projects if _doc_done(p, 'KSEB Connection'))
    warr_done       = sum(1 for p in projects if _doc_done(p, 'Warranty Card'))
    completed       = [p for p in projects if p.status in ('Completed', 'Closed')]
    inprog          = [p for p in projects if p.status == 'InProgress']
    total_work_amt  = sum(float(p.total_amount or 0) for p in projects)
 
    return jsonify({
        'kpis': {
            'Total':       len(projects),
            'Completed':   len(completed),
            'In Progress': len(inprog),
            'Work Amount': _inr_fmt(total_work_amt),
            'Feasibility': feas_done,
            'CD Payment':  cd_done,
            'KSEB Conn.':  conn_done,
            'Warranty':    warr_done,
        },
        'projects': [_project_to_dict_docstaff(p)
                     for p in sorted(projects, key=lambda x: x.created_at, reverse=True)],
    })


# ── Coordinator: printable HTML ───────────────────────────────────────────────
@app.route('/admin/coordinator_reports/print')
@login_required
@roles_required('admin','director','payments','office')
def print_coordinator_report():
    coord_id   = request.args.get('coordinator_id', type=int)
    coord_name = request.args.get('coordinator_name', '').strip()
    month      = request.args.get('month', type=int)
    year       = request.args.get('year',  type=int)

    if not (coord_id or coord_name):
        flash('Missing coordinator.', 'danger')
        return redirect(url_for('coordinator_reports'))

    label, coordinator, all_projects = _resolve_coord_projects(coord_id, coord_name)

    if month and year:
        month_start = date(year, month, 1)
        month_end = date(year, month, calendar.monthrange(year, month)[1])
        projects  = [p for p in all_projects
                     if month_start <= p.created_at.date() <= month_end and p.status != 'Cancelled']
        period    = f'{calendar.month_name[month]} {year}'
    else:
        projects = [p for p in all_projects if p.status != 'Cancelled']
        period   = 'All Works'

    projects   = sorted(projects, key=lambda x: x.created_at, reverse=True)
    total_val  = sum(float(p.total_amount    or 0) for p in projects)
    total_coll = sum(float(p.collected_amount or 0) for p in projects)

    return render_template('print_coordinator_report.html',
        coordinator=coordinator,
        projects=projects,
        period=period,
        total_val=total_val,
        total_coll=total_coll,
        total_pend=total_val - total_coll,
        generated=date.today().strftime('%d %b %Y'),
    )


# ── Doc staff: printable HTML ─────────────────────────────────────────────────
@app.route('/admin/docstaff_reports/print')
@login_required
@roles_required('admin','director','office')
def print_docstaff_report():
    staff_id = request.args.get('staff_id', type=int)
    month    = request.args.get('month', type=int)
    year     = request.args.get('year',  type=int)
    if not staff_id:
        flash('Missing staff member.', 'danger')
        return redirect(url_for('docstaff_reports'))
 
    staff        = User.query.get_or_404(staff_id)
    all_projects = Project.query.filter_by(doc_staff_id=staff_id).all()
 
    if month and year:
        month_start = date(year, month, 1)
        month_end = date(year, month, calendar.monthrange(year, month)[1])
        projects  = [p for p in all_projects
                     if month_start <= p.created_at.date() <= month_end and p.status != 'Cancelled']
        period    = f'{calendar.month_name[month]} {year}'
    else:
        projects = [p for p in all_projects if p.status != 'Cancelled']
        period   = 'All Works'
 
    projects  = sorted(projects, key=lambda x: x.created_at, reverse=True)
    feas_done = sum(1 for p in projects if _doc_done(p, 'Feasibility Receipt'))
    cd_done   = sum(1 for p in projects if _doc_done(p, 'CD Payment Receipt'))
    conn_done = sum(1 for p in projects if _doc_done(p, 'KSEB Connection'))
    mnre_done = sum(1 for p in projects if _doc_done(p, 'MNRE'))
 
    return render_template('print_docstaff_report.html',
        staff=staff,
        projects=projects,
        period=period,
        feas_done=feas_done,
        cd_done=cd_done,
        conn_done=conn_done,
        mnre_done=mnre_done,
        doc_done=_doc_done,
        generated=date.today().strftime('%d %b %Y'),
    )
from flask import abort
from io import BytesIO
@app.route('/coordinator/my_report')
@login_required
def coordinator_my_report():
    if current_user.role not in ('coordinator','director'):
        abort(403)
    now = datetime.now()
    return render_template('coordinator_my_report.html',
        current_month=now.month,
        current_year=now.year)


@app.route('/coordinator/my_report/preview_data')
@login_required
def coordinator_my_report_preview():
    if current_user.role not in ('coordinator','director'):
        abort(403)
    month = request.args.get('month', type=int)
    year  = request.args.get('year',  type=int)

    all_projects = Project.query.filter_by(coordinator_id=current_user.id).all()
    if month and year:
        month_start = date(year, month, 1)
        month_end = date(year, month, calendar.monthrange(year, month)[1])
        projects  = [p for p in all_projects
                     if month_start <= p.created_at.date() <= month_end and p.status != 'Cancelled']
    else:
        projects = [p for p in all_projects if p.status != 'Cancelled']
    total_val = sum(float(p.total_amount    or 0) for p in projects)
    collected = sum(float(p.collected_amount or 0) for p in projects)
    active    = [p for p in projects if p.status in ('InProgress','Delayed','Lead','OnHold')]
    completed = [p for p in projects if p.status in ('Completed','Closed')]

    return jsonify({
        'kpis': {
            'Total':     len(projects),
            'Active':    len(active),
            'Completed': len(completed),
            'Value':     _inr_fmt(total_val),
            'Collected': _inr_fmt(collected),
            'Pending':   _inr_fmt(total_val - collected),
        },
        'projects': [_project_to_dict_coord(p)
                     for p in sorted(projects, key=lambda x: x.created_at, reverse=True)],
    })


@app.route('/coordinator/my_report/download')
@login_required
def coordinator_my_report_download():
    if current_user.role not in ('coordinator','director'):
        abort(403)
    coord = User.query.get_or_404(current_user.id)
    month = request.args.get('month', type=int)
    year  = request.args.get('year',  type=int)
    if not all([month, year]) or not (1 <= month <= 12):
        flash('Invalid parameters.', 'danger')
        return redirect(url_for('coordinator_my_report'))
    projects = Project.query.filter_by(coordinator_id=current_user.id).all()
    path = build_coordinator_monthly_report(coord, projects, year, month, tempfile.gettempdir())
    return send_file(path, as_attachment=True,
        download_name=f"report_{coord.full_name.replace(' ','_')}_{year}_{month:02d}.xlsx",
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


@app.route('/coordinator/my_report/download_all')
@login_required
def coordinator_my_report_download_all():
    if current_user.role not in ('coordinator','director'):
        abort(403)
    coord    = User.query.get_or_404(current_user.id)
    projects = Project.query.filter_by(coordinator_id=current_user.id).all()
    path     = build_allworks_coordinator_report(coord, projects, tempfile.gettempdir())
    return send_file(path, as_attachment=True,
        download_name=f"AllWorks_{coord.username}.xlsx",
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


@app.route('/coordinator/my_report/print')
@login_required
def coordinator_my_report_print():
    coord    = User.query.get_or_404(current_user.id)
    month    = request.args.get('month', type=int)
    year     = request.args.get('year',  type=int)
    projects = Project.query.filter_by(coordinator_id=current_user.id).all()

    if month and year:
        month_start = date(year, month, 1)
        month_end = date(year, month, calendar.monthrange(year, month)[1])
        projects  = [p for p in projects
                     if month_start <= p.created_at.date() <= month_end and p.status != 'Cancelled']
        period    = f'{calendar.month_name[month]} {year}'
    else:
        projects = [p for p in projects if p.status != 'Cancelled']
        period   = 'All Works'

    projects   = sorted(projects, key=lambda x: x.created_at, reverse=True)
    total_val  = sum(float(p.total_amount    or 0) for p in projects)
    total_coll = sum(float(p.collected_amount or 0) for p in projects)

    return render_template('print_coordinator_report.html',
        coordinator=coord,
        projects=projects,
        period=period,
        total_val=total_val,
        total_coll=total_coll,
        total_pend=total_val - total_coll,
        generated=datetime.now().strftime('%d %b %Y, %I:%M %p'))
from solar_app_software.job_card_excel import build_job_card

@app.route('/job_card')
@login_required
def job_card_page():
    query = _clean(request.args.get('q', ''), 100)
    proj  = None

    if query:
        from sqlalchemy.orm import joinedload
        opts = [
            joinedload(Project.connection_details),
            joinedload(Project.loan_detail),
            joinedload(Project.panel_details),
            joinedload(Project.subsidy),
            joinedload(Project.onsite_progress),
            joinedload(Project.documents),
            joinedload(Project.expenses),
            joinedload(Project.coordinator),
            joinedload(Project.doc_staff),
            joinedload(Project.kseb_task),
        ]
        base = Project.query.options(*opts)
        if current_user.role == 'documents_k':
            base = base.filter(Project.doc_staff_id == current_user.id)

        proj = base.filter_by(project_code=query).first()
        if not proj:
            proj = (base
                    .join(Customer)
                    .filter(Customer.name.ilike(f'%{query}%'))
                    .order_by(Project.updated_at.desc())
                    .first())

    now = datetime.utcnow().strftime('%d-%m-%Y')
    return render_template('job_card.html', proj=proj, query=query, now=now)
 
@app.route('/api/job_card_search')
@login_required
@limiter.limit('60 per minute')
def api_job_card_search():
    q = _clean(request.args.get('q', ''), 80)
    if len(q) < 2:
        return jsonify([])

    query = (Project.query
             .join(Customer)
             .filter(
                 Customer.name.ilike(f'%{q}%') |
                 Project.project_code.ilike(f'%{q}%')
             ))
    if current_user.role == 'documents_k':
        query = query.filter(Project.doc_staff_id == current_user.id)

    results = query.order_by(Project.updated_at.desc()).limit(10).all()

    return jsonify([{
        'code':  p.project_code,
        'name':  p.customer.name,
        'stage': p.stage,
        'status': p.status,
        'kw':    float(p.inverter_capacity_kw),
    } for p in results])

@app.route('/projects/<int:pid>/job_card/download')
@login_required
def download_job_card(pid):
    from solar_app_software.job_card_excel import build_job_card
    from sqlalchemy.orm import joinedload
    import tempfile

    proj = (Project.query
            .options(
                joinedload(Project.connection_details),
                joinedload(Project.loan_detail),
                joinedload(Project.panel_details),
                joinedload(Project.subsidy),
                joinedload(Project.onsite_progress),
                joinedload(Project.documents),
                joinedload(Project.expenses),
                joinedload(Project.coordinator),
                joinedload(Project.doc_staff),
                joinedload(Project.kseb_task),
                joinedload(Project.assignments).joinedload(WorkerAssignment.worker),
            )
            .filter_by(id=pid)
            .first_or_404())
    if current_user.role == 'documents_k' and proj.doc_staff_id != current_user.id:
        flash('This project is not assigned to you.', 'danger')
        return redirect(url_for('job_card_page'))
    path = build_job_card(
        project=proj,
        output_path=os.path.join(
            tempfile.gettempdir(),
            f'JobCard_{proj.project_code}.xlsx'
        )
    )
    return send_file(
        path,
        as_attachment=True,
        download_name=f'JobCard_{proj.project_code}.xlsx',
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
@app.route('/admin/service_records/clear_all', methods=['POST'])
@login_required
@roles_required('admin')
def clear_all_service_records():
    count = ServiceRecord.query.delete()
    db.session.commit()
    log_action(0, f'All service records cleared by {current_user.full_name}: {count} deleted') if count else None
    flash(f'{count} service record(s) deleted. Only payment-completed projects will get fresh schedules from now on.', 'warning')
    return redirect(url_for('service_management'))

# ─────────────────────────────────────────────────────────────────────────────
# DB INIT & SEED
# ─────────────────────────────────────────────────────────────────────────────

# def seed_db():
#     if User.query.count() == 0:
#         roles_data = [
#             ('admin',   'admin@poweronplus.in',  'Admin User',    'admin',       'Admin@12345'),
#             ('anita',   'anita@poweronplus.in',  'Anita Nair',    'coordinator', 'Coord@12345'),
#             ('vinod',   'vinod@poweronplus.in',  'Vinod Menon',   'coordinator', 'Coord@12345'),
#             ('sreeja',  'sreeja@poweronplus.in', 'Sreeja K',      'documents',   'Docs@12345'),
#             ('priya',   'priya@poweronplus.in',  'Priya Das',     'documents',   'Docs@12345'),
#             ('rajan',   'pay@poweronplus.in',    'Rajan P',       'payments',    'Pay@12345'),
#             ('suresh',  'onsite@poweronplus.in', 'Suresh K',      'onsite',      'Site@12345'),
#             ('appteam', 'app@poweronplus.in',    'App Team Lead', 'appinstall',  'App@12345'),
#         ]
#         for uname, email, fname, role, pwd in roles_data:
#             u = User(username=uname, email=email, full_name=fname, role=role)
#             u.set_password(pwd)   # hashed via scrypt
#             db.session.add(u)

#         workers_data = [
#             ('Arun K', '9845001111', 'Panel Installation', 1200),
#             ('Biju M', '9845002222', 'Electrical Work',    1200),
#             ('Cijo P', '9845003333', 'Structural Work',    1200),
#         ]
#         for name, phone, skill, rate in workers_data:
#             db.session.add(Worker(name=name, phone=phone, skill=skill, rate_per_day=rate))
#         db.session.commit()
#         print('✓ Database seeded with default users and workers.')

#     if DocumentStage.query.count() == 0:
#         seed_stages = [
#             ('Customer KYC',        'always',    'ID Proof,Pass Book,Electricity Bill',              0),
#             ('Bank / Loan file',    'loan_self', 'GEO Tag Photo,Bank Stamp Paper,Bank File',         1),
#             ('Feasibility',         'always',    'Feasibility Receipt',                              2),
#             ('KSEB filing',         'always',    'KSEB Stamp Paper,B-Class Licence,KSEB File',       3),
#             ('Inspection & conn.',  'always',    'Inspection,CD Payment Receipt,KSEB Connection',    4),
#             ('Subsidy',             'dcr',       'Subsidy Request,Subsidy Redeem',                   5),
#             ('Project closure',     'always',    'Payment Completion,Warranty Card,App Installation',6),
#         ]
#         for name, cond, docs, order in seed_stages:
#             db.session.add(DocumentStage(name=name, condition=cond, docs=docs, sort_order=order))
#         db.session.commit()
#         print('Document stages seeded.')


@app.cli.command('mark_delayed')
def mark_delayed():
    cutoff = datetime.utcnow() - timedelta(days=180)
    stale  = Project.query.filter(Project.status == 'InProgress', Project.created_at <= cutoff).all()
    for proj in stale:
        proj.status = 'Delayed'
    db.session.commit()
    print(f'✓ Marked {len(stale)} project(s) as Delayed.')


if __name__ == '__main__':
    with app.app_context():
        db.create_all()
        # seed_db()
    
    debug = os.environ.get('FLASK_ENV') != 'production'
    app.run(debug=debug, port=5000)



